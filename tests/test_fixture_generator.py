"""Тесты генератора фикстур с профилями нагрузки (этап 2, M2).

Проверяются реальным round-trip через openpyxl (генерация + чтение), а не
моками — генератор чисто вычислительный, живой Р7 не нужен. Файлы пишутся
во временную директорию pytest (tmp_path), не в репозиторий.
"""
import pytest
from openpyxl import load_workbook

import r7_Testovarka as r7mod


def test_unknown_profile_raises_value_error(bare_r7, tmp_path):
    with pytest.raises(ValueError, match="неизвестный профиль"):
        bare_r7._generate_fixture(tmp_path / "x.xlsx", rows=10, profile="nonsense")


def test_missing_openpyxl_raises_runtime_error(bare_r7, tmp_path, monkeypatch):
    monkeypatch.setattr(r7mod, "EXCEL_OK", False)
    with pytest.raises(RuntimeError, match="openpyxl"):
        bare_r7._generate_fixture(tmp_path / "x.xlsx", rows=10, profile="flat")


@pytest.mark.parametrize("profile", r7mod.R7Testovarka.FIXTURE_PROFILES)
def test_generates_readable_file_with_expected_row_count(bare_r7, tmp_path, profile):
    path = tmp_path / f"{profile}.xlsx"
    result = bare_r7._generate_fixture(path, rows=50, profile=profile, seed=1)

    assert result == path
    assert path.exists()
    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == 51  # 50 строк данных + 1 заголовок


@pytest.mark.parametrize("profile", r7mod.R7Testovarka.FIXTURE_PROFILES)
def test_same_seed_gives_identical_content(bare_r7, tmp_path, profile):
    """Обязательное условие для сравнения версий Р7: одна и та же нагрузка,
    а не случайно разные файлы одного размера."""
    p1, p2 = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    bare_r7._generate_fixture(p1, rows=30, profile=profile, seed=7)
    bare_r7._generate_fixture(p2, rows=30, profile=profile, seed=7)

    rows1 = [[c.value for c in row] for row in load_workbook(p1).active.iter_rows()]
    rows2 = [[c.value for c in row] for row in load_workbook(p2).active.iter_rows()]
    assert rows1 == rows2


@pytest.mark.parametrize("profile", r7mod.R7Testovarka.FIXTURE_PROFILES)
def test_different_seed_gives_different_content(bare_r7, tmp_path, profile):
    p1, p2 = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    bare_r7._generate_fixture(p1, rows=30, profile=profile, seed=1)
    bare_r7._generate_fixture(p2, rows=30, profile=profile, seed=2)

    rows1 = [[c.value for c in row] for row in load_workbook(p1).active.iter_rows()]
    rows2 = [[c.value for c in row] for row in load_workbook(p2).active.iter_rows()]
    assert rows1 != rows2


def test_flat_profile_respects_cols(bare_r7, tmp_path):
    path = tmp_path / "flat.xlsx"
    bare_r7._generate_fixture(path, rows=10, profile="flat", cols=8)
    ws = load_workbook(path).active
    assert ws.max_column == 8


def test_flat_profile_has_no_formulas(bare_r7, tmp_path):
    """flat специально нагружает парсер значений, не движок пересчёта —
    в нём не должно быть ни одной формулы."""
    path = tmp_path / "flat.xlsx"
    bare_r7._generate_fixture(path, rows=20, profile="flat")
    ws = load_workbook(path).active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith("="), \
                    f"неожиданная формула в flat-профиле: {cell.coordinate}={cell.value!r}"


def test_formula_profile_chains_to_previous_row(bare_r7, tmp_path):
    """Каждая формула должна ссылаться на C предыдущей строки данных — граф
    зависимостей, а не независимые формулы."""
    path = tmp_path / "formula.xlsx"
    bare_r7._generate_fixture(path, rows=15, profile="formula", seed=1)
    ws = load_workbook(path).active
    for r in range(3, 17):  # строки данных начинаются со 2-й (i=1)
        formula = ws[f"C{r}"].value
        assert isinstance(formula, str) and formula.startswith("=")
        assert f"C{r - 1}" in formula, f"строка {r}: {formula!r} не ссылается на C{r-1}"


def test_styled_profile_varies_font_and_fill(bare_r7, tmp_path):
    """Стили должны реально различаться между строками — иначе это
    единственный уникальный стиль на весь файл, а не то, что нагружает
    таблицу стилей."""
    path = tmp_path / "styled.xlsx"
    bare_r7._generate_fixture(path, rows=30, profile="styled", seed=1)
    ws = load_workbook(path).active
    fills = {ws[f"B{r}"].fill.fgColor.rgb for r in range(2, 32)}
    bolds = {ws[f"B{r}"].font.bold for r in range(2, 32)}
    assert len(fills) > 1
    assert bolds == {True, False}


def test_mixed_profile_has_formulas_and_some_styling(bare_r7, tmp_path):
    path = tmp_path / "mixed.xlsx"
    bare_r7._generate_fixture(path, rows=30, profile="mixed", seed=1)
    ws = load_workbook(path).active
    formula_count = sum(
        1 for r in range(3, 32)
        if isinstance(ws[f"D{r}"].value, str) and ws[f"D{r}"].value.startswith("=")
    )
    styled_count = sum(
        1 for r in range(2, 32)
        if ws[f"B{r}"].fill and ws[f"B{r}"].fill.fgColor.rgb != "00000000"
    )
    assert formula_count > 20   # почти все строки данных
    assert 0 < styled_count < 30  # часть, не все — каждая пятая по конструкции

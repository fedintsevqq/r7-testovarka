# -*- coding: utf-8 -*-
"""
R7-Testovarka Light – управление версиями + стресс-тест таблиц
"""

import os
import sys
import subprocess
import winreg
import time
import threading
import shutil
import re
import json
import hashlib
import csv
import ctypes
import platform
import html
import statistics
import random
import math
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import webbrowser

def get_base_dir() -> Path:
    """Возвращает базовую директорию приложения (портативный режим).

    В режиме PyInstaller (.exe, sys.frozen == True) — папка рядом с .exe.
    В режиме Python-скрипта — папка, содержащая .py файл.
    Использование sys.executable вместо sys.argv[0] надёжнее при запуске
    через ярлык или другой лаунчер.
    """
    if getattr(sys, 'frozen', False):
        # PyInstaller: sys.executable указывает на собранный .exe
        return Path(sys.executable).parent
    # Обычный запуск: берём папку скрипта, а не текущую рабочую директорию
    return Path(__file__).resolve().parent


BASE_DIR = get_base_dir()

# Консоль Windows по умолчанию — cp1251/cp866, а не UTF-8: print() с эмодзи
# (используются ниже в диагностике опциональных зависимостей) падает на такой
# консоли с UnicodeEncodeError и рушит запуск ещё до создания UI. sys.stdout
# бывает и None (pythonw.exe без консоли) — reconfigure на None кидает
# AttributeError, поэтому весь блок в try/except.
try:
    if sys.stdout is not None:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if sys.stderr is not None:
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# Библиотеки для автоматизации
try:
    import pyautogui
    PYAUTOGUI_OK = True
    # Аварийный выход: во время многоминутного автотеста клавиатура занята
    # программой, а мышь — нет. Инструмент нигде не двигает курсор сам
    # (только клики по текущей позиции — moveTo/dragTo в коде не используются),
    # поэтому включённый FAILSAFE ничего не ломает и не срабатывает случайно, а
    # даёт оператору физический способ прервать сценарий: увести мышь в угол
    # экрана поднимает pyautogui.FailSafeException.
    pyautogui.FAILSAFE = True
    # PAUSE по умолчанию 0.1 с и добавляется ПОСЛЕ КАЖДОГО вызова pyautogui.
    # Для замеров это чистый шум: хоткей из двух клавиш стоил 0.5 с сна
    # (4 события × interval + PAUSE) независимо от того, что делает Р7-Офис,
    # и именно эта константа, а не производительность Р7, определяла результат
    # 13 тестов. Всю действительно необходимую паузу задаём явно через
    # R7Testovarka._pace(), чтобы её можно было вычесть из замера.
    pyautogui.PAUSE = 0
except ImportError:
    PYAUTOGUI_OK = False
    print("⚠️ Установите pyautogui: pip install pyautogui")

try:
    import pyperclip
except ImportError:
    pyperclip = None
    print("⚠️ Установите pyperclip: pip install pyperclip")

try:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("⚠️ Установите openpyxl: pip install openpyxl")

try:
    import win32gui
    import win32con
    WIN32_OK = True
except ImportError:
    WIN32_OK = False
    print("⚠️ Установите pywin32: pip install pywin32")

try:
    import psutil
    PSUTIL_OK = True
except ImportError:
    PSUTIL_OK = False
    print("⚠️ Установите psutil: pip install psutil")

# Опциональный CDP-триггер готовности редактора (кнопка «Жирный» в DOM —
# см. r7_webdriver_connector.py и коммит 7978206). Полностью необязателен:
# без пакетов requests/websocket-client (или самого модуля) программа
# работает как раньше, на win32gui/CPU-логике из _wait_until_r7_ready.
try:
    from r7_webdriver_connector import (
        R7WebDriverConnector,
        r7_launch_debug_args,
        DEFAULT_CDP_PORT,
        WEBDRIVER_OK,
    )
except ImportError as e:
    print(f"⚠️ Ошибка импорта r7_webdriver_connector: {e}")
    WEBDRIVER_OK = False

print(f"🔍 WEBDRIVER_OK после импорта: {WEBDRIVER_OK} (файл: {__file__}, cwd: {os.getcwd()})")


COLORS = {
    "bg":            "#1E1E2E",  # основной фон
    "bg_card":       "#2A2A3E",  # фон карточек/фреймов
    "accent":        "#6C63FF",  # акцент: кнопки, активные элементы, заголовки
    "accent_hover":  "#5750D9",  # затемнение акцента при наведении
    "text":          "#E0E0E0",  # основной текст
    "text_secondary":"#A0A0B0",  # вторичный текст
    "border":        "#3A3A5A",  # границы/разделители, фон обычных кнопок
    "border_hover":  "#4A4A6A",  # фон кнопок при наведении
    "log_bg":        "#1A1A2E",  # фон лога
    "success":       "#4CAF50",  # INFO / ✅
    "warn":          "#FF9800",  # WARN / ⚠️
    "error":         "#F44336",  # ERROR / ❌
}
FONT_UI  = ("Segoe UI", 10)
FONT_LOG = ("Consolas", 9)
DEFAULT_TEST_RUNS = 7  # число прогонов по умолчанию для нового/несохранённого теста.
                       # Было 3 — мало для медианы/MAD на операциях короче
                       # разрешения детектора простоя, где значение бимодально
                       # (см. MIN_RUNS_FOR_STATS в run_test_with_runs и отчёт
                       # по нагрузочному тестированию, 25.08.2026). Диапазон
                       # UI (Spinbox from_=1, to=10) не менялся — 7 в него укладывается.

MEASURE_SCHEMA_VERSION = 2  # версия схемы JSON-результатов (performance_full_*.json).
                            # 1 (файлы до 25.08.2026, без этого поля): CPU-пороги
                            # детекторов простоя сравнивались с СЫРОЙ суммой
                            # cpu_percent() по процессам Р7, без деления на число
                            # ядер — момент «Р7 простаивает» зависел от железа
                            # стенда; итоговое время операции — среднее (avg) по
                            # прогонам, чувствительное к выбросам на бимодальных
                            # операциях. 2 (текущая): пороги нормированы на
                            # psutil.cpu_count() (см. R7Testovarka._cpu_count),
                            # итоговое время — медиана с MAD как мерой разброса
                            # (см. run_test_with_runs), первый прогон отбрасывается
                            # как прогрев. Старые файлы без этого поля читать как
                            # версию 1 — сравнивать 1 и 2 напрямую нельзя, разные
                            # величины.


def _col_letter(index):
    """Буквенное имя столбца по его номеру: 1 → A, 5 → E, 27 → AA.

    Нужно для CDP-пути тестов вставки: клавиатурная версия ходит по листу
    стрелками (Ctrl+Home, затем N раз «вправо»), а api адресует ячейки
    ссылками вида "A1:E1" — их и собирает эта функция.

    Args:
        index: Номер столбца, начиная с 1. Значения меньше 1 приводятся к 1.

    Returns:
        str: Буквенное имя столбца.
    """
    if index < 1:
        index = 1
    name = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        name = chr(ord("A") + rem) + name
    return name


def _linear_slope(points):
    """Наклон прямой методом наименьших квадратов по точкам (x, y).

    Не тянет numpy ради одной формулы — двухпроходная сумма по спискам
    длиной в десятки-сотни точек (частота семплера — раз в секунду, soak
    на несколько часов даёт тысячи, не миллионы) не оправдывает вес
    зависимости.

    Args:
        points: Непустая последовательность (x, y) — минимум 2 точки.

    Returns:
        float | None: Наклон (y на единицу x), либо None, если точек
        меньше 2 или все x совпадают (вертикальная выборка — наклон не
        определён, а не бесконечность).
    """
    pts = list(points)
    n = len(pts)
    if n < 2:
        return None
    mean_x = sum(x for x, _ in pts) / n
    mean_y = sum(y for _, y in pts) / n
    num = sum((x - mean_x) * (y - mean_y) for x, y in pts)
    den = sum((x - mean_x) ** 2 for x, _ in pts)
    if den == 0:
        return None
    return num / den


# Первая калибровка (не подтверждена многочасовым живым прогоном — см.
# отчёт по нагрузочному тестированию, раздел про soak-тесты, Stage 3).
# 5 МБ/час — заведомо выше шума однократных сборок мусора (JS-heap
# пилообразно колеблется на десятки МБ между циклами GC; линейная регрессия
# по достаточному числу точек эти колебания усредняет, но порог всё равно
# должен быть больше типичной амплитуды одного цикла, а не любого дрейфа).
LEAK_SLOPE_MB_PER_HOUR = 5.0
LEAK_MIN_SAMPLES = 30  # меньше — наклон недостоверен (шум одной точки GC)


# ПРОВЕРЕНО НА ЖИВОМ Р7 (25.08.2026): CDP-метрика "Documents" (доступна как
# doc_count в замерах ResourceSampler) — это ЧИСЛО ВНУТРЕННИХ DOM-ДОКУМЕНТОВ
# CEF-рендерера (фреймы, служебные контексты), а не «сколько файлов открыл
# пользователь». На живом прогоне с ОДНИМ открытым файлом и без единого
# действия пользователя она сама уехала 228 → 232 за 4 секунды простоя —
# то есть небольшой дрейф этой метрики нормален и не означает, что
# пользователь открыл второй документ. Поэтому «стабильность» ниже — это
# допуск на дрейф (доля от начального значения), а не точное равенство.
DOC_COUNT_STABLE_TOLERANCE_FRAC = 0.10  # первая калибровка по одному
                                        # короткому живому замеру (дрейф
                                        # 4/228 ≈ 1.8%), с запасом на более
                                        # длинный soak — требует уточнения
                                        # по итогам многочасового прогона
                                        # (Stage 3)


def detect_leak(samples, key="heap_mb",
                threshold_mb_per_hour=LEAK_SLOPE_MB_PER_HOUR,
                min_samples=LEAK_MIN_SAMPLES,
                doc_stable_tolerance_frac=DOC_COUNT_STABLE_TOLERANCE_FRAC):
    """Оценивает наличие утечки по ряду замеров ResourceSampler.

    Критерий: наклон линейной регрессии выше threshold_mb_per_hour ПРИ
    стабильном числе документов (doc_count не выходил за пределы допуска
    doc_stable_tolerance_frac от своего начального значения). Без второго
    условия рост heap из-за открытия новых документов читался бы как
    утечка — это не утечка, а ожидаемый рост данных. Допуск, а не точное
    равенство — см. DOC_COUNT_STABLE_TOLERANCE_FRAC.

    Args:
        samples: Список dict вида {"t": unix-время, key: значение в МБ,
            "doc_count": внутренний счётчик DOM-документов CEF | None}.
            Формат — тот же, что отдаёт ResourceSampler.snapshot().
        key: Какое поле замера анализировать — "heap_mb" (JS-куча
            рендерера, основной сигнал утечки внутри Р7) или "rss_mb"
            (RSS процесса — грубее, аллокатор маскирует освобождённую
            память, но не требует CDP).
        threshold_mb_per_hour: Порог наклона.
        min_samples: Минимум точек с непустым key, иначе наклон недостоверен.
        doc_stable_tolerance_frac: Допустимый дрейф doc_count относительно
            первого замера, доля (0.10 = ±10%).

    Returns:
        dict: {"leak": bool | None, "slope_mb_per_hour": float | None,
        "n_samples": int, "verdict": str}. leak=None означает «не удалось
        оценить» (мало точек или вырожденная выборка) — это НЕ «утечки нет»,
        вызывающий код должен различать эти два случая.
    """
    points = [(s["t"], s[key]) for s in samples if s.get(key) is not None]
    if len(points) < min_samples:
        return {"leak": None, "slope_mb_per_hour": None, "n_samples": len(points),
                "verdict": f"недостаточно замеров для оценки ({len(points)} < {min_samples})"}

    t0 = points[0][0]
    hours_series = [((t - t0) / 3600.0, v) for t, v in points]
    slope = _linear_slope(hours_series)
    if slope is None:
        return {"leak": None, "slope_mb_per_hour": None, "n_samples": len(points),
                "verdict": "наклон не определён (все замеры на одном времени)"}

    doc_counts = [s["doc_count"] for s in samples if s.get("doc_count") is not None]
    if doc_counts:
        baseline = doc_counts[0]
        tolerance = max(1.0, abs(baseline) * doc_stable_tolerance_frac)
        stable_docs = (max(doc_counts) - min(doc_counts)) <= tolerance
    else:
        stable_docs = True  # doc_count не собирался (нет CDP) — не блокируем вердикт
    over_threshold = slope > threshold_mb_per_hour
    slope_r = round(slope, 3)

    if over_threshold and stable_docs:
        verdict = (f"ЕСТЬ УТЕЧКА: {slope_r:.2f} МБ/час при стабильном "
                  f"числе документов ({len(points)} замеров)")
        leak = True
    elif over_threshold and not stable_docs:
        verdict = (f"наклон {slope_r:.2f} МБ/час выше порога, но число "
                  f"документов менялось за период выборки — не утечка, "
                  f"а рост данных")
        leak = False
    else:
        verdict = f"утечки не обнаружено (наклон {slope_r:.2f} МБ/час)"
        leak = False

    return {"leak": leak, "slope_mb_per_hour": slope_r, "n_samples": len(points),
            "verdict": verdict}


class ResourceSampler(threading.Thread):
    """Фоновый семплер RAM/JS-heap процессов Р7 — снимает точку раз в
    interval секунд, пока не остановлен, независимо от того, что в этот
    момент делает основной поток теста.

    ЗАЧЕМ ОТДЕЛЬНЫЙ ПОТОК, А НЕ ЗАМЕР В КОНЦЕ ТЕСТА: утечка — это НАКЛОН
    ряда во времени, а не одна точка. Существующие _sample_r7_resources
    снимают RAM один раз после каждой операции — этого достаточно для
    отчёта по одному прогону, но не для soak-теста в несколько часов,
    где интересен именно дрейф.

    ЗАЧЕМ JS-HEAP, А НЕ ТОЛЬКО RSS ПРОЦЕССА: утечку внутри самого редактора
    RSS процесса маскирует поведением аллокатора — память, освобождённая
    JS, не всегда сразу возвращается ОС (см. Performance.getMetrics в
    r7_webdriver_connector.py). RSS снимается всегда (psutil, не требует
    CDP); JS-heap — только если передан подключённый коннектор.

    Использование:
        sampler = ResourceSampler(get_procs=self._get_r7_processes,
                                   connector=self._webdriver_connector)
        sampler.start()
        ...
        sampler.stop()
        sampler.join(timeout=5)
        verdict = detect_leak(sampler.snapshot())
    """

    def __init__(self, get_procs, connector=None, interval=1.0, log_cb=None):
        """Args:
            get_procs: Callable() -> list[psutil.Process] — например,
                self._get_r7_processes. Вызывается заново на каждом
                замере (не список, зафиксированный один раз при
                создании) — процессы Р7 могут появляться/исчезать
                (x2t, новые окна) за время жизни семплера.
            connector: R7WebDriverConnector текущего запуска, либо None —
                тогда собирается только rss_mb, heap_mb/doc_count всегда
                None.
            interval: Пауза между замерами, сек.
            log_cb: Функция логирования; по умолчанию no-op (сэмплер живёт
                в отдельном потоке, лишний шум на каждую секунду soak-теста
                не нужен — ошибки одного замера не логируются, только
                накапливаются как None в ряду).
        """
        super().__init__(daemon=True)
        self._get_procs = get_procs
        self._connector = connector
        self._interval = interval
        self.log_cb = log_cb or (lambda msg: None)
        self._stop_event = threading.Event()
        self._lock = threading.Lock()
        self.samples = []

    def run(self):
        # Первый замер — сразу, не через interval секунд: иначе короткий
        # прогон (соак остановили раньше) рискует не набрать ни одной точки.
        self._sample_once()
        while not self._stop_event.wait(self._interval):
            self._sample_once()

    def _sample_once(self):
        """Снимает одну точку ряда. Ошибка одного источника (psutil упал,
        CDP не ответил) не должна ронять весь поток — соответствующее поле
        остаётся None, семплер продолжает работать дальше."""
        row = {"t": time.time(), "rss_mb": None, "heap_mb": None, "doc_count": None}
        try:
            procs = self._get_procs()
            if procs:
                row["rss_mb"] = sum(p.memory_info().rss for p in procs) / (1024 * 1024)
        except Exception as e:
            self.log_cb(f"⚠️ ResourceSampler: RSS не снят ({type(e).__name__}: {e})")

        if self._connector is not None and getattr(self._connector, "connected", False):
            try:
                metrics = self._connector.performance_metrics(timeout=2.0)
            except Exception as e:
                metrics = None
                self.log_cb(f"⚠️ ResourceSampler: JS-heap не снят ({type(e).__name__}: {e})")
            if metrics:
                if "JSHeapUsedSize" in metrics:
                    row["heap_mb"] = metrics["JSHeapUsedSize"] / (1024 * 1024)
                if "Documents" in metrics:
                    row["doc_count"] = metrics["Documents"]

        with self._lock:
            self.samples.append(row)

    def stop(self):
        """Останавливает цикл run(). Не блокирует — вызывающий код сам
        решает, ждать ли завершения потока (join)."""
        self._stop_event.set()

    def snapshot(self):
        """Копия накопленного ряда замеров — безопасно вызывать из другого
        потока, пока семплер продолжает работать (список копируется под
        тем же _lock, которым run() защищает append)."""
        with self._lock:
            return list(self.samples)


def _normal_cdf(z):
    """Функция распределения стандартного нормального закона Φ(z).

    math.erf — часть стандартной библиотеки с Python 3.2, точное (не
    приближённое) вычисление; отдельной зависимости не требует.
    """
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


def _mann_whitney_u(x, y):
    """Критерий Манна-Уитни (Уилкоксона для двух независимых выборок) —
    ручная реализация без scipy: нормальное приближение с поправкой на
    связи (ties) и непрерывность.

    ПОЧЕМУ БЕЗ SCIPY: scipy.stats.mannwhitneyu с версии 1.7 по умолчанию
    считает ТОЧНЫЙ p-value для маленьких выборок без связей (см.
    документацию scipy — проверено context7 при разработке этой функции).
    Здесь выборки как раз маленькие (5-10 прогонов на версию) — точный
    расчёт был бы честнее нормального приближения. Но добавлять scipy как
    зависимость означает тянуть за собой numpy — десятки МБ веса для
    десктопного инструмента, который сейчас распространяется как компактный
    portable-архив (см. requirements.txt — там только лёгкие пакеты). При
    N=5..10 и одном сравнении на пару версий разница между точным расчётом
    и нормальным приближением с поправкой на непрерывность для практических
    выводов ("регрессия"/"без изменений") пренебрежимо мала — реализация
    ниже проверена в тестах против брутфорс-перебора всех перестановок
    (единственного надёжного способа получить точный p-value без scipy) на
    малых N и даёт близкий результат.

    Args:
        x, y: Две независимые выборки чисел (settle_ms/api_ms одной
            операции на двух версиях Р7). Порядок важен для знака
            результата (используется compare_runs при интерпретации), но
            не для самого p-value — тест двусторонний.

    Returns:
        tuple[float, float]: (U1, p_two_sided). U1 — статистика для x
        относительно y. p_two_sided — 1.0 в вырожденном случае (обе выборки
        совпадают полностью, разброса нет — заведомо «нет оснований для
        вывода о различии»), не 0.0/NaN.
    """
    n1, n2 = len(x), len(y)
    n = n1 + n2
    combined = sorted([(v, 0) for v in x] + [(v, 1) for v in y], key=lambda p: p[0])

    # Ранги 1..n, усреднённые внутри групп связей (одинаковых значений).
    ranks = [0.0] * n
    tie_sizes = []
    i = 0
    while i < n:
        j = i
        while j < n and combined[j][0] == combined[i][0]:
            j += 1
        avg_rank = (i + 1 + j) / 2.0  # ранги i+1..j (1-based), их среднее
        for k in range(i, j):
            ranks[k] = avg_rank
        tie_sizes.append(j - i)
        i = j

    rank_sum_x = sum(r for r, (_, grp) in zip(ranks, combined) if grp == 0)
    u1 = rank_sum_x - n1 * (n1 + 1) / 2.0

    mean_u = n1 * n2 / 2.0
    tie_correction = sum(t ** 3 - t for t in tie_sizes)
    # n*(n-1) == 0 невозможно при n1,n2 >= 1 (n >= 2) — знаменатель безопасен.
    var_u = (n1 * n2 / 12.0) * ((n + 1) - tie_correction / (n * (n - 1)))
    if var_u <= 0:
        return u1, 1.0  # все значения совпадают — различать нечего

    sigma_u = math.sqrt(var_u)
    diff = u1 - mean_u
    continuity = 0.5 if diff > 0 else (-0.5 if diff < 0 else 0.0)
    z = (diff - continuity) / sigma_u
    p = 2.0 * (1.0 - _normal_cdf(abs(z)))
    return u1, min(1.0, max(0.0, p))


MIN_RUNS_FOR_COMPARISON = 5  # минимум прогонов на КАЖДУЮ версию — меньше
                             # критерий Манна-Уитни статистически ненадёжен
COMPARISON_MIN_EFFECT_PCT = 10.0  # практическая величина эффекта: значимый,
                                  # но <10% сдвиг — шум даже на починенной
                                  # метрике (см. отчёт по нагрузочному
                                  # тестированию про разброс api_ms)
COMPARISON_ALPHA = 0.05


def compare_runs(base_times, new_times,
                 min_effect_pct=COMPARISON_MIN_EFFECT_PCT,
                 alpha=COMPARISON_ALPHA,
                 min_runs=MIN_RUNS_FOR_COMPARISON):
    """Сравнивает длительности одной операции на двух версиях Р7 и выносит
    вердикт: регрессия, ускорение или без изменений.

    Регрессия/ускорение объявляются, только если ОБА условия выполнены:
      1. Статистическая значимость (критерий Манна-Уитни, p < alpha) —
         разница не объясняется случайным разбросом между прогонами.
      2. Практическая значимость (|относительная разница медиан| >
         min_effect_pct) — на достаточно большом N даже 2%-й сдвиг станет
         "статистически значимым", хотя для реального решения он не имеет
         веса (см. отчёт по нагрузочному тестированию, раздел про
         автодетект регрессий).

    Args:
        base_times, new_times: Списки длительностей одной операции — старая
            и новая версия Р7 соответственно. Порядок значим для знака
            effect_pct и для того, что считается "регрессией" (новая версия
            медленнее) против "ускорения" (новая версия быстрее).
        min_effect_pct: Порог практической значимости, %.
        alpha: Порог статистической значимости.
        min_runs: Минимум прогонов на каждую версию.

    Returns:
        dict: {"verdict": "РЕГРЕССИЯ" | "УСКОРЕНИЕ" | "без изменений" |
        "недостаточно прогонов", "median_base": float | None,
        "median_new": float | None, "effect_pct": float | None,
        "p_value": float | None, "n_base": int, "n_new": int}.
    """
    n_base, n_new = len(base_times), len(new_times)
    if n_base < min_runs or n_new < min_runs:
        return {"verdict": "недостаточно прогонов", "median_base": None,
                "median_new": None, "effect_pct": None, "p_value": None,
                "n_base": n_base, "n_new": n_new}

    median_base = statistics.median(base_times)
    median_new = statistics.median(new_times)
    effect_pct = (((median_new - median_base) / median_base) * 100.0
                 if median_base else 0.0)

    _, p_value = _mann_whitney_u(base_times, new_times)

    significant = p_value < alpha
    if significant and effect_pct > min_effect_pct:
        verdict = "РЕГРЕССИЯ"
    elif significant and effect_pct < -min_effect_pct:
        verdict = "УСКОРЕНИЕ"
    else:
        verdict = "без изменений"

    return {"verdict": verdict, "median_base": median_base, "median_new": median_new,
            "effect_pct": round(effect_pct, 1), "p_value": round(p_value, 4),
            "n_base": n_base, "n_new": n_new}


class R7Testovarka:
    TEST_DEFINITIONS = [
        "Выделение всех ячеек (Ctrl+A)",
        "Копирование всех ячеек (Ctrl+C)",
        "Вставка большого массива (Ctrl+V)",
        "Добавление нового листа",
        "Добавление столбца (горячие клавиши)",
        "Добавление столбца (меню Вставка)",
        "Вставка 1 ячейки (горячие клавиши)",
        "Вставка 5 ячеек (горячие клавиши)",
        "Вставка 1 ячейки (ПКМ)",
        "Вставка 5 ячеек (ПКМ)",
        "Функция ВПР (50K строк)",
        "Удаление столбца (Del)",
        "Сохранение в PDF (конвертация x2t)",
    ]

    # ── Пороги определения «документ открыт» ────────────────────────────────
    # Одни на все три режима (одиночный тест, тест своего файла, Batch), чтобы
    # они больше не разъезжались, как разъехались STABLE_SECS=10 и STABLE_SECS=8
    # у двух прежних копий ожидания загрузки.
    READY_POLL_SEC          = 0.15   # шаг опроса
    READY_RESPONSIVE_MS     = 300    # окно прокачало очередь быстрее — оно отзывчиво
    # ДО 25.08.2026 (measure_schema 1) здесь сравнивалась СЫРАЯ сумма
    # cpu_percent() по всем процессам Р7, без деления на число ядер — момент
    # «Р7 простаивает» зависел от количества ядер стенда: то же реальное
    # состояние (например, один поток пересчёта на 100%) на 4-ядерной машине
    # даёт сумму ~100, а на 16-ядерной — тоже ~100, но 100/16 ядер это совсем
    # другая доля мощности машины. Нормировка (см. _cpu_count()) переводит
    # порог в шкалу Task Manager (0–100% = вся машина): READY_IDLE_CPU_PCT=3.0
    # означает «меньше ~3% суммарной мощности», а не «меньше 8» в сырых
    # процентах одного ядра. См. measure_schema в JSON-результатах.
    READY_IDLE_CPU_PCT      = 3.0    # нормированный CPU процессов Р7 ниже — простой
    READY_IDLE_SAMPLES      = 20     # столько простоев подряд → документ открыт (≈3 с)
    READY_PROC_REFRESH_SEC  = 1.0    # как часто пересобирать список процессов (ловим x2t)
    READY_MIN_BUSY_SEC      = 0.5    # не выносить вердикт раньше — даём Р7 начать работу

    # Доп. триггер: кнопка «Жирный» на панели инструментов. Работает только
    # если она существует как отдельное нативное окно Win32 (класс "Button"
    # или "ToolbarButton") — см. предупреждение в docstring _wait_for_bold_button.
    # На практике это условие не выполняется ни для CEF-панели (текущая
    # сборка — HTML в одном render-окне), ни для классического Win32
    # ToolbarWindow32 (общий контрол сам рисует кнопки, у них тоже нет
    # отдельного HWND) — реалистичного билда, где сработает эта ветка, не
    # определено; см. docstring _is_bold_button_visible.
    # Однобуквенные метки ("b", "ж") сознательно из спецификации — риск: на
    # гипотетической сборке, где EnumChildWindows всё же находит что-то
    # подходящее по классу, ЛЮБАЯ кнопка с такой короткой подписью (не
    # обязательно именно "Жирный") будет принята без дополнительной проверки.
    BOLD_BUTTON_LABELS     = ("b", "ж", "жирный", "bold")  # регистронезависимо
    BOLD_BUTTON_CLASSES    = ("Button", "ToolbarButton")
    BOLD_BUTTON_POLL_SEC   = 0.1
    BOLD_BUTTON_TIMEOUT_SEC = 3.0
    # Верхняя граница именно на ПОДКЛЮЧЕНИЕ к CDP-порту (connector.connect()
    # внутри _wait_for_bold_button_cdp), не на весь бюджет BOLD_BUTTON_TIMEOUT_SEC.
    # _prepare_webdriver_launch создаёт коннектор всегда, когда порт 8080
    # свободен, — независимо от того, откроет ли его сама сборка Р7. Если
    # не откроет, connect() без этой границы опрашивал бы /json циклом до
    # 2 с впустую (порт закрыт => _pick_target() сразу None => sleep(poll_sec)
    # по кругу), а _wait_for_bold_button_cdp вызывается уже ПОСЛЕ того, как
    # остальные признаки готовности совпали — то есть это время инфлировало
    # бы прямо замер «Открытие файла» на каждой сборке без реального CDP.
    BOLD_BUTTON_CDP_CONNECT_TIMEOUT_SEC = 0.5

    # ── Замер отдельной операции ────────────────────────────────────────────
    # Операция считается завершённой, когда Р7 перестал быть занятым. Занятость
    # определяется по двум признакам сразу: окно не прокачивает очередь
    # сообщений ИЛИ процессы Р7 грузят CPU (плюс отдельно — жив ли конвертер x2t).
    OP_POLL_SEC         = 0.05   # шаг опроса состояния Р7
    OP_RESPONSIVE_MS    = 40     # окно не ответило за это — считаем занятым
    # Нормированная шкала — см. комментарий у READY_IDLE_CPU_PCT. Порог выше,
    # чем у READY_IDLE_CPU_PCT (3.0): окно опроса здесь вчетверо короче
    # (OP_CPU_WINDOW_SEC=0.20 против READY_PROC_REFRESH_SEC=1.0), а короткое
    # окно усреднения даёт больше дрожания на одном и том же реальном CPU.
    OP_IDLE_CPU_PCT     = 4.0    # нормированный CPU процессов Р7 ниже — не занято
    OP_CPU_WINDOW_SEC   = 0.20   # окно усреднения CPU: квант GetProcessTimes ≈15.6 мс,
                                 # на окне 50 мс это давало бы шум в десятки процентов
    OP_IDLE_SAMPLES     = 6      # подряд «не занято» → операция завершена (0.3 с)
    OP_START_GRACE_SEC  = 1.00   # ждём начала работы столько, прежде чем признать
                                 # операцию слишком быстрой для измерения.
                                 # В замер это ожидание НЕ попадает — стоит только
                                 # времени прогона, поэтому взято с запасом
    OP_PDF_GRACE_SEC    = 6.00   # для экспорта в PDF: x2t стартует не сразу после
                                 # Enter в диалоге «Сохранить как»
    OP_PROC_REFRESH_SEC = 0.50   # пересбор списка процессов (ловим x2t)
    OP_MAX_WAIT_SEC     = 180    # предохранитель на одну операцию
    OP_SELECT_ALL_MAX_SEC = 20   # отдельный, куда более короткий предохранитель
                                 # для Ctrl+A: выделив 25 млн ячеек, Р7 считает
                                 # по ним агрегаты в статусной строке и держит
                                 # CPU занятым десятками секунд. Общие 180 с
                                 # выглядели как зависание приложения; честнее
                                 # отметить операцию как timeout и идти дальше
    OP_KEY_PACE         = 0.08   # пауза после клавиш, меняющих состояние (буфер, лист)
    OP_MENU_PACE        = 0.12   # пауза на отрисовку меню или диалога
    OP_DIALOG_PACE      = 0.60   # отрисовка МОДАЛЬНОГО диалога («Вставить ячейки»).
                                 # OP_MENU_PACE=0.12 для него мало: модалка Р7 —
                                 # HTML внутри CEF, и на нагруженном документе она
                                 # не успевает появиться за 120 мс. Enter уходил в
                                 # сетку, а диалог оставался висеть (см. PR #4:
                                 # второй Enter добавили, но гонку не убрали)
    OP_DIALOG_ATTEMPTS  = 3      # столько раз подтверждаем модалку (см. _confirm_modal_enter)
    CLOSE_CDP_RETRY_SEC = 1.00   # как часто опрашивать CDP при закрытии Р7
                                 # (_close_r7_gracefully): реже шага цикла в 0.2 с,
                                 # чтобы не спамить websocket-запросами и логом

    # ── Статистика по прогонам одной операции (run_test_with_runs) ─────────
    # Среднее по 3 прогонам на бимодальной величине (см. измеренный разброс
    # api_ms/settle_ms) — способ увидеть регрессию там, где её нет: один
    # выброс сдвигает среднее непропорционально. Медиана устойчивее, MAD
    # (Median Absolute Deviation) — устойчивая мера разброса рядом с ней.
    MIN_RUNS_FOR_STATS = 2  # меньше — отбрасывать первый прогон (прогрев) уже
                            # нечем заменить: с runs=1 остался бы 0 прогонов

    def __init__(self, root):
        """Initializes the main application window and state.

        Args:
            root: The tkinter root window.
        """
        self.root = root
        self.root.title("R7-Testovarka Light")
        self.root.resizable(True, True)
        # Ниже сетка карточек на вкладке «Производительность» (Canvas шириной
        # 380px) и лог рядом с ней уже не помещаются вменяемо — без явного
        # предела окно можно было сжать до состояния, где всё наезжает друг
        # на друга.
        self.root.minsize(760, 560)
        self._apply_default_geometry()

        self.distributives_folder = BASE_DIR / "Distributives"
        self.distributives_folder.mkdir(exist_ok=True)

        self.test_files_folder = BASE_DIR / "TestFiles"
        self.test_files_folder.mkdir(exist_ok=True)

        self.reports_folder = BASE_DIR / "Reports"
        self.reports_folder.mkdir(exist_ok=True)

        self.current_version_info = None
        self.distributives = []
        self.selected_distributive = None
        self._cached_r7_path = None
        self._cached_cpu_count = None  # psutil.cpu_count(), см. _cpu_count()
        self._paced_total = 0.0    # сумма преднамеренных пауз внутри текущего замера
        self._pending_modal_confirm = False  # модалку «Вставить ячейки» надо
                                             # добить Enter'ами уже вне замера
                                             # (см. _flush_pending_modal_confirm)
        self._op_start_grace = None  # операция может попросить больше времени на старт
        self._op_max_wait = None     # ...и свой, более короткий, предохранитель
        self._webdriver_connector = None   # R7WebDriverConnector текущего запуска Р7, либо None
        self._current_webdriver_port = None  # CDP-порт текущего запуска, либо None
        self._op_via_cdp = False     # операция текущего замера ушла через api,
                                     # а не клавишами (влияет на трактовку below_floor)
        self._cdp_api_ms = 0.0       # сумма api_ms по всем шагам текущего замера —
                                     # синхронное время внутри рендерера, не зависящее
                                     # от опроса CPU детектором простоя (см. _cdp_sequence)
        self._pending_cdp_verify = None  # отложенная проверка CDP-операции
                                          # (см. _flush_pending_cdp_verify)
        self._cdp_ui_baseline = None  # DOM-снимок до первой операции — см. _capture_cdp_ui_baseline
        self.test_vars = {}   # populated by _build_perf_tab
        self.test_runs = {}   # populated by _build_perf_tab — IntVar per test, 1-10 runs
        self.perf_stop_event = threading.Event()
        self._perf_running = False   # защита от повторного запуска, пока прогон идёт
        self._batch_running = False  # тот же самый флаг для Batch-режима — оба
                                      # шлют клавиши в Р7-Офис и не должны идти одновременно

        self.setup_ui()
        self.refresh_distributives()
        self.detect_current_version()

    # ---------------------- UI ----------------------
    # Желаемый размер окна при старте. Числа не на глаз: собранному UI нужно
    # winfo_reqwidth x winfo_reqheight = 1149x669 (вкладка «Производительность»
    # одна требует 1125 по ширине — там сетка карточек Canvas 380px и лог
    # стоят рядом). Прежние 800x600 обрезали её на 325px по ширине — отсюда и
    # «интерфейс обрезан». Ниже — требуемое плюс запас на будущие виджеты.
    DEFAULT_WIN_W = 1220
    DEFAULT_WIN_H = 780

    def _apply_default_geometry(self):
        """Ставит стартовый размер окна, вписывая его в реальный экран.

        Жёсткое geometry("800x600") обрезало интерфейс, но и просто увеличить
        константу нельзя: на ноутбуке с 1366x768 окно 1180x780 не поместится
        и часть уедет за край. Поэтому желаемый размер ограничивается
        размером экрана минус поля под панель задач, а окно центрируется.
        Значения ниже minsize не опускаемся — тогда пусть лучше вылезет за
        край, чем виджеты наедут друг на друга.
        """
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
            # Поля: рамки окна по бокам и панель задач снизу.
            w = max(760, min(self.DEFAULT_WIN_W, screen_w - 80))
            h = max(560, min(self.DEFAULT_WIN_H, screen_h - 120))
            x = max(0, (screen_w - w) // 2)
            y = max(0, (screen_h - h) // 3)  # чуть выше центра — визуально ровнее
            self.root.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            # winfo_* теоретически может отказать до полной инициализации Tk —
            # окно без явной геометрии всё равно откроется, просто по умолчанию.
            self.root.geometry(f"{self.DEFAULT_WIN_W}x{self.DEFAULT_WIN_H}")

    def _apply_dark_theme(self):
        """Настраивает тёмную тему через ttk.Style.

        Тема 'clam' выбрана намеренно: нативные темы Windows ('vista'/
        'winnative') рисуют кнопки/вкладки/скроллбары средствами ОС и
        игнорируют цветовые переопределения ttk.Style для многих опций —
        подтверждено документацией Tk. 'clam' — собственный рендерер Tk,
        поддерживающий полную кастомизацию цвета для всех использованных
        здесь виджетов.
        """
        self.root.configure(bg=COLORS["bg"])
        style = ttk.Style(self.root)
        style.theme_use("clam")

        style.configure(".", background=COLORS["bg"], foreground=COLORS["text"],
                         font=FONT_UI)
        style.configure("TFrame", background=COLORS["bg"])
        style.configure("Card.TFrame", background=COLORS["bg_card"])
        style.configure("TLabel", background=COLORS["bg"], foreground=COLORS["text"])
        style.configure("Card.TLabel", background=COLORS["bg_card"], foreground=COLORS["text"])
        style.configure("Secondary.TLabel", background=COLORS["bg"],
                         foreground=COLORS["text_secondary"])
        style.configure("Header.TLabel", background=COLORS["bg"],
                         foreground=COLORS["accent"], font=("Segoe UI", 16, "bold"))
        style.configure("StatusOk.TLabel", background=COLORS["bg"], foreground=COLORS["success"])
        style.configure("StatusErr.TLabel", background=COLORS["bg"], foreground=COLORS["error"])

        style.configure("TLabelframe", background=COLORS["bg"], foreground=COLORS["text"],
                         bordercolor=COLORS["border"])
        style.configure("TLabelframe.Label", background=COLORS["bg"], foreground=COLORS["text_secondary"])

        style.configure("TButton", background=COLORS["border"], foreground=COLORS["text"],
                         bordercolor=COLORS["border"], focusthickness=0, padding=6)
        style.map("TButton",
                  background=[("active", COLORS["border_hover"]), ("pressed", COLORS["border_hover"])])
        style.configure("Accent.TButton", background=COLORS["accent"], foreground="#FFFFFF",
                         font=("Segoe UI", 12, "bold"), padding=10)
        style.map("Accent.TButton",
                  background=[("active", COLORS["accent_hover"]), ("pressed", COLORS["accent_hover"])])

        style.configure("TCheckbutton", background=COLORS["bg"], foreground=COLORS["text"])
        style.map("TCheckbutton", background=[("active", COLORS["bg"])])
        style.configure("Card.TCheckbutton", background=COLORS["bg_card"], foreground=COLORS["text"])
        style.map("Card.TCheckbutton", background=[("active", COLORS["bg_card"])])

        style.configure("TSpinbox", fieldbackground=COLORS["bg_card"], background=COLORS["bg_card"],
                         foreground=COLORS["text"], arrowcolor=COLORS["text"])

        style.configure("TNotebook", background=COLORS["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", background=COLORS["bg"], foreground=COLORS["text_secondary"],
                         padding=(14, 8), borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", COLORS["bg"])],
                  foreground=[("selected", COLORS["accent"])])

        style.configure("TScrollbar", background=COLORS["border"], troughcolor=COLORS["bg"],
                         bordercolor=COLORS["bg"], arrowcolor=COLORS["text_secondary"])
        style.map("TScrollbar", background=[("active", COLORS["border_hover"])])

        style.configure("Treeview", background=COLORS["bg_card"], fieldbackground=COLORS["bg_card"],
                         foreground=COLORS["text"], bordercolor=COLORS["border"], rowheight=26)
        style.configure("Treeview.Heading", background=COLORS["border"], foreground=COLORS["text"],
                         relief="flat")
        style.map("Treeview",
                  background=[("selected", COLORS["accent"])],
                  foreground=[("selected", "#FFFFFF")])

        style.configure("Horizontal.TProgressbar", background=COLORS["accent"],
                         troughcolor=COLORS["bg_card"], bordercolor=COLORS["bg"])

    def setup_ui(self):
        """Builds the main UI layout with notebook tabs and status bar."""
        self._apply_dark_theme()

        main = ttk.Frame(self.root, padding="10")
        main.pack(fill=tk.BOTH, expand=True)

        # ── Шапка: логотип + статус вместо стандартного заголовка окна ────────
        header = ttk.Frame(main)
        header.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(header, text="⚡ R7 Testovarka", style="Header.TLabel").pack(side=tk.LEFT)
        self.lbl_status_dot = ttk.Label(header, text="●  Готов", style="StatusOk.TLabel")
        self.lbl_status_dot.pack(side=tk.RIGHT)
        # «Тень» под шапкой: одна тёмная линия — ttk.Style не умеет рисовать
        # настоящую размытую тень, это ближайшее достижимое приближение.
        shadow = tk.Frame(main, height=2, bg=COLORS["border"])
        shadow.pack(fill=tk.X, pady=(0, 8))

        info = ttk.Frame(main, style="Card.TFrame", padding="10")
        info.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(info, text="Текущая версия", style="Secondary.TLabel").pack(anchor=tk.W)
        self.lbl_current = ttk.Label(info, text="Не определена", style="Card.TLabel",
                                     font=("Segoe UI", 16, "bold"))
        self.lbl_current.pack(anchor=tk.W, pady=(2, 0))

        self.notebook = ttk.Notebook(main)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        self.tab_versions = ttk.Frame(self.notebook)
        self.tab_perf = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_versions, text="📦 Версии")
        self.notebook.add(self.tab_perf, text="⚡ Производительность")

        self._build_versions_tab()
        self._build_perf_tab()

        self.status_var = tk.StringVar(value="Готов")
        status = ttk.Label(self.root, textvariable=self.status_var, anchor=tk.W, padding=(8, 4),
                           style="Secondary.TLabel")
        status.pack(side=tk.BOTTOM, fill=tk.X)

    def _build_versions_tab(self):
        """Builds the distributives table and install controls."""
        ttk.Label(self.tab_versions, text="Дистрибутивы", style="Secondary.TLabel").pack(
            anchor=tk.W, pady=(4, 6))
        frame = ttk.Frame(self.tab_versions, style="Card.TFrame")
        frame.pack(fill=tk.BOTH, expand=True)

        scroll = ttk.Scrollbar(frame)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree = ttk.Treeview(
            frame, columns=("name", "version", "size"), show="headings",
            selectmode="browse", yscrollcommand=scroll.set)
        self.tree.heading("name", text="Имя")
        self.tree.heading("version", text="Версия")
        self.tree.heading("size", text="Размер (МБ)")
        self.tree.column("name", width=320, anchor=tk.W)
        self.tree.column("version", width=110, anchor=tk.CENTER)
        self.tree.column("size", width=110, anchor=tk.CENTER)
        self.tree.pack(fill=tk.BOTH, expand=True)
        scroll.config(command=self.tree.yview)

        self.lbl_file_info = ttk.Label(self.tab_versions, text="", style="Secondary.TLabel")
        self.lbl_file_info.pack(anchor=tk.W, pady=5)

        self.quiet_install_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.tab_versions, text="Тихая установка",
                        variable=self.quiet_install_var).pack(anchor=tk.W, pady=(0, 5))

        btn_frame = ttk.Frame(self.tab_versions)
        btn_frame.pack(fill=tk.X, pady=10)
        self.btn_install = ttk.Button(btn_frame, text="📥 Установить", style="Accent.TButton",
                                      command=self.install_selected, state=tk.DISABLED)
        self.btn_install.pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🔄 Обновить", command=self.refresh_distributives).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📁 Добавить", command=self.add_distributive).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📂 Открыть папку", command=self.open_distributives_folder).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🔐 Проверить хеш-суммы", command=self.check_hashes).pack(side=tk.LEFT, padx=5)

        self.tree.bind('<<TreeviewSelect>>', self.on_select_distributive)

    def _build_perf_tab(self):
        """Builds the performance tab: dark log + card grid of tests + run controls."""
        top = ttk.Frame(self.tab_perf)
        top.pack(fill=tk.BOTH, expand=True, pady=(0, 4))

        # ── Log (dark, tagged by severity) ────────────────────────────────────
        log_frame = ttk.Frame(top)
        log_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.test_log = tk.Text(log_frame, font=FONT_LOG, bg=COLORS["log_bg"],
                                fg=COLORS["text"], insertbackground=COLORS["text"],
                                borderwidth=0, highlightthickness=0)
        self.test_log.tag_configure("INFO", foreground=COLORS["success"])
        self.test_log.tag_configure("WARN", foreground=COLORS["warn"])
        self.test_log.tag_configure("ERROR", foreground=COLORS["error"])
        scroll_log = ttk.Scrollbar(log_frame, command=self.test_log.yview)
        self.test_log.configure(yscrollcommand=scroll_log.set)
        scroll_log.pack(side=tk.RIGHT, fill=tk.Y)
        self.test_log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ── Test selection: scrollable grid of cards (2 columns) ──────────────
        sel_outer = ttk.LabelFrame(top, text="Выберите тесты", padding="4")
        sel_outer.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))

        cv = tk.Canvas(sel_outer, width=380, borderwidth=0, highlightthickness=0,
                       bg=COLORS["bg"])
        vsb = ttk.Scrollbar(sel_outer, orient="vertical", command=cv.yview)
        cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        cv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        inner = ttk.Frame(cv)
        cv_win = cv.create_window((0, 0), window=inner, anchor="nw")

        def _on_card_enter(event):
            """Highlights a test card (and its label) on mouse-over.

            Bound once and reused for every card — `event.widget` is the
            card frame itself, so no per-iteration closure is needed.
            """
            event.widget.configure(bg=COLORS["border_hover"])
            for w in event.widget.winfo_children():
                if isinstance(w, tk.Label):
                    w.configure(bg=COLORS["border_hover"])

        def _on_card_leave(event):
            """Restores a test card's normal background when the mouse leaves it."""
            event.widget.configure(bg=COLORS["bg_card"])
            for w in event.widget.winfo_children():
                if isinstance(w, tk.Label):
                    w.configure(bg=COLORS["bg_card"])

        saved = self._load_test_selection()
        self.test_vars = {}
        self.test_runs = {}
        CARD_COLS = 2
        for idx, name in enumerate(self.TEST_DEFINITIONS):
            entry = saved.get(name, {"enabled": True, "runs": DEFAULT_TEST_RUNS})
            var = tk.BooleanVar(value=entry.get("enabled", True))
            runs_var = tk.IntVar(value=entry.get("runs", DEFAULT_TEST_RUNS))

            card = tk.Frame(inner, bg=COLORS["bg_card"], bd=1, relief=tk.SOLID,
                            highlightbackground=COLORS["border"], highlightthickness=1)
            card.grid(row=idx // CARD_COLS, column=idx % CARD_COLS,
                      padx=4, pady=4, sticky="nsew")
            card.bind("<Enter>", _on_card_enter)
            card.bind("<Leave>", _on_card_leave)

            row1 = ttk.Frame(card, style="Card.TFrame")
            row1.pack(fill=tk.X, padx=6, pady=(6, 2))
            ttk.Checkbutton(row1, variable=var, style="Card.TCheckbutton").pack(side=tk.LEFT)
            ttk.Spinbox(row1, from_=1, to=10, increment=1, width=4,
                        textvariable=runs_var).pack(side=tk.LEFT, padx=(4, 0))
            lbl = tk.Label(card, text=name, bg=COLORS["bg_card"], fg=COLORS["text"],
                          wraplength=150, anchor=tk.W, justify=tk.LEFT,
                          font=FONT_UI)
            lbl.pack(fill=tk.X, padx=6, pady=(0, 6))

            self.test_vars[name] = var
            self.test_runs[name] = runs_var

        for c in range(CARD_COLS):
            inner.columnconfigure(c, weight=1)

        def _on_inner_configure(event):
            cv.configure(scrollregion=cv.bbox("all"))
        def _on_canvas_configure(event):
            cv.itemconfig(cv_win, width=event.width)

        inner.bind("<Configure>", _on_inner_configure)
        cv.bind("<Configure>", _on_canvas_configure)

        mini = ttk.Frame(sel_outer)
        mini.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(mini, text="☑ Все", width=7,
                   command=lambda: [v.set(True) for v in self.test_vars.values()]).pack(side=tk.LEFT)
        ttk.Button(mini, text="☐ Снять", width=7,
                   command=lambda: [v.set(False) for v in self.test_vars.values()]).pack(side=tk.LEFT, padx=3)

        # ── Progress bar ────────────────────────────────────────────────────
        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(self.tab_perf, variable=self.progress_var, maximum=100,
                        mode="determinate").pack(fill=tk.X, padx=2, pady=(4, 0))

        # ── Bottom action buttons ───────────────────────────────────────────
        btn_frame = ttk.Frame(self.tab_perf)
        btn_frame.pack(fill=tk.X, pady=4)
        self.btn_run_perf = ttk.Button(
            btn_frame, text="▶ Запустить выбранные тесты", style="Accent.TButton",
            command=self.run_spreadsheet_test)
        self.btn_run_perf.pack(side=tk.LEFT, padx=5)
        self.btn_stop_perf = ttk.Button(
            btn_frame, text="⏹ Остановить", command=self._request_stop_perf_test,
            state=tk.DISABLED)
        self.btn_stop_perf.pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📊 Сравнить размеры файлов",
                   command=self.compare_file_sizes).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📊 Сравнить версии",
                   command=self.compare_versions).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📈 Тренды",
                   command=self.show_trends).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🚀 Batch-режим (все версии)",
                   command=self.run_batch_mode).pack(side=tk.LEFT, padx=5)

    # ---------------------- Управление версиями ----------------------
    # Реестр читается и для HKLM (машинные установки, 64- и 32-битная ветка),
    # и для HKCU (установка "только для текущего пользователя") — раньше
    # HKCU не проверялся вовсе, и такие установки Р7-Офис не находились.
    _UNINSTALL_REGISTRY_ROOTS = (
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_CURRENT_USER,  r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
    )

    def _read_current_version_from_registry(self):
        """Чистое чтение реестра — без обращения к виджетам Tk.

        winreg — обычный Python-модуль, потокобезопасен как любой другой;
        в отличие от него, Tk-виджеты (self.lbl_current) не гарантированно
        безопасны для изменения не из главного потока. Разделение на «прочитать»
        (эта функция, любой поток) и «показать» (detect_current_version,
        только главный поток) нужно ровно поэтому — метод вызывается и из
        главного потока (при старте), и из фоновых (_batch_worker,
        install_selected.worker).

        Returns:
            dict | None: {"name", "version", "uninstall_string",
            "quiet_uninstall_string"} для первой найденной записи Р7-Офис,
            либо None, если ничего не найдено.
        """
        for root, reg_path in self._UNINSTALL_REGISTRY_ROOTS:
            try:
                key = winreg.OpenKey(root, reg_path, 0, winreg.KEY_READ)
            except OSError:
                continue
            try:
                for i in range(winreg.QueryInfoKey(key)[0]):
                    try:
                        sub = winreg.EnumKey(key, i)
                    except OSError:
                        continue
                    try:
                        subkey = winreg.OpenKey(key, sub)
                    except OSError:
                        continue
                    try:
                        name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                        if "Р7-Офис" in name or "R7-Office" in name:
                            ver = winreg.QueryValueEx(subkey, "DisplayVersion")[0]
                            info = {
                                "name": name,
                                "version": ver,
                                "uninstall_string": winreg.QueryValueEx(subkey, "UninstallString")[0],
                            }
                            try:
                                info["quiet_uninstall_string"] = \
                                    winreg.QueryValueEx(subkey, "QuietUninstallString")[0]
                            except OSError:
                                info["quiet_uninstall_string"] = None
                            return info
                    except OSError:
                        pass
                    finally:
                        winreg.CloseKey(subkey)
            finally:
                winreg.CloseKey(key)
        return None

    def detect_current_version(self):
        """Reads Windows registry and updates the "Текущая версия" label.

        self.current_version_info обновляется синхронно в вызывающем потоке —
        это обычное присваивание Python, оно безопасно из любого потока и
        нужно немедленно там, где detect_current_version вызывается из
        фонового потока и код сразу же читает результат (например,
        _batch_worker). Обновление самого виджета — единственная часть,
        которую нельзя делать не из главного потока, — маршалится туда через
        root.after(), если вызов пришёл не из главного потока.
        """
        info = self._read_current_version_from_registry()
        self.current_version_info = info

        def _update_label():
            if info:
                self.lbl_current.config(
                    text=f"{info['name']} ({info['version']})", foreground=COLORS["success"])
            else:
                self.lbl_current.config(text="Не установлена", foreground=COLORS["warn"])

        if threading.current_thread() is threading.main_thread():
            _update_label()
        else:
            self.root.after(0, _update_label)

    def refresh_distributives(self):
        """Rescans the Distributives folder and refreshes the table."""
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        self.distributives = []
        files = list(self.distributives_folder.glob("*.msi")) + list(self.distributives_folder.glob("*.exe"))
        if not files:
            self.btn_install.config(state=tk.DISABLED)
            self.status_var.set("Дистрибутивы не найдены")
            return
        files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        for f in files:
            ver = self._extract_version(f.stem) or "—"
            size_mb = round(f.stat().st_size / (1024 * 1024), 1)
            self.distributives.append({"path": f, "name": f.name})
            self.tree.insert("", tk.END, iid=str(len(self.distributives) - 1),
                              values=(f.name, ver, size_mb))
        self.status_var.set(f"Найдено: {len(files)}")

    def _extract_version(self, filename):
        """Extracts a version string like v2026.1.3 from a filename.

        Args:
            filename: The installer filename stem (without extension).

        Returns:
            str: Version string like 'v2026.1.3', or None if not found.
        """
        match = re.search(r'(\d+\.\d+(?:\.\d+)*)', filename)
        return f"v{match.group(1)}" if match else None

    def on_select_distributive(self, event):
        """Handles Treeview selection — enables Install button and shows file size."""
        sel = self.tree.selection()
        if sel and self.distributives:
            idx = int(sel[0])
            self.selected_distributive = self.distributives[idx]
            self.btn_install.config(state=tk.NORMAL)
            mb = self.selected_distributive["path"].stat().st_size / (1024 * 1024)
            self.lbl_file_info.config(text=f"{self.selected_distributive['name']} ({mb:.1f} МБ)")
        else:
            self.btn_install.config(state=tk.DISABLED)

    # msiexec.exe возвращает 3010 при успешном завершении, если требуется
    # перезагрузка — это тоже успех, а не ошибка.
    _MSIEXEC_SUCCESS_CODES = (0, 3010)

    def _build_uninstall_command(self, info):
        """Строит команду тихого удаления из данных реестра.

        UninstallString для MSI-пакетов обычно выглядит как
        "MsiExec.exe /I{GUID}" — это задокументированное поведение Windows
        Installer: /I означает «установить/переустановить», и с флагом
        /quiet, добавленным поверх, получается тихий РЕМОНТ установки, а не
        удаление. Настоящая тихая деинсталляция требует /X. Windows отдельно
        хранит QuietUninstallString с уже верным /X{GUID} — используем её,
        если она есть; иначе чиним /I на /X сами.

        Args:
            info: self.current_version_info.

        Returns:
            str: Готовая командная строка для subprocess.Popen(..., shell=False).
        """
        quiet_str = info.get("quiet_uninstall_string")
        if quiet_str:
            return quiet_str
        cmd = info["uninstall_string"]
        if "msiexec" in cmd.lower():
            fixed = re.sub(r'(?i)/I(\{[0-9A-Fa-f-]+\})', r'/X\1', cmd)
            cmd = fixed
        return cmd + " /quiet /norestart"

    def uninstall_current_version(self):
        """Silently uninstalls the currently detected R7-Office version.

        Returns:
            bool: True если удаление подтверждено (код возврата 0/3010, либо
            версия изначально не была установлена). False при таймауте или
            ненулевом коде возврата — в этом случае каталоги программы НЕ
            удаляются, чтобы не рассинхронизировать файлы с реестром.
        """
        if not self.current_version_info:
            return True
        self.status_var.set("Удаление...")
        cmd = self._build_uninstall_command(self.current_version_info)
        try:
            # shell=False: командная строка уже полностью собрана, а без
            # обёртки cmd.exe proc.kill() ниже завершает реальный процесс
            # деинсталлятора, а не промежуточный cmd.exe.
            proc = subprocess.Popen(cmd, shell=False)
        except OSError as e:
            self.status_var.set(f"⚠️ Не удалось запустить удаление: {e}")
            return False
        try:
            proc.wait(timeout=60)
        except subprocess.TimeoutExpired:
            proc.kill()
            self.status_var.set("⚠️ Удаление не завершилось за 60 сек, процесс завершён принудительно")
            return False

        if proc.returncode not in self._MSIEXEC_SUCCESS_CODES:
            self.status_var.set(f"⚠️ Удаление завершилось с кодом {proc.returncode}")
            return False

        time.sleep(3)
        for p in [r"C:\Program Files\R7-Office", r"C:\Program Files (x86)\R7-Office"]:
            if os.path.exists(p):
                shutil.rmtree(p, ignore_errors=True)
        return True

    def install_version(self, path, quiet=True):
        """Installs an R7-Office distributive.

        Args:
            path: Path object pointing to the .msi or .exe installer.
            quiet: If True (default), adds /quiet and installs silently.
                If False, the installer shows its normal UI.

        Returns:
            bool: True on success (return code 0 or 3010), False if the
            process timed out or exited with any other code.
        """
        self.status_var.set(f"Установка {path.name}...")
        if path.suffix == ".msi":
            cmd = ["msiexec", "/i", str(path), "/norestart"]
        else:
            cmd = [str(path)]
        if quiet:
            cmd.append("/quiet")
        # Тихая установка не требует участия пользователя — 5 минут с запасом.
        # Интерактивная показывает мастер установки, который пользователь
        # проходит вручную, поэтому таймаут увеличен, чтобы не убить процесс
        # посреди диалогов (EULA, выбор папки и т.д.).
        timeout_sec = 300 if quiet else 1800
        # shell=False: список аргументов не требует обёртки cmd.exe, и без неё
        # proc.kill() по таймауту завершает реальный установщик, а не cmd.exe.
        try:
            proc = subprocess.Popen(cmd, shell=False)
        except OSError as e:
            self.status_var.set(f"⚠️ Не удалось запустить установку: {e}")
            return False
        try:
            proc.wait(timeout=timeout_sec)
        except subprocess.TimeoutExpired:
            proc.kill()
            self.status_var.set(
                f"⚠️ Установка не завершилась за {timeout_sec // 60} мин, процесс завершён принудительно")
            return False
        if proc.returncode not in self._MSIEXEC_SUCCESS_CODES:
            self.status_var.set(f"⚠️ Установка завершилась с кодом {proc.returncode}")
            return False
        time.sleep(3)
        self.detect_current_version()
        return True

    def install_selected(self):
        """Confirms and launches uninstall + install in a background thread."""
        if not self.selected_distributive:
            return
        if self.current_version_info:
            if not messagebox.askyesno("Подтверждение",
                                       f"Удалить текущую и установить\n{self.selected_distributive['name']}?"):
                return
        self.btn_install.config(state=tk.DISABLED)
        quiet = self.quiet_install_var.get()

        def worker():
            uninstalled = self.uninstall_current_version()
            installed = False
            if uninstalled:
                installed = self.install_version(self.selected_distributive["path"], quiet=quiet)

            if installed:
                self.root.after(0, lambda: messagebox.showinfo("Готово", "Установка завершена"))
            elif not uninstalled:
                self.root.after(0, lambda: messagebox.showerror(
                    "Ошибка", "Не удалось удалить текущую версию — установка отменена.\n"
                             "Подробности в строке статуса."))
            else:
                self.root.after(0, lambda: messagebox.showerror(
                    "Ошибка", "Установка не завершилась успешно.\n"
                             "Подробности в строке статуса."))
            self.root.after(0, self.refresh_distributives)
            self.root.after(0, self.detect_current_version)
            self.root.after(0, lambda: self.btn_install.config(state=tk.NORMAL))
        threading.Thread(target=worker, daemon=True).start()

    def add_distributive(self):
        """Opens a file dialog to copy installers into the Distributives folder."""
        files = filedialog.askopenfilenames(filetypes=[("Installer", "*.msi *.exe")])
        for f in files:
            dst = self.distributives_folder / Path(f).name
            shutil.copy2(f, dst)
        self.refresh_distributives()

    def open_distributives_folder(self):
        """Opens the Distributives folder in Windows Explorer."""
        os.startfile(str(self.distributives_folder))

    # ---------------------- Настройки тестов ----------------------
    def _load_test_selection(self):
        """Loads saved test-selection state from selected_tests.json.

        Accepts both the old shape ({name: bool}) and the current one
        ({name: {"enabled": bool, "runs": int}}), upgrading the old one
        in memory so files saved by earlier versions of the app keep working.

        Returns:
            dict: Mapping test_name → {"enabled": bool, "runs": int}.
        """
        path = BASE_DIR / "selected_tests.json"
        if not path.exists():
            return {}
        try:
            with open(path, encoding="utf-8") as f:
                raw = json.load(f)
        except Exception:
            return {}
        upgraded = {}
        for name, value in raw.items():
            if isinstance(value, dict):
                upgraded[name] = {
                    "enabled": bool(value.get("enabled", True)),
                    "runs": int(value.get("runs", DEFAULT_TEST_RUNS)),
                }
            else:
                # Старый формат: значение — просто bool.
                upgraded[name] = {"enabled": bool(value), "runs": DEFAULT_TEST_RUNS}
        return upgraded

    def _save_test_selection(self):
        """Persists the current checkbox + run-count state to selected_tests.json."""
        path = BASE_DIR / "selected_tests.json"
        try:
            data = {
                name: {"enabled": var.get(), "runs": self.test_runs[name].get()}
                for name, var in self.test_vars.items()
            }
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.add_test_log(f"⚠️ Не удалось сохранить настройки тестов: {e}")

    # ---------------------- Лог ----------------------
    def add_test_log(self, msg):
        """Appends a timestamped, severity-colored message to the performance log.

        Severity is inferred from the leading emoji already used consistently
        throughout the codebase (❌/⚠️ for errors/warnings, everything else
        default) — no call site elsewhere in the file needs to change.

        Args:
            msg: The text to append.
        """
        try:
            if msg.startswith("❌"):
                tag = "ERROR"
            elif msg.startswith("⚠️"):
                tag = "WARN"
            else:
                tag = "INFO"
            line = f"[{datetime.now():%H:%M:%S}] {msg}\n"
            self.test_log.insert(tk.END, line, tag)
            self.test_log.see(tk.END)
            # update_idletasks (не update!): перерисовывает накопившиеся
            # изменения без обработки очереди событий Tk. add_test_log
            # вызывается сотнями раз за прогон из фоновых потоков — update()
            # заходил бы в главный цикл Tk и обрабатывал там события, включая
            # нажатия кнопок, реентерабельно посреди стека фонового потока.
            self.root.update_idletasks()
        except Exception:
            print(msg)

    def _set_perf_progress(self, done, total):
        """Updates the Performance tab's progress bar (0-100%). Safe to call
        even if the widget doesn't exist yet or the app is in another mode.
        Marshals the actual Tk update onto the main thread via root.after,
        since this is called from the worker thread during a test run."""
        try:
            pct = 100 * done / total if total else 0
            self.root.after(0, lambda: self.progress_var.set(pct))
        except Exception:
            pass

    # ---------------------- Стресс-тест таблиц ----------------------
    def _reset_perf_buttons(self):
        """Возвращает кнопки вкладки «Производительность» в состояние покоя.

        Вызывается из главного потока (через root.after) в finally-обёртке
        рабочего потока — при любом исходе: нормальном завершении,
        досрочной остановке или исключении.
        """
        self._perf_running = False
        try:
            self.btn_run_perf.config(state=tk.NORMAL)
            self.btn_stop_perf.config(state=tk.DISABLED)
        except Exception:
            pass

    def _request_stop_perf_test(self):
        """Обработчик кнопки «⏹ Остановить»: просит рабочий поток прерваться
        между операциями. Р7-Офис закрывается штатно, отчёт по уже
        выполненным операциям всё равно сохраняется."""
        self.perf_stop_event.set()
        self.btn_stop_perf.config(state=tk.DISABLED)
        self.add_test_log("⏹ Запрошена остановка теста...")

    def run_spreadsheet_test(self):
        """Entry point for the stress test — validates prerequisites then launches worker thread."""
        if self._perf_running:
            messagebox.showwarning("Тест уже выполняется",
                                   "Дождитесь завершения текущего прогона или нажмите «Остановить».")
            return
        if self._batch_running:
            messagebox.showwarning("Выполняется Batch-режим",
                                   "Оба режима управляют клавиатурой Р7-Офис и не могут "
                                   "работать одновременно. Дождитесь завершения Batch-режима.")
            return
        if not ctypes.windll.shell32.IsUserAnAdmin():
            messagebox.showerror(
                "Ошибка прав",
                "Стресс-тест требует запуска от имени администратора.\n"
                "Перезапустите программу с правами администратора."
            )
            return
        if not self.current_version_info:
            messagebox.showwarning("Нет версии", "Р7-Офис не установлен или не определён.")
            return
        if not PYAUTOGUI_OK or not pyperclip or not EXCEL_OK or not WIN32_OK:
            missing = []
            if not PYAUTOGUI_OK: missing.append("pyautogui")
            if not pyperclip: missing.append("pyperclip")
            if not EXCEL_OK: missing.append("openpyxl")
            if not WIN32_OK: missing.append("pywin32")
            messagebox.showerror("Ошибка",
                                 f"Отсутствуют библиотеки:\n{', '.join(missing)}\n"
                                 f"Установите: pip install " + " ".join(missing))
            return
        enabled = ({n for n, v in self.test_vars.items() if v.get()}
                   if self.test_vars else set(self.TEST_DEFINITIONS))
        if not enabled:
            messagebox.showwarning("Нет тестов", "Выберите хотя бы один тест для выполнения.")
            return
        # Снимок self.test_runs на главном потоке — как enabled_tests, чтобы
        # фоновый поток не трогал Tk-переменные напрямую.
        runs_snapshot = {n: v.get() for n, v in self.test_runs.items()}
        self._save_test_selection()

        self.perf_stop_event.clear()
        self._perf_running = True
        self.btn_run_perf.config(state=tk.DISABLED)
        self.btn_stop_perf.config(state=tk.NORMAL)

        def _worker():
            try:
                self._spreadsheet_worker(enabled, runs_snapshot, self.perf_stop_event)
            finally:
                # root.after — восстановление кнопок делает виджеты только
                # из главного потока. Покрывает любой исход: нормальное
                # завершение, досрочный return, необработанное исключение.
                self.root.after(0, self._reset_perf_buttons)

        threading.Thread(target=_worker, daemon=True).start()

    def _spreadsheet_worker(self, enabled_tests=None, test_runs=None, stop_event=None):
        """Runs selected spreadsheet performance tests sequentially and saves reports.

        Args:
            enabled_tests: Set of test-name strings to execute. None → all tests.
            test_runs: Dict of test-name → run count, snapshotted from
                self.test_runs on the main thread. None → empty dict.
            stop_event: threading.Event — установка прерывает прогон между
                операциями (и между повторами внутри одной операции). Р7-Офис
                при этом закрывается штатно, отчёт по уже выполненным
                операциям сохраняется. None → создаётся локально, никогда не
                устанавливается (для вызовов в обход UI).
        """
        if enabled_tests is None:
            enabled_tests = set(self.TEST_DEFINITIONS)
        if test_runs is None:
            test_runs = {}
        if stop_event is None:
            stop_event = threading.Event()
        self.add_test_log("\n🚀 ЗАПУСК СТРЕСС-ТЕСТА ТАБЛИЦ")

        # ----- 1. Поиск тестового файла -----
        def find_test_file():
            """Searches known directories for the 50K-row test spreadsheet.

            Returns:
                Path: Path to the found file, or None.
            """
            patterns = ["файл-для-теста-Р7-офис-50К*.xlsx", "файл-для-теста-Р7-офис-50К*.xls", "*50К*.xlsx"]
            search_dirs = [self.test_files_folder, BASE_DIR, Path.home() / "Downloads", Path.home() / "Загрузки", Path.cwd()]
            for sd in search_dirs:
                if not sd.exists():
                    continue
                for pat in patterns:
                    for f in sd.glob(pat):
                        return f
            return None

        test_file = find_test_file()
        if not test_file:
            self.add_test_log("❌ Тестовый файл не найден.")
            return
        self.add_test_log(f"✅ Найден файл: {test_file}")

        # ----- 2. Вспомогательные функции для окон -----
        def find_r7_window():
            """Returns the hwnd of the first visible R7-Office window, or None."""
            import win32gui
            wins = []
            def enum_cb(hwnd, _):
                if win32gui.IsWindowVisible(hwnd):
                    title = win32gui.GetWindowText(hwnd)
                    if "Р7-Офис" in title or test_file.stem in title:
                        wins.append(hwnd)
            win32gui.EnumWindows(enum_cb, wins)
            return wins[0] if wins else None

        def wait_for_window(title_part, timeout=60):
            """Polls for a visible window containing title_part, sets it foreground when found.

            Args:
                title_part: Substring to search for in window titles.
                timeout: Maximum seconds to wait.

            Returns:
                bool: True if window found, False on timeout.
            """
            import win32gui
            start = time.time()
            while time.time() - start < timeout:
                wins = []
                def enum_cb(hwnd, _):
                    if win32gui.IsWindowVisible(hwnd):
                        title = win32gui.GetWindowText(hwnd)
                        if title_part.lower() in title.lower():
                            wins.append(hwnd)
                win32gui.EnumWindows(enum_cb, wins)
                if wins:
                    win32gui.SetForegroundWindow(wins[0])
                    return True
                time.sleep(0.5)
            return False

        def maximize_window():
            """Maximizes the R7-Office window if found.

            Returns:
                bool: True if window was found and maximized.
            """
            import win32gui, win32con
            hwnd = find_r7_window()
            if hwnd:
                win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
                time.sleep(0.5)
                return True
            return False

        def focus_window():
            """Brings the R7-Office window to the foreground.

            Returns:
                bool: True if window was found and focused.
            """
            import win32gui
            hwnd = find_r7_window()
            if hwnd:
                try:
                    win32gui.SetForegroundWindow(hwnd)
                except Exception:
                    pass
                time.sleep(0.3)
                return True
            return False

        def close_update_dialog(search_timeout=0):
            return self._close_update_dialog_if_exists(search_timeout=search_timeout)

        def safe_hotkey(*keys):
            """Отправляет сочетание клавиш без задержек.

            Прежний interval=0.1 добавлял 0.1 сек между каждым нажатием и
            отпусканием: хоткей из двух клавиш стоил 0.4 сек, из трёх — 0.6 сек,
            и это время целиком попадало в замер. Пауза, где она действительно
            нужна, задаётся явно через self._pace().
            """
            pyautogui.hotkey(*keys)

        def safe_press(key, presses=1, pace=0.0):
            """Нажимает клавишу одна или несколько раз.

            Args:
                key: Имя клавиши в терминах pyautogui.
                presses: Сколько раз нажать.
                pace: Пауза между нажатиями (через _pace, вычитается из замера).
                    Нужна при навигации по меню, где Р7 не успевает отрисовать
                    следующий пункт.
            """
            for _ in range(presses):
                pyautogui.press(key)
                if pace:
                    self._pace(pace)

        def post_action_delay(seconds=0.5):
            """Waits after an operation completes — called outside measure() timing window."""
            time.sleep(seconds)

        # ----- 3. Запуск Р7 и замер времени открытия -----
        r7_path = self._find_r7_path()
        if not r7_path:
            self.add_test_log("❌ Р7-Офис не найден.")
            return

        self.add_test_log(f"🔄 Запуск Р7-Офис с файлом: {test_file.name}")
        # Порт проверяется ДО старта секундомера — иначе TCP-connect_ex
        # внутри _prepare_webdriver_launch попадает в open_elapsed, хоть и
        # не относится к скорости открытия файла.
        debug_args = self._prepare_webdriver_launch(filename_hint=test_file.name)
        open_start = time.time()
        subprocess.Popen([r7_path, str(test_file), *debug_args], shell=True)

        if not wait_for_window(test_file.stem, timeout=60) and not wait_for_window("Р7-Офис", timeout=10):
            self.add_test_log("❌ Окно Р7 не появилось.")
            return

        # Подготовка окна к тесту (разворот, фокус, снятие диалога обновления)
        # к скорости открытия файла отношения не имеет — засекаем её отдельно
        # и вычитаем из open_elapsed, иначе она уезжает прямо в результат.
        _setup_start = time.time()
        maximize_window()
        focus_window()
        # Один проход без опроса: дальше диалог обновления ловит фоновый монитор.
        close_update_dialog(search_timeout=0)
        _setup_elapsed = time.time() - _setup_start

        # Фоновый мониторинг окна обновления на весь период теста
        _upd_stop = threading.Event()
        threading.Thread(
            target=self._monitor_update_dialog,
            args=(_upd_stop,),
            daemon=True,
        ).start()
        self.add_test_log("🔍 Запущен мониторинг окна обновления (проверка каждые 2 сек)")

        # Фоновый семплер ресурсов (этап 2, H3) — создаётся здесь (не внутри
        # try ниже), тем же паттерном, что и _upd_stop чуть выше: чтобы имя
        # было гарантированно определено к моменту finally, даже если try
        # упадёт на первой же строке. .start() — позже, у начала прогона
        # операций (см. там же), здесь ещё рано: RAM/CPU только формируются
        # открытием файла, замерять эту фазу как часть теста не нужно.
        _resource_sampler = ResourceSampler(
            get_procs=self._get_r7_processes,
            connector=self._webdriver_connector,
            interval=1.0,
            log_cb=self.add_test_log,
        )

        try:
            data_ready = self._wait_until_r7_ready(find_r7_window, timeout=120)
            open_elapsed = time.time() - open_start - _setup_elapsed
            self.add_test_log(
                f"✅ Файл открыт за {open_elapsed:.2f} сек "
                f"({'данные загружены' if data_ready else 'таймаут — возможна частичная загрузка'};"
                f" подготовка окна {_setup_elapsed:.2f} сек не учтена)")

            if not focus_window():
                _upd_stop.set()
                self.add_test_log("❌ Окно Р7-Офис недоступно после открытия файла — тест прерван")
                return

            # Подключаемся к CDP до снятия базового снимка: без соединения
            # снимок был бы пустым, и вычитать из дампов меню стало бы нечего.
            self._cdp_ensure_connected()
            # Базовый DOM-снимок ДО первой операции — см. _cdp_dump_ui и
            # _capture_cdp_ui_baseline (issue #9).
            self._capture_cdp_ui_baseline()
            # Один раз за запуск: найден ли внутренний api редактора. От этого
            # зависит, пойдут тесты через CDP или клавишами.
            self._cdp_log_api_info()

            # ----- 3.5 Мониторинг ресурсов ------------------------------------------------
            self._r7_pids = None  # сбросить кэш перед новым поиском
            self._x2t_logged_pids = set()  # сбросить дедуп x2t перед новым тестом
            r7_procs = self._get_r7_processes()
            if PSUTIL_OK and r7_procs:
                try:
                    _init_ram = round(
                        sum(p.memory_info().rss for p in r7_procs) / (1024 * 1024), 1
                    )
                    pids_str = ", ".join(str(p.pid) for p in r7_procs)
                    self.add_test_log(
                        f"🔍 Поиск процесса Р7: найдено {len(r7_procs)} процессов "
                        f"(PID: {pids_str}), суммарная RAM = {_init_ram:.1f} МБ"
                    )
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    self.add_test_log(f"🔍 Найдено {len(r7_procs)} процессов Р7, RAM недоступна")
            else:
                self.add_test_log(
                    "⚠️ Процесс Р7 не найден — замеры RAM/CPU будут недоступны"
                    if PSUTIL_OK else
                    "⚠️ psutil не установлен — замеры RAM/CPU недоступны"
                )

            # ----- 4. Тесты ----------------------------------------------------------------
            sample0 = self._sample_r7_resources(r7_procs)
            results = [{
                "name": "Открытие файла", "time": open_elapsed, "error": None,
                "ram":            sample0["ram_mb"]       if sample0 else None,
                "cpu":            sample0["cpu_raw_pct"]   if sample0 else None,
                "cpu_normalized": sample0["cpu_norm_pct"]  if sample0 else None,
                "threads":        sample0["threads"]       if sample0 else None,
                "uptime_sec":     sample0["uptime_sec"]    if sample0 else None,
            }]

            def run_test_with_runs(name, func, runs):
                """Runs `func` `runs` times, logging each pass, and appends one
                averaged result to `results` (or nothing if the test is disabled).

                The resource sample (RAM/CPU/threads/uptime) is taken once, after
                the LAST pass — sampling on every pass would multiply the
                0.1s-per-process cpu_percent() blocking cost by `runs` for no
                benefit, since RAM/CPU during repeated identical operations don't
                need a separate reading per pass the way timing does. The process
                list IS still refreshed right before that single sample, exactly
                like the sibling batch worker's measure() does — otherwise
                short-lived processes such as x2t, spawned during one of the
                `runs` passes, would be missed by a stale r7_procs snapshot (the
                same bug the shipped fix in that measure() exists to prevent).
                """
                nonlocal r7_procs
                if name not in enabled_tests:
                    return
                runs = max(1, int(runs))
                self.add_test_log(f"⏳ Тест: {name} (прогон 1/{runs})...")
                try:
                    focus_window()
                except Exception as e:
                    self.add_test_log(f"   ⚠️ Не удалось установить фокус: {e}")

                pass_times = []
                api_ms_values = []    # синхронное время api по прогонам, ушедшим через CDP
                error = None
                below_floor = False   # хоть один прогон оказался ниже порога измерения
                for i in range(runs):
                    if stop_event.is_set():
                        self.add_test_log(f"⏹ {name}: остановлено пользователем "
                                          f"(выполнено прогонов: {i}/{runs})")
                        break
                    if i > 0:
                        self.add_test_log(f"⏳ Тест: {name} (прогон {i + 1}/{runs})...")
                    # Секундомер запускается перед отправкой клавиш и
                    # останавливается, когда Р7 перестал быть занятым, — за вычетом
                    # собственных пауз, накопленных в _paced_total.
                    self._paced_total = 0.0
                    self._op_start_grace = None
                    self._op_max_wait = None
                    self._op_via_cdp = False
                    self._cdp_api_ms = 0.0
                    start = time.time()
                    try:
                        func()
                    except Exception as e:
                        error = str(e)
                        self.add_test_log(f"   ❌ прогон {i + 1}: ошибка — {e}")
                        break
                    done_ts, status = self._wait_operation_done(find_r7_window)
                    if status == "timeout":
                        elapsed = time.time() - start - self._paced_total
                    else:
                        elapsed = max(0.0, done_ts - start - self._paced_total)
                    pass_times.append(elapsed)
                    if self._op_via_cdp:
                        api_ms_values.append(self._cdp_api_ms)
                    # Замер закрыт — только теперь добиваем модалку «Вставить
                    # ячейки», если она не успела появиться внутри операции,
                    # и доводим отложенную проверку CDP-операции (её round-trip
                    # не должен попадать в цифру). Зеркалится в measure()
                    # Batch-режима.
                    self._flush_pending_modal_confirm()
                    self._flush_pending_cdp_verify()
                    post_action_delay()
                    # api_ms — субмиллисекундное разрешение (см. _cdp_sequence),
                    # печатается рядом с elapsed, а не вместо него: elapsed
                    # по-прежнему то, что видит пользователь (settle_ms).
                    _api_note = (f" [api: {self._cdp_api_ms:.2f} мс]"
                                if self._op_via_cdp else "")
                    if status == "below_floor":
                        below_floor = True
                        _grace = self._op_start_grace or self.OP_START_GRACE_SEC
                        if self._op_via_cdp:
                            # На CDP-пути это не «ноль, который мы не умеем
                            # измерить»: Runtime.evaluate возвращается только
                            # когда api отработал, поэтому цифра — реальная
                            # длительность вызова. Просто Р7 после него не
                            # успел стать занятым.
                            self.add_test_log(
                                f"   ⏱ прогон {i + 1}: {elapsed:.3f} сек{_api_note} — "
                                f"вызов api отработал синхронно, Р7 не стал занятым")
                        else:
                            self.add_test_log(
                                f"   ⏱ прогон {i + 1}: {elapsed:.3f} сек — Р7 не был занят "
                                f"дольше {_grace:.1f} сек, операция ниже порога измерения")
                    elif status == "timeout":
                        self.add_test_log(
                            f"   ⚠️ прогон {i + 1}: {elapsed:.3f} сек{_api_note} — "
                            f"Р7 так и не освободился")
                    else:
                        self.add_test_log(f"   ✅ прогон {i + 1}: {elapsed:.3f} сек{_api_note}")

                if not pass_times:
                    results.append({"name": name, "time": 0.0, "error": error,
                                     "ram": None, "cpu": None, "cpu_normalized": None,
                                     "threads": None, "uptime_sec": None,
                                     "runs": [], "avg": 0.0, "min": 0.0, "max": 0.0,
                                     "median": 0.0, "mad": 0.0, "n_runs": 0,
                                     "first_run_discarded": False,
                                     "below_floor": False, "api_ms": None})
                    return

                # avg/min/max — старые ключи, без изменений (обратная
                # совместимость с уже сохранёнными performance_full_*.json и
                # их читателями). Среднее по 3 прогонам на бимодальной
                # величине — способ увидеть регрессию там, где её нет: один
                # выброс сдвигает его непропорционально (см. отчёт по
                # нагрузочному тестированию, 25.08.2026). Поэтому headline-
                # значение ("time" ниже) — медиана, а не avg_t.
                avg_t = sum(pass_times) / len(pass_times)
                min_t = min(pass_times)
                max_t = max(pass_times)

                # Первый прогон отбрасывается как прогрев (JIT/кэши файловой
                # системы/CDP-соединения) — но только если после этого
                # останется что усреднять (см. MIN_RUNS_FOR_STATS).
                first_run_discarded = len(pass_times) >= self.MIN_RUNS_FOR_STATS
                stats_times = pass_times[1:] if first_run_discarded else pass_times
                median_t = statistics.median(stats_times)
                mad_t = self._mad(stats_times)

                # Среднее api_ms только по прогонам, ушедшим через CDP — на
                # клавиатурном пути api_ms не существует, и подмешивать сюда
                # его отсутствие как ноль исказило бы среднее вниз.
                avg_api_ms = (round(sum(api_ms_values) / len(api_ms_values), 3)
                              if api_ms_values else None)
                _avg_api_note = (f", api {avg_api_ms:.2f} мс" if avg_api_ms is not None else "")
                _discard_note = " (1-й отброшен)" if first_run_discarded else ""
                self.add_test_log(
                    f"   📊 медиана {median_t:.3f} сек (MAD {mad_t:.3f}) — "
                    f"{len(stats_times)}/{len(pass_times)} прогонов{_discard_note}, "
                    f"среднее {avg_t:.3f} сек (мин {min_t:.3f}, макс "
                    f"{max_t:.3f}{_avg_api_note})")

                # Обновляем список процессов перед замером — как в measure()
                # соседнего batch-воркера, иначе короткоживущий x2t может быть
                # пропущен устаревшим снимком.
                self._r7_pids = None
                r7_procs = self._get_r7_processes()
                sample = self._sample_r7_resources(r7_procs)
                self._log_resources(sample)

                results.append({
                    "name": name, "time": median_t, "error": error,
                    "ram":            sample["ram_mb"]      if sample else None,
                    "cpu":            sample["cpu_raw_pct"]  if sample else None,
                    "cpu_normalized": sample["cpu_norm_pct"] if sample else None,
                    "threads":        sample["threads"]      if sample else None,
                    "uptime_sec":     sample["uptime_sec"]    if sample else None,
                    "runs": pass_times, "avg": avg_t, "min": min_t, "max": max_t,
                    "median": median_t, "mad": mad_t, "n_runs": len(stats_times),
                    "first_run_discarded": first_run_discarded,
                    "below_floor": below_floor, "api_ms": avg_api_ms,
                })

            # Все паузы ниже идут через self._pace() — они нужны для надёжности
            # автоматизации, но вычитаются из замера. Прежние time.sleep() внутри
            # этих функций попадали в результат напрямую.
            KEY_PACE  = self.OP_KEY_PACE
            MENU_PACE = self.OP_MENU_PACE

            def copy_paste_hotkey(cell_count, paste_offset):
                # CDP: выделить A1:<N>1, скопировать, уйти вправо, вставить —
                # мимо фокуса и клавиатуры (см. _cdp_copy_paste).
                if self._cdp_copy_paste(cell_count, paste_offset, key_pace=KEY_PACE):
                    return
                safe_hotkey('ctrl', 'home')
                for _ in range(cell_count - 1):
                    pyautogui.hotkey('shift', 'right')
                safe_hotkey('ctrl', 'c')
                self._pace(KEY_PACE)          # даём буферу обмена наполниться
                pyautogui.press('right', presses=paste_offset)
                safe_hotkey('ctrl', 'v')

            def copy_paste_context(cell_count, paste_offset):
                # CDP: то же выделение и копирование, но вставка — со сдвигом
                # ячеек вниз, то есть ровно то, что делает пункт контекстного
                # меню «Вставить ячейки» и следующая за ним модалка выбора
                # сдвига. Модалки на этом пути не возникает.
                if self._cdp_copy_paste(cell_count, paste_offset, shift="down",
                                        key_pace=KEY_PACE):
                    return
                safe_hotkey('ctrl', 'home')
                for _ in range(cell_count - 1):
                    pyautogui.hotkey('shift', 'right')
                pyautogui.click(button='right')
                self._pace(MENU_PACE)         # отрисовка контекстного меню
                # Один раз за прогон снимаем состав меню: навигация ниже идёт
                # стрелками вслепую, и лишний пункт в меню уводит счётчик.
                # Р7 сейчас простаивает с раскрытым меню, поэтому длительность
                # дампа корректно вычитается из замера.
                self._cdp_dump_ui("контекстное меню ячейки (копирование)",
                                  charge_pace=True)
                # Точное попадание по подписи надёжнее счёта стрелок вслепую —
                # но работает, только если меню есть в DOM (см. issue #9).
                if not self._cdp_click_context_item(("копировать", "copy")):
                    safe_press('down', 2, pace=MENU_PACE)
                    safe_press('enter')
                self._pace(MENU_PACE)
                pyautogui.press('right', presses=paste_offset)
                pyautogui.click(button='right')
                self._pace(MENU_PACE)
                self._cdp_dump_ui("контекстное меню ячейки (вставка)",
                                  charge_pace=True)
                if not self._cdp_click_context_item(
                        ("вставить скопированные ячейки", "вставить ячейки",
                         "insert copied cells", "insert cells")):
                    safe_press('down', 3, pace=MENU_PACE)
                    safe_press('enter')
                # Р7-Офис показывает модалку «Вставить ячейки» — подтверждаем её.
                # Зеркалится в paste_pkm() Batch-режима.
                self._confirm_modal_enter()

            def add_column(method='hotkey'):
                # На CDP-пути оба варианта («горячие клавиши» и «меню Вставка»)
                # сводятся к одному вызову asc_insertCells — меню там нет.
                if self._cdp_add_column(key_pace=KEY_PACE):
                    return
                safe_hotkey('ctrl', 'pageup')
                self._pace(KEY_PACE)          # переключение листа
                pyautogui.press('right')
                if method == 'hotkey':
                    safe_hotkey('ctrl', 'shift', '=')
                else:
                    safe_hotkey('alt', 'i')
                    self._pace(MENU_PACE)     # раскрытие меню «Вставка»
                    safe_press('c')

            def paste_big():
                if self._cdp_paste_big(key_pace=KEY_PACE):
                    return
                safe_hotkey('shift', 'f11')
                self._pace(KEY_PACE)          # даём создаться новому листу
                safe_hotkey('ctrl', 'v')

            def vlookup():
                safe_hotkey('ctrl', 'pagedown')
                self._pace(KEY_PACE)          # переключение листа
                safe_hotkey('ctrl', 'home')
                pyperclip.copy('=VLOOKUP(A2;Лист1!A:B;2;FALSE)')
                safe_hotkey('ctrl', 'v')
                self._pace(KEY_PACE)          # формула должна попасть в ячейку
                safe_press('enter')
                safe_hotkey('ctrl', 'shift', 'down')
                safe_hotkey('ctrl', 'd')

            def select_all():
                """Ctrl+A с укороченным предохранителем.

                Выделив весь лист, Р7 пересчитывает агрегаты статусной строки по
                всем ячейкам и на большом файле держит CPU занятым десятками
                секунд. С общими OP_MAX_WAIT_SEC=180 это выглядело как зависание
                инструмента. Ограничиваем ожидание OP_SELECT_ALL_MAX_SEC: если
                Р7 не успел — операция честно помечается timeout, и прогон идёт
                дальше вместо трёхминутной паузы. Зеркалится в Batch-режиме.
                """
                self._op_max_wait = self.OP_SELECT_ALL_MAX_SEC
                if self._cdp_select_all():
                    return
                safe_hotkey('ctrl', 'a')

            def copy_all():
                """Ctrl+C по текущему выделению (после теста Ctrl+A — по всему листу)."""
                if self._cdp_copy():
                    return
                safe_hotkey('ctrl', 'c')

            def add_sheet():
                """Новый лист: Shift+F11 либо asc_addWorksheet через CDP."""
                if self._cdp_add_sheet():
                    return
                safe_hotkey('shift', 'f11')

            def del_column():
                safe_hotkey('ctrl', 'home')
                pyautogui.press('right')
                pyautogui.press('delete')

            def save_as_pdf():
                """Экспортирует текущий файл в PDF — запускает x2t (конвертер).

                Приоритет — хоткей Ctrl+Shift+S (Save As в Р7-Офис). Если диалог
                «Сохранить как» не появился за 3 сек, откатываемся на меню
                Файл → Сохранить как (Alt+F, затем навигация вниз и Enter) —
                конкретный пункт меню не проверен вживую на реальном Р7-Офис,
                это задокументированный запасной путь, требующий ручной проверки.

                Ожидание диалога — реакция Р7, поэтому оно остаётся в замере.
                Вычитается только безрезультатное ожидание перед запасным путём.
                Само время конвертации ловит _wait_operation_done: пока жив процесс
                x2t, операция считается незавершённой.
                """
                tmp_pdf = str(Path(os.environ.get("TEMP", ".")) /
                              f"temp_export_x2t_{int(time.time())}.pdf")
                # x2t стартует не мгновенно после Enter — просим детектор подождать
                # его дольше обычного, иначе экспорт будет помечен «ниже порога».
                self._op_start_grace = self.OP_PDF_GRACE_SEC

                safe_hotkey('ctrl', 'shift', 's')
                _t_dlg = time.time()
                if not self._wait_for_window_title(("сохранить как", "save as"), timeout=3.0):
                    # Диалог не открылся — эти 3 сек не время Р7, а наша неудача.
                    self._paced_total += time.time() - _t_dlg
                    self.add_test_log("   ⚠️ Ctrl+Shift+S не открыл диалог, пробуем меню Файл")
                    safe_hotkey('alt', 'f')
                    self._pace(MENU_PACE)
                    safe_press('down', 3, pace=MENU_PACE)
                    safe_press('enter')
                    self._wait_for_window_title(("сохранить как", "save as"), timeout=3.0)

                pyperclip.copy(tmp_pdf)
                safe_hotkey('ctrl', 'a')
                safe_hotkey('ctrl', 'v')
                self._pace(KEY_PACE)
                safe_press('enter')

            _test_ops = [
                ("Выделение всех ячеек (Ctrl+A)",      select_all),
                ("Копирование всех ячеек (Ctrl+C)",     copy_all),
                ("Вставка большого массива (Ctrl+V)",    paste_big),
                ("Добавление нового листа",              add_sheet),
                ("Добавление столбца (горячие клавиши)", lambda: add_column('hotkey')),
                ("Добавление столбца (меню Вставка)",    lambda: add_column('menu')),
                ("Вставка 1 ячейки (горячие клавиши)",   lambda: copy_paste_hotkey(1, 10)),
                ("Вставка 5 ячеек (горячие клавиши)",    lambda: copy_paste_hotkey(5, 15)),
                ("Вставка 1 ячейки (ПКМ)",               lambda: copy_paste_context(1, 10)),
                ("Вставка 5 ячеек (ПКМ)",                lambda: copy_paste_context(5, 15)),
                ("Функция ВПР (50K строк)",              vlookup),
                ("Удаление столбца (Del)",               del_column),
                ("Сохранение в PDF (конвертация x2t)",   save_as_pdf),
            ]

            def _update_status(text):
                """Safely updates the status bar from this worker thread —
                marshals onto the main thread via root.after and swallows
                errors from a window closed mid-run."""
                try:
                    self.root.after(0, lambda: self.status_var.set(text))
                except Exception:
                    pass

            # Прогресс и статус считаются только по включённым тестам — раньше
            # цикл шёл по всем 13 операциям и показывал «⚙ Название — N/13»
            # даже для снятых чекбоксом тестов, которые run_test_with_runs
            # молча пропускает.
            _active_ops = [op for op in _test_ops if op[0] in enabled_tests]
            # Семплер (запущен раньше, до try — см. комментарий там же)
            # начинает копить точки именно с этого момента: до сих пор RAM/CPU
            # ещё формировались самим открытием файла (переходный процесс), а
            # detect_leak() интересует дрейф ВО ВРЕМЯ теста, не старт.
            _resource_sampler.start()

            _run_start = time.time()
            self._set_perf_progress(0, len(_active_ops))
            for _i, (_name, _func) in enumerate(_active_ops, start=1):
                if stop_event.is_set():
                    self.add_test_log(
                        f"⏹ Остановлено пользователем ({_i - 1}/{len(_active_ops)} тестов выполнено)")
                    break
                _update_status(
                    f"⚙ {_name} — {_i}/{len(_active_ops)} "
                    f"(прошло {time.time() - _run_start:.0f} сек)")
                run_test_with_runs(_name, _func, test_runs.get(_name, DEFAULT_TEST_RUNS))
                self._set_perf_progress(_i, len(_active_ops))
            if not stop_event.is_set():
                _update_status(
                    f"✅ Готово: {len(_active_ops)}/{len(_active_ops)} "
                    f"(всего {time.time() - _run_start:.0f} сек)")
            self._cleanup_x2t_temp_pdfs()

            # ----- 5. Статистика ресурсов --------------------------------------------------
            ram_vals      = [r["ram"] for r in results if r.get("ram") is not None]
            cpu_vals      = [r["cpu"] for r in results if r.get("cpu") is not None]
            cpu_norm_vals = [r["cpu_normalized"] for r in results if r.get("cpu_normalized") is not None]
            peak_ram = max(ram_vals) if ram_vals else None
            avg_ram  = round(sum(ram_vals) / len(ram_vals), 1) if ram_vals else None
            min_ram  = min(ram_vals) if ram_vals else None
            peak_cpu = max(cpu_vals) if cpu_vals else None
            peak_cpu_norm = max(cpu_norm_vals) if cpu_norm_vals else None
            avg_cpu_norm  = round(sum(cpu_norm_vals) / len(cpu_norm_vals), 1) if cpu_norm_vals else None
            if peak_ram is not None:
                self.add_test_log(
                    f"📊 Пик RAM: {peak_ram:.1f} МБ  Средн: {avg_ram:.1f} МБ  Мин: {min_ram:.1f} МБ")
            if peak_cpu is not None:
                self.add_test_log(
                    f"📊 Пик CPU: {peak_cpu:.1f}% (сырое)  {peak_cpu_norm:.1f}% (норм., "
                    f"{psutil.cpu_count() if PSUTIL_OK else '?'} ядер)")

            # ── Детектор утечек (этап 2, H3) ────────────────────────────────────
            # Останавливаем сразу после операций теста, до сохранения отчётов и
            # закрытия Р7 — семплер должен покрывать сам прогон, не переходные
            # процессы вокруг него. finally ниже вызовет stop() повторно на
            # случай исключения выше (идемпотентно, безопасно).
            _resource_sampler.stop()
            _resource_sampler.join(timeout=5)
            leak_verdict = detect_leak(_resource_sampler.snapshot())
            if leak_verdict["leak"] is True:
                self.add_test_log(f"⚠️ {leak_verdict['verdict']}")
            elif leak_verdict["leak"] is False:
                self.add_test_log(f"✅ {leak_verdict['verdict']}")
            # leak is None (мало замеров — короткий прогон/мало включённых
            # тестов) — логировать нечего, это ожидаемо, не предупреждение.

            # ----- 6. Сохранение отчётов ---------------------------------------------------
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Timestamp в имени — иначе каждый следующий прогон затирает Excel-
            # и HTML-отчёт предыдущего (performance_full_*.json и так уже был
            # уникальным на прогон, эти два — нет).
            REPORT_FILE = self.reports_folder / f"Performance_Report_{ts}.xlsx"
            HTML_REPORT_PATH = REPORT_FILE.with_suffix(".html")
            try:
                REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)

                # Excel
                from openpyxl import Workbook as WB
                wb = WB()
                ws = wb.active
                ws.title = "Результаты"
                ws.append(["Операция", "Время (сек)", "RAM (МБ)", "CPU (%)", "Ошибка"])
                for r in results:
                    ws.append([r["name"], round(r["time"], 2),
                               r.get("ram") or "", r.get("cpu") or "",
                               r.get("error") or ""])
                wb.save(str(REPORT_FILE))
                self.add_test_log(f"📊 Excel-отчёт сохранён: {REPORT_FILE}")

                # JSON (полные данные для последующего сравнения версий)
                json_path = self.reports_folder / f"performance_full_{ts}.json"
                full_data = {
                    "timestamp": ts,
                    "measure_schema": MEASURE_SCHEMA_VERSION,
                    "version": self.current_version_info.get("name") if self.current_version_info else None,
                    "test_file": str(test_file),
                    "system": self._build_system_info(),
                    "summary": {
                        "peak_ram_mb": peak_ram,
                        "avg_ram_mb": avg_ram,
                        "min_ram_mb": min_ram,
                        "peak_cpu_pct": peak_cpu,
                        "peak_cpu_normalized_pct": peak_cpu_norm,
                        "avg_cpu_normalized_pct": avg_cpu_norm,
                        "leak_detection": leak_verdict,
                    },
                    "results": results,
                }
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(full_data, f, indent=2, ensure_ascii=False)
                self.add_test_log(f"📄 JSON-данные сохранены: {json_path.name}")

                # HTML
                version_str = (self.current_version_info.get("name")
                               if self.current_version_info else None)
                html_content = self._generate_html_report(
                    results, test_file, open_elapsed, version_str,
                    ram_vals, cpu_vals, peak_ram, avg_ram, min_ram, peak_cpu,
                )
                with open(HTML_REPORT_PATH, "w", encoding="utf-8") as f:
                    f.write(html_content)
                self.add_test_log(f"📄 HTML-отчёт сохранён: {HTML_REPORT_PATH}")

            except Exception as e:
                self.add_test_log(f"⚠️ Ошибка сохранения отчётов: {e}")

            # ----- 7. Закрытие -------------------------------------------------------------
            _upd_stop.set()
            self.add_test_log("🔍 Мониторинг окна обновления остановлен")
            self.add_test_log("🔚 Закрытие Р7-Офис...")
            self._close_r7_gracefully(find_r7_window())
            self.add_test_log("🏁 Тест завершён.")

            # ----- 8. Диалог после теста ---------------------------------------------------
            self.root.after(0, lambda: self._show_post_test_dialog(HTML_REPORT_PATH, ts))
        finally:
            # Поток-монитор диалога обновления не должен пережить эту функцию —
            # раньше _upd_stop.set() стоял в линейном коде, и любое исключение
            # выше оставляло монитор сканировать все окна системы до закрытия
            # приложения. CDP/Selenium-соединение и семплер ресурсов — тот же
            # случай: должны остановиться независимо от того, как функция
            # завершилась. _resource_sampler.stop() безопасно вызывать даже
            # если .start() выше так и не случился (исключение до него) —
            # это просто Event.set(), не требует живого потока.
            _upd_stop.set()
            _resource_sampler.stop()
            self._close_webdriver_connector()

    # ---------------------- Вспомогательные методы (ресурсы, отчёты) ------

    # Подстроки "r7"/"р7" ловили и собственные сборки инструмента —
    # R7-Testovarka.exe и R7Manager.exe совпадают с той же маской, что и
    # editors_helper.exe. Список сужен до реальных бинарников Р7-Офис.
    # Имена процессов Р7-Офис. "editors" здесь ОБЯЗАТЕЛЕН и добавлен не для
    # полноты: 25.08.2026 выяснилось, что главный процесс приложения на живой
    # сборке называется editors.exe, и под прежнюю маску он не подходил
    # (проверка "editors_helper" in "editors.exe" даёт False). Последствия
    # были втройне неприятными:
    #   1. _terminate_r7_processes не мог завершить Р7 вообще: убивались
    #      только дочерние рендереры editors_helper.exe, а живой родитель
    #      немедленно создавал их заново. Документированный «штатный
    #      запасной путь» через форс-килл на этой сборке не работал, и
    #      попытка закрыть Р7 уходила в бесконечный цикл respawn.
    #   2. Замеры RAM/CPU теряли главный процесс целиком (84 МБ RSS в
    #      наблюдавшемся случае) — все прежние цифры памяти занижены.
    #   3. _wait_operation_done определял «Р7 простаивает» по неполному
    #      набору процессов, то есть мог объявить операцию законченной,
    #      пока главный процесс ещё работал.
    # Дерево процессов на сборке 2026.2.2.x:
    #   DesktopEditors.exe (лаунчер, сразу завершается)
    #     └── editors.exe (главный процесс приложения)
    #           └── editors_helper.exe × N (рендереры CEF)
    #
    # Сравнение идёт по ТОЧНОМУ имени (см. _matches_r7_process), а не по
    # вхождению подстроки: маска "editors" как подстрока цепляла бы любой
    # сторонний процесс со словом editors в имени, а инструмент по этому
    # списку не только меряет, но и убивает процессы.
    _R7_PROCESS_NAMES = ("editors.exe", "editors_helper.exe", "desktopeditors.exe")

    # x2t — конвертер документов, отдельный короткоживущий процесс. Его имя
    # содержит версию (x2t.exe, x2t64.exe и т.п.), поэтому он единственный,
    # кто сопоставляется по префиксу, а не по точному имени.
    _R7_PROCESS_PREFIXES = ("x2t",)

    @classmethod
    def _matches_r7_process(cls, name):
        """True, если имя процесса принадлежит Р7-Офис.

        Args:
            name: Имя процесса, как его отдаёт psutil (с расширением).

        Returns:
            bool
        """
        low = (name or "").lower()
        if low in cls._R7_PROCESS_NAMES:
            return True
        return any(low.startswith(pref) for pref in cls._R7_PROCESS_PREFIXES)

    def _get_r7_processes(self, log_cb=None):
        """Returns list of psutil.Process objects for all R7-Office related processes.

        Сопоставление — по точному имени (_R7_PROCESS_NAMES) плюс префикс для
        x2t (_R7_PROCESS_PREFIXES), см. _matches_r7_process. Маски "r7"/"р7"
        сознательно отсутствуют: под них попадали собственные процессы
        инструмента, R7-Testovarka.exe и R7Manager.exe. Свой PID и PID родителя исключаются явно — второй рубеж
        защиты на случай, если Р7-Офис когда-нибудь переименует исполняемый файл
        во что-то похожее на маску.

        x2t — внутренний конвертер документов Р7-Офис, отдельный процесс, который
        может давать заметный вклад в общую RAM/CPU при открытии/сохранении файлов.
        If self._r7_pids is set (from a previous call), tries direct PID lookup first.

        Args:
            log_cb: Callback for the x2t detection line; defaults to self.add_test_log.
        """
        if log_cb is None:
            log_cb = self.add_test_log

        if not PSUTIL_OK:
            return []

        if not hasattr(self, "_x2t_logged_pids"):
            self._x2t_logged_pids = set()

        own_pid = os.getpid()
        try:
            parent_pid = psutil.Process(own_pid).ppid()
        except Exception:
            parent_pid = None
        excluded_pids = {own_pid} | ({parent_pid} if parent_pid else set())

        # Fast path: try previously discovered PIDs directly
        if getattr(self, "_r7_pids", None):
            procs = []
            for pid in self._r7_pids:
                if pid in excluded_pids:
                    continue
                try:
                    p = psutil.Process(pid)
                    p.name()  # raises NoSuchProcess if dead
                    procs.append(p)
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            if procs:
                return procs

        # Full scan
        found = []
        try:
            for proc in psutil.process_iter(["name", "pid"]):
                try:
                    pid = proc.info.get("pid")
                    if pid in excluded_pids:
                        continue
                    name = (proc.info.get("name") or "").lower()
                    if self._matches_r7_process(name):
                        found.append(proc)
                        if "x2t" in name:
                            if pid not in self._x2t_logged_pids:
                                log_cb(
                                    f"🔧 Обнаружен процесс конвертации x2t: "
                                    f"PID={pid}, имя={proc.info.get('name')}"
                                )
                                self._x2t_logged_pids.add(pid)
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except Exception:
            pass

        # Cache PIDs for subsequent fast-path calls
        self._r7_pids = [p.pid for p in found]
        return found

    def _cpu_count(self):
        """Число логических ядер, закэшированное на экземпляр.

        `psutil.cpu_count()` не меняется на живой машине за время одного
        прогона, а вызывающий код (детекторы простоя) опрашивает его на
        каждой итерации цикла (каждые 0.05–0.15 с) — кэш убирает системный
        вызов из горячего пути. `or 1` — на случай, если psutil не смог
        определить число ядер (документированная возможность API, а не
        гипотетический случай): без подстраховки деление на None упало бы.

        Тот же паттерн, что у `_find_r7_path`/`self._cached_r7_path` —
        прямая проверка атрибута, выставленного в `__init__`, а не
        `getattr(..., None)`.

        Returns:
            int: число логических ядер, минимум 1.
        """
        if self._cached_cpu_count:
            return self._cached_cpu_count
        self._cached_cpu_count = (psutil.cpu_count() or 1) if PSUTIL_OK else 1
        return self._cached_cpu_count

    def _build_system_info(self):
        """Окружение прогона для JSON-результатов — общий код для обоих
        воркеров (одиночный тест и Batch), раньше продублированный дословно
        в двух местах.

        Returns:
            dict: {"os", "ram_total_gb", "cpu_model", "cpu_cores_logical"}.
        """
        sys_mem_gb = (round(psutil.virtual_memory().total / (1024 ** 3), 1)
                     if PSUTIL_OK else None)
        return {
            "os": platform.platform(),
            "ram_total_gb": sys_mem_gb,
            "cpu_model": platform.processor() or None,
            # Нужно, чтобы сравнивать нормированный CPU (measure_schema 2,
            # см. OP_IDLE_CPU_PCT) между стендами осмысленно — без числа
            # ядер нормированный процент сам по себе не восстановить обратно
            # в сырую загрузку.
            "cpu_cores_logical": self._cpu_count(),
        }

    @staticmethod
    def _mad(values):
        """Median Absolute Deviation — устойчивая мера разброса, пара к
        медиане (в статистике нет готовой функции для этого — sample stdev
        есть, MAD нет, см. модуль statistics). Не масштабируется константой
        1.4826 (переводящей MAD в оценку, сравнимую со стандартным
        отклонением для нормального распределения) — здесь скорость
        операций Р7 ничем не гарантированно нормальна, само значение MAD
        интересно как «типичное отклонение от медианы» в секундах, не как
        оценка сигмы.

        Args:
            values: Непустая последовательность чисел.

        Returns:
            float: MAD. 0.0, если все значения совпадают.
        """
        center = statistics.median(values)
        return statistics.median(abs(v - center) for v in values)

    def _sample_r7_resources(self, procs):
        """Снимает агрегированные метрики RAM/CPU/потоков/аптайма по списку процессов Р7.

        CPU нормализуется делением на psutil.cpu_count(): «сырое» значение psutil
        может превышать 100% на многоядерных системах (сумма по всем ядрам), а
        «норм.» приводит его к шкале 0–100%, как в диспетчере задач Windows.

        Args:
            procs: список psutil.Process, обычно результат _get_r7_processes().

        Returns:
            dict | None: {"ram_mb", "cpu_raw_pct", "cpu_norm_pct", "threads",
            "uptime_sec"}, или None если psutil недоступен или ни один процесс не жив.
        """
        if not (PSUTIL_OK and procs):
            return None

        total_ram_mb  = 0.0
        total_cpu_raw = 0.0
        total_threads = 0
        oldest_create = None
        alive = 0
        now = time.time()

        for p in procs:
            # Память: основной метрик; процесс считается "живым" если RAM читается успешно
            try:
                total_ram_mb += p.memory_info().rss / (1024 * 1024)
                alive += 1
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

            # CPU: суммируем по всем процессам Р7 (редактор + x2t могут работать
            # одновременно), а не берём max — max одного процесса занижал бы
            # реальную суммарную нагрузку на систему.
            try:
                total_cpu_raw += p.cpu_percent(interval=0.1)
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

            # Потоки: каждый вызов независим
            try:
                total_threads += p.num_threads()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

            # Время создания: каждый вызов независим
            try:
                create_ts = p.create_time()
                if oldest_create is None or create_ts < oldest_create:
                    oldest_create = create_ts
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

        if alive == 0:
            return None

        cpu_count = self._cpu_count()
        return {
            "ram_mb":       round(total_ram_mb, 1),
            "cpu_raw_pct":  round(total_cpu_raw, 1),
            "cpu_norm_pct": round(total_cpu_raw / cpu_count, 1),
            "threads":      total_threads,
            "uptime_sec":   round(now - oldest_create, 1) if oldest_create is not None else None,
        }

    def _log_resources(self, sample, log_cb=None):
        """Форматированный вывод одного замера ресурсов с цветовой индикацией CPU.

        Индикатор считается по нормализованному CPU (0–100%, все ядра):
        🟢 < 50% — обычная нагрузка, 🟡 50–79.9% — средняя, 🔴 ≥ 80% — высокая.

        Args:
            sample: dict из _sample_r7_resources(), либо None (тогда ничего не пишет).
            log_cb: функция логирования; по умолчанию self.add_test_log.
        """
        if sample is None:
            return
        if log_cb is None:
            log_cb = self.add_test_log

        cpu_norm = sample.get("cpu_norm_pct")
        if cpu_norm is None:
            icon = "⚪"
        elif cpu_norm < 50:
            icon = "🟢"
        elif cpu_norm < 80:
            icon = "🟡"
        else:
            icon = "🔴"

        uptime = sample.get("uptime_sec")
        uptime_str = f"{uptime:.0f} сек" if uptime is not None else "—"

        log_cb(
            f"   📊 RAM: {sample['ram_mb']:.1f} МБ  "
            f"CPU: {sample['cpu_raw_pct']:.1f}% (норм. {cpu_norm if cpu_norm is not None else '—'}%) {icon}  "
            f"Потоки: {sample['threads']}  Аптайм: {uptime_str}"
        )

    # ---------------------- Замер операции ----------------------

    def _pace(self, seconds):
        """Преднамеренная пауза внутри измеряемой операции.

        Нужна там, где Р7-Офис физически не успевает за клавиатурой: между
        открытием меню и выбором пункта, между Ctrl+C и Ctrl+V и т.п. В отличие
        от прежних time.sleep() и pyautogui.PAUSE, это время накапливается в
        self._paced_total и вычитается из результата замера — то есть пауза
        обеспечивает надёжность автоматизации, но не попадает в цифру
        производительности.

        Args:
            seconds: Длительность паузы.
        """
        if seconds <= 0:
            return
        t0 = time.time()
        time.sleep(seconds)
        self._paced_total += time.time() - t0

    def _confirm_modal_enter(self, pace=None):
        """Подтверждает модалку «Вставить ячейки» — часть, которая обязана
        находиться ВНУТРИ окна замера.

        Диалог сдвига ячеек — HTML-модалка внутри CEF, а не окно ОС: win32gui её
        не видит (тот же случай, что и кнопка «Жирный», см.
        r7_webdriver_connector.py), поэтому дождаться её появления штатным
        _wait_for_window_title нельзя — остаётся слепой Enter. На OP_MENU_PACE
        (0.12 с) это гонка: на нагруженном документе модалка не успевает
        отрисоваться, Enter уходит в сетку, диалог остаётся висеть и ломает все
        последующие операции прогона. Именно это возвращало баг, который PR #4
        считал закрытым.

        Почему здесь ровно один Enter, а повторы — в _flush_pending_modal_confirm.
        Пауза ниже проходит через _pace() и вычитается из замера, и это корректно
        только пока Р7 действительно простаивает, показывая модалку. Сразу после
        подтверждающего Enter начинается настоящая работа (вставка ячеек) — сон и
        вычитание в этот момент отняли бы у результата время, которое Р7 реально
        работал: операция короче секунды выдавала бы 0/below_floor. Поэтому func()
        возвращает управление сразу после подтверждения, а страховочные повторы
        уходят за границу замера.

        Args:
            pace: Пауза на отрисовку модалки; по умолчанию OP_DIALOG_PACE.
        """
        if pace is None:
            pace = self.OP_DIALOG_PACE
        # Р7 в этот момент простаивает, ожидая ввода, — вычитать паузу корректно.
        self._pace(pace)
        pyautogui.press('enter')
        # Дальше начинается работа Р7: замер должен идти без наших пауз.
        self._pending_modal_confirm = True

    def _flush_pending_modal_confirm(self, log_cb=None):
        """Досылает страховочные Enter'ы по модалке — уже ВНЕ окна замера.

        Нужно на случай, когда модалка не успела появиться за OP_DIALOG_PACE:
        тогда подтверждающий Enter из _confirm_modal_enter ушёл в сетку, модалка
        всплыла позже и висит. Здесь она добивается, не искажая цифру: замер к
        этому моменту уже закрыт (_wait_operation_done отработал), поэтому пауза
        обычная time.sleep, а не _pace.

        Лишний Enter безвреден: если модалки нет, он лишь сдвигает активную
        ячейку на строку вниз, а следующий тест всё равно начинается с Ctrl+Home.
        Последовательность самоисправляющаяся — какой бы из Enter'ов ни совпал с
        появлением модалки, она закроется.

        Вызывать в обоих воркерах сразу после _wait_operation_done (см. правило
        зеркалирования в CLAUDE.md).

        Args:
            log_cb: Функция логирования; по умолчанию self.add_test_log.
        """
        if not self._pending_modal_confirm:
            return
        self._pending_modal_confirm = False
        # Диагностика (один раз за прогон): что на самом деле висит на экране
        # после подтверждающего Enter. Именно эти подписи нужны, чтобы чинить
        # слепую навигацию стрелками по контекстному меню — сейчас счётчик
        # `down` подобран вслепую и уезжает, стоит меню обзавестись пунктом.
        # Стоит вне замера, поэтому на цифры не влияет.
        self._cdp_dump_ui("после подтверждения модалки «Вставить ячейки»", log_cb=log_cb)
        for _ in range(max(0, self.OP_DIALOG_ATTEMPTS - 1)):
            time.sleep(self.OP_DIALOG_PACE)
            try:
                pyautogui.press('enter')
            except Exception as e:
                (log_cb or self.add_test_log)(
                    f"   ⚠️ Не удалось дослать Enter по модалке: {e}")
                return

    def _wait_for_window_title(self, substrings, timeout=3.0):
        """Ждёт появления видимого окна с подходящим заголовком.

        Время ожидания — это реакция Р7-Офис, поэтому оно НЕ вычитается из
        замера (в отличие от _pace). Вызывающий код вычитает его сам, если
        ожидание оказалось безрезультатным.

        Args:
            substrings: Подстроки заголовка (без учёта регистра).
            timeout: Максимум секунд ожидания.

        Returns:
            bool: True, если окно появилось.
        """
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self._win_title_contains(*substrings):
                return True
            time.sleep(self.OP_POLL_SEC)
        return False

    def _wait_operation_done(self, hwnd, log_cb=None, start_grace=None):
        """Ждёт, пока Р7-Офис закончит обрабатывать только что отправленную операцию.

        Прежде замер операции заканчивался на последнем нажатии клавиши:
        pyautogui только кладёт события во входную очередь и возвращается сразу,
        поэтому «время операции» равнялось сумме собственных задержек скрипта.
        Здесь секундомер останавливается по реальному признаку — Р7 перестал
        быть занятым.

        Занятость определяется по трём признакам (любой означает «занят»):
          1. Окно не прокачивает очередь сообщений за OP_RESPONSIVE_MS.
          2. Процессы Р7 грузят CPU не ниже OP_IDLE_CPU_PCT.
          3. Жив конвертер x2t — им идёт экспорт в PDF.

        Возвращается момент НАЧАЛА простоя, а не конец окна подтверждения,
        поэтому OP_IDLE_SAMPLES не добавляется к результату замера.

        Args:
            hwnd: Дескриптор окна Р7 либо функция его поиска.
            log_cb: Функция логирования; по умолчанию self.add_test_log.
            start_grace: Сколько ждать начала работы Р7. По умолчанию
                OP_START_GRACE_SEC; экспорт в PDF просит больше через
                self._op_start_grace, потому что x2t стартует с задержкой.

        Returns:
            tuple[float | None, str]: (момент завершения, статус).
              "ok"          — работа началась и закончилась;
              "below_floor" — Р7 не стал занятым за OP_START_GRACE_SEC, то есть
                              операция быстрее порога измерения;
              "timeout"     — не дождались за OP_MAX_WAIT_SEC (момент — None).
        """
        if log_cb is None:
            log_cb = self.add_test_log

        if start_grace is None:
            start_grace = getattr(self, "_op_start_grace", None) or self.OP_START_GRACE_SEC

        # Предохранитель на операцию — как и start_grace, его может укоротить
        # сама тест-функция через self._op_max_wait (см. select_all).
        max_wait = getattr(self, "_op_max_wait", None) or self.OP_MAX_WAIT_SEC
        start    = time.time()
        deadline = start + max_wait

        cur_hwnd     = None if callable(hwnd) else hwnd
        tracked      = {}     # pid -> (psutil.Process с «прогретым» CPU, имя)
        last_refresh = 0.0
        last_cpu_at  = 0.0
        last_cpu     = 0.0
        seen_busy    = False
        idle_streak  = 0
        idle_since   = None

        while time.time() < deadline:
            now = time.time()

            if callable(hwnd):
                if not (cur_hwnd and WIN32_OK and win32gui.IsWindow(cur_hwnd)):
                    cur_hwnd = hwnd()

            if PSUTIL_OK and now - last_refresh >= self.OP_PROC_REFRESH_SEC:
                last_refresh = now
                self._r7_pids = None
                for p in self._get_r7_processes(log_cb=log_cb):
                    if p.pid in tracked:
                        continue
                    try:
                        name = (p.name() or "").lower()
                        p.cpu_percent(None)
                        tracked[p.pid] = (p, name)
                    except (psutil.NoSuchProcess, psutil.AccessDenied):
                        pass

            # x2t проверяем на каждом опросе — он короткоживущий, и лишние
            # 0.2 сек ожидания его смерти уехали бы прямо в замер PDF-экспорта.
            converter_alive = False
            for pid, (p, name) in list(tracked.items()):
                if "x2t" not in name:
                    continue
                try:
                    if p.is_running():
                        converter_alive = True
                    else:
                        tracked.pop(pid, None)
                except Exception:
                    tracked.pop(pid, None)

            if PSUTIL_OK and now - last_cpu_at >= self.OP_CPU_WINDOW_SEC:
                last_cpu_at = now
                total = 0.0
                dead  = []
                for pid, (p, _name) in tracked.items():
                    try:
                        total += p.cpu_percent(None)
                    except (psutil.NoSuchProcess, psutil.AccessDenied):
                        dead.append(pid)
                for pid in dead:
                    tracked.pop(pid, None)
                # Нормировка на число ядер — см. комментарий у OP_IDLE_CPU_PCT
                # (measure_schema 2): без неё порог зависел от числа ядер
                # стенда, а не от реальной занятости Р7.
                last_cpu = total / self._cpu_count()

            responsive = self._window_responsive(cur_hwnd, self.OP_RESPONSIVE_MS)
            busy = (not responsive) or converter_alive or (
                PSUTIL_OK and last_cpu >= self.OP_IDLE_CPU_PCT)

            if busy:
                seen_busy   = True
                idle_streak = 0
                idle_since  = None
            else:
                if idle_streak == 0:
                    idle_since = now
                idle_streak += 1
                if seen_busy and idle_streak >= self.OP_IDLE_SAMPLES:
                    return idle_since, "ok"
                if not seen_busy and now - start >= start_grace:
                    # Р7 вообще не стал занятым — операция быстрее, чем мы умеем мерить.
                    return start, "below_floor"

            time.sleep(self.OP_POLL_SEC)

        log_cb(f"   ⚠️ Р7-Офис не освободился за {max_wait:.0f} сек")
        return None, "timeout"

    # ---------------------- Готовность документа ----------------------

    def _window_responsive(self, hwnd, timeout_ms=None):
        """True, если окно вынуло сообщение из очереди за timeout_ms.

        WM_NULL ничего не делает, но SendMessageTimeout возвращается ровно
        тогда, когда окно прокачало очередь сообщений — это прямое измерение
        занятости UI-потока.

        Прежний зонд через pyautogui.hotkey('ctrl','End') так не умел:
        keybd_event только кладёт событие во входную очередь и возвращается
        сразу, не дожидаясь обработки. Поэтому его длительность всегда была
        одинаковой (≈0.3 сек — сумма interval и PAUSE самого pyautogui) и о
        состоянии Р7-Офис не говорила ничего.

        Args:
            hwnd: Дескриптор окна Р7-Офис. None или отсутствие pywin32 → True
                (проверка пропускается, решение принимается по CPU).
            timeout_ms: Порог ожидания; по умолчанию READY_RESPONSIVE_MS.

        Returns:
            bool
        """
        if not WIN32_OK or not hwnd:
            return True
        if timeout_ms is None:
            timeout_ms = self.READY_RESPONSIVE_MS
        try:
            res = win32gui.SendMessageTimeout(
                hwnd, win32con.WM_NULL, 0, 0,
                win32con.SMTO_ABORTIFHUNG, int(timeout_ms))
        except Exception:
            # pywintypes.error с ERROR_TIMEOUT — окно не разгребает очередь
            return False
        # pywin32 возвращает (result, lresult); result == 0 — тоже таймаут
        if isinstance(res, tuple):
            return bool(res[0])
        return True

    def _find_bold_button_hwnd(self, hwnd):
        """Ищет окно кнопки «Жирный» среди ВСЕХ потомков hwnd (рекурсивно,
        через EnumChildWindows — не FindWindowEx с проверкой только прямых
        детей: реальная кнопка, если она вообще существует как нативное
        окно, почти наверняка вложена глубже одного уровня).

        Совпадением считается окно с классом из BOLD_BUTTON_CLASSES и
        текстом из BOLD_BUTTON_LABELS (без учёта регистра) — независимо от
        текущего состояния enabled/disabled, это отдельная проверка.

        Args:
            hwnd: Окно Р7-Офис, в поддереве которого искать кнопку.

        Returns:
            int | None: hwnd найденной кнопки, либо None.
        """
        if not (WIN32_OK and hwnd):
            return None

        import win32gui

        needles = set(self.BOLD_BUTTON_LABELS)
        found = [None]

        def _walk(h, _):
            if found[0] is not None:
                return
            try:
                cls = win32gui.GetClassName(h)
                if cls not in self.BOLD_BUTTON_CLASSES:
                    return
                # "&" — маркер мнемоники Win32 (подчёркивает следующую букву
                # при Alt), не часть подписи: подпись "&B" на экране выглядит
                # как "B". Без снятия "&" сравнение "&b" == "b" не совпало бы,
                # и кнопка с настоящим акселератором осталась бы незамеченной.
                text = win32gui.GetWindowText(h).replace("&", "").strip().lower()
                if text in needles:
                    found[0] = h
            except Exception:
                pass

        try:
            win32gui.EnumChildWindows(hwnd, _walk, None)
        except Exception:
            return None
        return found[0]

    def _is_bold_button_visible(self, hwnd):
        """Проверяет, доступна ли на панели инструментов Р7 кнопка «Жирный»
        (найдена через _find_bold_button_hwnd и IsWindowEnabled() — True).

        ПРОВЕРЕНО ЭКСПЕРИМЕНТАЛЬНО (не предположение): на установленной здесь
        сборке (2026.2.2.x) панель инструментов — не набор нативных Win32-
        виджетов, а HTML внутри окна рендера CEF. Полный дамп дерева дочерних
        окон главного hwnd через win32gui.EnumChildWindows дал 40 узлов, все
        классов Qt5152QWindowIcon / CefBrowserWindow / Chrome_WidgetWin_0 /
        Chrome_RenderWidgetHostHWND — ни одного "Button" или "ToolbarButton".
        Сама кнопка реально существует, но в DOM: запуск с флагом
        --ascdesktop-support-debug-info открывает отладочный порт Chrome
        DevTools Protocol (см. вывод "DevTools listening on ws://..." в
        консоли), через который она была найдена как
        <button id="id-toolbar-btn-bold" class="btn btn-toolbar"> внутри
        iframe apps/spreadsheeteditor/main/ — на 3 уровня вложенности глубже
        top-level окна, недостижима никаким Win32 API в принципе.

        Метод оставлен на случай другой версии/сборки Р7, где панель
        нарисована классическими Win32-виджетами; _wait_until_r7_ready
        корректно откатывается на CPU+WM_NULL, если кнопка не находится (и,
        как показывает проверка выше, на этой сборке будет откатываться
        всегда).

        Args:
            hwnd: Окно Р7-Офис, в поддереве которого искать кнопку.

        Returns:
            bool: True, если кнопка найдена и включена.
        """
        btn = self._find_bold_button_hwnd(hwnd)
        if btn is None:
            return False
        try:
            import win32gui
            return bool(win32gui.IsWindowEnabled(btn))
        except Exception:
            return False

    def _wait_for_bold_button(self, hwnd, timeout=None):
        """Ждёт, пока кнопка «Жирный» на панели инструментов Р7 станет
        доступна, либо не истечёт timeout.

        Сначала дешёвая разовая проверка существования окна вообще
        (_find_bold_button_hwnd). Если его нет — как на CEF-сборках, см.
        _is_bold_button_visible, — возвращает False немедленно, не тратя
        timeout впустую: кнопки нет, ждать нечего. Иначе опрашивает
        _is_bold_button_visible каждые BOLD_BUTTON_POLL_SEC, пока кнопка не
        станет доступна — окно ищется заново на каждом опросе (не
        кэшируется), это устойчивее к случаю, если панель успеет
        перестроиться, и стоит того при бюджете в единицы секунд.

        Отдельная, ограниченная по времени фаза внутри _wait_until_r7_ready —
        не основной цикл опроса готовности. Вызывается не более одного раза
        за вызов _wait_until_r7_ready (см. его docstring).

        Args:
            hwnd: Окно Р7-Офис.
            timeout: Секунд ожидания; по умолчанию BOLD_BUTTON_TIMEOUT_SEC.

        Returns:
            bool: True, если кнопка стала доступна в пределах timeout.
        """
        # Сначала — дешёвая разовая проверка существования окна вообще.
        # Если его нет (как на CEF-сборках, см. _is_bold_button_visible),
        # выходим сразу: незачем тратить timeout на опрос несуществующего
        # окна, IsWindowEnabled() на каждой итерации всё равно даст ту же
        # ошибку.
        if self._find_bold_button_hwnd(hwnd) is None:
            return False

        if timeout is None:
            timeout = self.BOLD_BUTTON_TIMEOUT_SEC
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self._is_bold_button_visible(hwnd):
                return True
            time.sleep(self.BOLD_BUTTON_POLL_SEC)
        return False

    # ── CDP-триггер готовности (кнопка «Жирный» в DOM) ─────────────────────
    # В отличие от _wait_for_bold_button (win32gui) — реально видит кнопку:
    # панель инструментов Р7 рисуется как HTML внутри CEF-рендера, а не
    # набором нативных Win32-виджетов (см. r7_webdriver_connector.py и
    # коммит 7978206). Требует, чтобы Р7 в этом запуске был стартован с
    # --ascdesktop-support-debug-info (см. _prepare_webdriver_launch) — без
    # этого self._webdriver_connector остаётся None, и весь блок ниже
    # молча ничего не делает, оставляя работу win32gui/CPU-логике.

    @staticmethod
    def _cdp_port_free(port, timeout=0.2):
        """Проверяет, свободен ли TCP-порт на localhost.

        Args:
            port: Порт для проверки.
            timeout: Секунд на попытку подключения.

        Returns:
            bool: True, если порт свободен (никто не слушает на нём).
        """
        import socket
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(timeout)
            return s.connect_ex(("127.0.0.1", port)) != 0

    def _prepare_webdriver_launch(self, log_cb=None, filename_hint=None):
        """Готовит CDP-подключение к следующему запуску Р7-Офис: подбирает
        порт, создаёт (но не подключает — порт ещё не открыт, процесс не
        запущен) self._webdriver_connector и возвращает доп. аргументы
        командной строки, которые нужно добавить к [r7_path, file] в Popen.

        Вызывать непосредственно перед subprocess.Popen, который запускает
        Р7 — записывает состояние (self._webdriver_connector,
        self._current_webdriver_port) для этого конкретного запуска.

        Args:
            filename_hint: Имя открываемого файла (test_file.name) —
                передаётся в R7WebDriverConnector и используется при выборе
                CDP-цели, если в момент подключения окажется открыто больше
                одного документа (см. R7WebDriverConnector.filename_hint,
                H5). В сегодняшней архитектуре (один документ на запуск Р7)
                на выбор цели не влияет — единственная doctype=-цель
                находится и без фильтра; готовит почву для многодокументных
                сценариев.

        Порт по умолчанию — DEFAULT_CDP_PORT (8080), подтверждённый на живом
        Р7-Офис. Если он занят (например, завис процесс от прошлого
        прогона), пробует DEFAULT_CDP_PORT+1, +2 с явным
        --remote-debugging-port=<port> — этот путь НЕ подтверждён
        эмпирически (см. docstring r7_webdriver_connector.py: похоже, что
        сама Р7 порт из флага не читает и всегда слушает 8080) — но
        передать его безопасно, он будет просто проигнорирован, если Р7 его
        не понимает, и _wait_for_bold_button_cdp тогда не найдёт живой
        порт и молча откатится на win32gui/CPU.

        Если и 8080, и запасные заняты — CDP для этого запуска отключается
        (self._webdriver_connector = None), тест идёт по обычному пути.
        Это безопасный откат, не ошибка.

        Args:
            log_cb: Функция логирования; по умолчанию self.add_test_log.

        Returns:
            list[str]: доп. аргументы для subprocess.Popen (пустой список,
                если WebDriver недоступен или свободного порта не нашлось).
        """
        if log_cb is None:
            log_cb = self.add_test_log
        self._webdriver_connector = None
        self._current_webdriver_port = None
        if not WEBDRIVER_OK:
            return []

        if self._cdp_port_free(DEFAULT_CDP_PORT):
            self._current_webdriver_port = DEFAULT_CDP_PORT
            self._webdriver_connector = R7WebDriverConnector(
                DEFAULT_CDP_PORT, log_cb=log_cb, filename_hint=filename_hint)
            return r7_launch_debug_args()

        log_cb(f"⚠️ CDP-порт {DEFAULT_CDP_PORT} занят — пробую запасные "
               f"(--remote-debugging-port не подтверждён на реальной Р7)")
        for candidate in (DEFAULT_CDP_PORT + 1, DEFAULT_CDP_PORT + 2):
            if self._cdp_port_free(candidate):
                self._current_webdriver_port = candidate
                self._webdriver_connector = R7WebDriverConnector(
                    candidate, log_cb=log_cb, filename_hint=filename_hint)
                return r7_launch_debug_args(port=candidate)

        log_cb("⚠️ Свободный CDP-порт не найден — WebDriver-триггер отключён для этого запуска")
        return []

    def _close_webdriver_connector(self):
        """Закрывает CDP/Selenium-соединение текущего запуска Р7, если оно
        было открыто (self._webdriver_connector, см.
        _prepare_webdriver_launch). Безопасно вызывать всегда — в том числе
        когда WebDriver в этом запуске не использовался вовсе.

        Вызывать из finally там же, где останавливается монитор диалога
        обновления (_upd_stop.set()) — оба ресурса живут на весь запуск Р7
        и должны быть освобождены, даже если тест упал с исключением.
        """
        if self._webdriver_connector is not None:
            self._webdriver_connector.close()
            self._webdriver_connector = None
        self._current_webdriver_port = None

    def _wait_for_bold_button_cdp(self, timeout, log_cb):
        """Пробует подтвердить готовность через CDP-коннектор текущего
        запуска (self._webdriver_connector). Основной триггер — пробуется
        ПЕРЕД win32gui-версией (_wait_for_bold_button) в _wait_until_r7_ready:
        в отличие от неё, реально видит DOM внутри CEF-рендера.

        Любая ошибка (нет соединения, порт не открылся, исключение
        Selenium/websocket) не фатальна — метод просто возвращает False, и
        вызывающий код откатывается на win32gui/CPU-логику.

        Args:
            timeout: Секунд на подключение и опрос кнопки суммарно.
            log_cb: Функция логирования.

        Returns:
            bool: True, если кнопка «Жирный» доступна (найдена и не disabled).
        """
        connector = self._webdriver_connector
        if connector is None:
            # Диагностика: без этой строки в логе неотличимы два разных
            # случая — "коннектор создан, но CDP не ответил" и "коннектор
            # вообще не был создан при запуске" (WEBDRIVER_OK=False, либо
            # все кандидаты портов заняты — см. _prepare_webdriver_launch).
            log_cb(
                f"🔌 WebDriver: CDP-коннектор не создан для этого запуска "
                f"(WEBDRIVER_OK={WEBDRIVER_OK}, "
                f"порт={self._current_webdriver_port}) — пропускаю CDP-триггер"
            )
            return False

        deadline = time.time() + timeout
        try:
            log_cb(f"🔌 WebDriver: попытка подключения к CDP на порту {connector.port}...")
            connect_timeout = max(0.1, min(self.BOLD_BUTTON_CDP_CONNECT_TIMEOUT_SEC, timeout))
            if not connector.connect(timeout=connect_timeout):
                log_cb(f"⚠️ CDP недоступен на порту {connector.port} (порт не открылся "
                       f"или Р7 запущен без --ascdesktop-support-debug-info), использую fallback")
                return False

            cdp_start = time.time()
            while time.time() < deadline:
                state = connector.bold_button_state()
                if state and state.get("found") and not state.get("disabled"):
                    elapsed = time.time() - cdp_start
                    log_cb(f"✅ Кнопка 'Жирный' доступна (CDP, {elapsed:.2f} с)")
                    return True
                time.sleep(self.BOLD_BUTTON_POLL_SEC)

            log_cb("⚠️ CDP: кнопка не стала доступна за отведённое время, использую fallback")
            return False
        except Exception as e:
            log_cb(f"⚠️ CDP недоступен, использую fallback ({type(e).__name__}: {e})")
            return False

    def _wait_until_r7_ready(self, hwnd, timeout=120, log_cb=None):
        """Ждёт, пока Р7-Офис закончит открывать документ.

        Готовность подтверждается двумя независимыми признаками одновременно:
          1. Окно отзывчиво — SendMessageTimeout(WM_NULL) проходит быстрее
             READY_RESPONSIVE_MS.
          2. Процессы Р7 простаивают — суммарный CPU держится ниже
             READY_IDLE_CPU_PCT подряд READY_IDLE_SAMPLES замеров, и при этом
             не запущен конвертер x2t.

        Одного признака мало. Р7 грузит данные в фоновом потоке и остаётся
        отзывчивым во время загрузки — отзывчивость сама по себе сработала бы
        слишком рано. Простой CPU без отзывчивости, наоборот, может совпасть с
        зависшим окном.

        Отдельно проверяется x2t: при открытии .xlsx редактор запускает его
        конвертировать файл, причём уже ПОСЛЕ появления своего окна. Между
        затишьем редактора и стартом конвертера есть пауза, в которую метод
        иначе объявил бы готовность — поэтому список процессов пересобирается
        раз в READY_PROC_REFRESH_SEC, появление нового процесса сбрасывает
        счётчик подтверждения, а живой x2t запрещает вердикт независимо от CPU.

        Опрос идёт с шагом READY_POLL_SEC (0.15 сек) вместо прежних ≈1.6 сек на
        итерацию, а подтверждение занимает ≈3 сек вместо 18 сек, которые уходили
        на BASE_WAIT и окно стабилизации Ctrl+End-зонда.

        Доп. триггер (не чаще раза за вызов): в момент первого совпадения
        признаков 1 и 2 метод пробует, по порядку, ДВА независимых способа
        подтвердить готовность по кнопке «Жирный» на панели инструментов:

          a) _wait_for_bold_button_cdp — через CDP-коннектор текущего
             запуска (r7_webdriver_connector.py), если Р7 был стартован с
             --ascdesktop-support-debug-info (см. _prepare_webdriver_launch).
             Видит кнопку по-настоящему: панель инструментов — DOM внутри
             CEF-рендера, а не набор нативных Win32-виджетов. Бюджет —
             BOLD_BUTTON_TIMEOUT_SEC на подключение и опрос суммарно.
          b) _wait_for_bold_button — win32gui, EnumChildWindows. На
             установленной здесь сборке (2026.2.2.x) экспериментально
             подтверждён НЕработающим (см. его docstring): кнопка реально
             существует как <button id="id-toolbar-btn-bold">, но на 3
             уровня вложенности глубже top-level окна, вне досягаемости
             Win32 API — этот способ не находит окно кнопки и почти не
             стоит времени, оставлен на случай сборки с классическими
             Win32-виджетами на панели.

        Если оба способа не сработали (не найдены/не стали enabled) —
        падаем на обычное накопление READY_IDLE_SAMPLES по CPU и WM_NULL,
        как и раньше. Любая ошибка CDP/Selenium не фатальна для теста —
        только для этого способа подтверждения готовности.

        Ограничение: длительная пауза на вводе-выводе выглядит так же, как
        простой. Пороги вынесены в константы класса — если открытие очень
        больших файлов начнёт определяться преждевременно, поднимать надо
        READY_IDLE_SAMPLES.

        Args:
            hwnd: Окно Р7-Офис для зонда отзывчивости. Либо дескриптор, либо
                функция без аргументов, возвращающая дескриптор — во втором
                случае окно перерешивается, если прежнее перестало существовать.
                None → проверка отзывчивости пропускается.
            timeout: Максимум секунд ожидания.
            log_cb: Функция логирования; по умолчанию self.add_test_log.

        Returns:
            bool: True — готовность подтверждена, False — таймаут или падение Р7.
        """
        if log_cb is None:
            log_cb = self.add_test_log

        # Диагностика CDP-триггера (временная, для разбора несрабатывания —
        # видно сразу, дошло ли вообще до попытки подключения, ещё до того,
        # как base_idle впервые станет True).
        log_cb(
            f"🔌 WebDriver: WEBDRIVER_OK={WEBDRIVER_OK}, "
            f"коннектор={'создан (порт ' + str(self._current_webdriver_port) + ')' if self._webdriver_connector else 'не создан'}"
        )

        start    = time.time()
        deadline = start + timeout

        if not PSUTIL_OK:
            # Без psutil остаётся только отзывчивость окна. Этого мало, чтобы
            # поймать фоновую загрузку, поэтому добавляем короткую фиксированную
            # выдержку и честно пишем об этом в лог.
            log_cb("⚠️ psutil недоступен — готовность определяется только по отзывчивости окна")
            while time.time() < deadline:
                h = hwnd() if callable(hwnd) else hwnd
                if self._window_responsive(h):
                    time.sleep(1.0)
                    return True
                time.sleep(self.READY_POLL_SEC)
            return False

        log_cb("⏳ Ожидание готовности документа (отзывчивость окна + простой CPU)...")

        tracked = {}   # pid -> (psutil.Process с «прогретым» CPU, имя процесса)

        def _adopt():
            """Добавляет в tracked новые процессы Р7. Возвращает их число.

            Первый cpu_percent(None) у процесса задаёт базу отсчёта и всегда
            возвращает 0.0, поэтому он делается здесь, а не в замере.
            Имя запоминается сразу, чтобы не дёргать name() на каждом опросе.
            """
            added = 0
            self._r7_pids = None   # форсируем полное сканирование, чтобы поймать x2t
            for p in self._get_r7_processes(log_cb=log_cb):
                if p.pid in tracked:
                    continue
                try:
                    name = (p.name() or "").lower()
                    p.cpu_percent(None)
                    tracked[p.pid] = (p, name)
                    added += 1
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            return added

        _adopt()
        had_procs         = bool(tracked)
        last_refresh      = time.time()
        idle_streak       = 0
        peak_cpu          = 0.0
        cur_hwnd          = None if callable(hwnd) else hwnd
        bold_button_tried = False   # проба кнопки «Жирный» — не чаще раза за вызов

        while time.time() < deadline:
            time.sleep(self.READY_POLL_SEC)
            now = time.time()

            # Если передана функция поиска окна — перерешиваем hwnd только
            # когда прежний перестал быть окном (Р7 может заменить top-level
            # окно после сплэша). В обычном случае обхода окон не происходит.
            if callable(hwnd):
                if not (cur_hwnd and WIN32_OK and win32gui.IsWindow(cur_hwnd)):
                    cur_hwnd = hwnd()

            # Пересобираем список процессов раз в READY_PROC_REFRESH_SEC: x2t
            # стартует уже после появления окна редактора, и без обновления
            # списка его загрузка осталась бы невидимой.
            if now - last_refresh >= self.READY_PROC_REFRESH_SEC:
                last_refresh = now
                if _adopt():
                    # Появился новый процесс — начинаем подтверждение заново.
                    idle_streak = 0
                    had_procs = True

            total_cpu = 0.0
            converter_alive = False
            dead = []
            for pid, (p, name) in tracked.items():
                try:
                    total_cpu += p.cpu_percent(None)
                    if "x2t" in name:
                        converter_alive = True
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    dead.append(pid)
            for pid in dead:
                tracked.pop(pid, None)
            # Нормировка на число ядер — см. комментарий у READY_IDLE_CPU_PCT
            # (measure_schema 2). Локальный peak_cpu этой функции — только для
            # диагностических строк лога ниже, в JSON-результаты не попадает
            # (там свой peak_cpu, из _sample_r7_resources).
            total_cpu /= self._cpu_count()

            # Процессы Р7 были и исчезли — приложение упало. Ждать до конца
            # таймаута (по умолчанию 120 сек) в этом случае бессмысленно.
            if had_procs and not tracked:
                log_cb("❌ Все процессы Р7-Офис исчезли — приложение завершилось "
                       "или упало во время открытия файла")
                return False

            peak_cpu   = max(peak_cpu, total_cpu)
            responsive = self._window_responsive(cur_hwnd)

            # Пока жив x2t — документ ещё конвертируется, каким бы низким ни
            # был CPU в этот момент (конвертер умеет ждать ввод-вывод).
            base_idle = (responsive and tracked and not converter_alive
                         and total_cpu < self.READY_IDLE_CPU_PCT
                         and now - start >= self.READY_MIN_BUSY_SEC)

            # Доп. триггер — кнопка «Жирный»: пробуем ровно один раз, в момент
            # первого обнаружения простоя (признаки 1 и 2 уже совпали). Если
            # она доступна, объявляем готовность сразу, без обычных ~3 сек
            # накопления READY_IDLE_SAMPLES. Не найдена/недоступна — падаем
            # обратно на старую логику CPU+WM_NULL, дальше уже не пробуем
            # (кнопки может просто не существовать как нативного окна — см.
            # предупреждение в _is_bold_button_visible).
            if base_idle and idle_streak == 0 and not bold_button_tried:
                bold_button_tried = True
                log_cb("⏳ Ожидание кнопки 'Жирный'...")

                # CDP-триггер (см. _wait_for_bold_button_cdp) — пробуется
                # первым: в отличие от win32gui ниже, реально видит кнопку в
                # DOM внутри CEF-рендера. Активен только если Р7 в этом
                # запуске был стартован с debug-флагом (self._webdriver_
                # connector не None — см. _prepare_webdriver_launch);
                # иначе _wait_for_bold_button_cdp возвращает False мгновенно.
                cdp_timeout = max(0.0, min(self.BOLD_BUTTON_TIMEOUT_SEC, deadline - time.time()))
                if self._wait_for_bold_button_cdp(cdp_timeout, log_cb):
                    log_cb(
                        f"   📊 Документ открыт за {time.time() - start:.2f} сек "
                        f"ожидания: кнопка «Жирный» доступна (CDP)")
                    return True

                # Ограничиваем пробу оставшимся бюджетом deadline, а не берём
                # полный BOLD_BUTTON_TIMEOUT_SEC безусловно — иначе вызов с
                # небольшим timeout мог бы превысить его на неучтённые
                # секунды, если простой обнаружился ближе к концу окна.
                btn_timeout = max(0.0, min(self.BOLD_BUTTON_TIMEOUT_SEC, deadline - time.time()))
                if self._wait_for_bold_button(cur_hwnd, timeout=btn_timeout):
                    log_cb("✅ Кнопка 'Жирный' доступна")
                    log_cb(
                        f"   📊 Документ открыт за {time.time() - start:.2f} сек "
                        f"ожидания: кнопка «Жирный» на панели инструментов доступна")
                    return True
                # "Не найдена" и "найдена, но не включилась за отведённое
                # время" — разные диагнозы: в первом случае кнопки как окна
                # нет вообще (см. _is_bold_button_visible), во втором она
                # есть, но fallback-логика всё равно продолжит работу как
                # обычно — сообщение не должно вводить в заблуждение при
                # разборе логов.
                if self._find_bold_button_hwnd(cur_hwnd) is None:
                    log_cb("⚠️ Кнопка 'Жирный' не найдена, использую fallback")
                else:
                    log_cb(
                        f"⚠️ Кнопка 'Жирный' найдена, но не стала доступна за "
                        f"{btn_timeout:.1f} сек, использую fallback")

            if base_idle:
                idle_streak += 1
            else:
                idle_streak = 0

            if idle_streak >= self.READY_IDLE_SAMPLES:
                log_cb(
                    f"   📊 Документ открыт за {now - start:.2f} сек ожидания: "
                    f"CPU процессов Р7 упал до {total_cpu:.1f}% "
                    f"(пик {peak_cpu:.1f}%), окно отзывчиво")
                return True

        log_cb(
            f"⚠️ Таймаут {timeout} сек: готовность не подтверждена "
            f"(пик CPU за ожидание {peak_cpu:.1f}%), продолжаем тест")
        return False

    @staticmethod
    def _json_for_script(obj, **kwargs):
        r"""json.dumps(), но безопасный для вставки прямо внутрь <script>...</script>.

        Строковое значение, содержащее буквальную последовательность
        "</script", закрыло бы окружающий тег раньше времени — HTML-парсер
        браузера не знает, что находится внутри JS-строкового литерала, и
        видит закрывающий тег буквально. Версия/имя теста, попадающие сюда,
        приходят из простых текстов (реестр, simpledialog, JSON-файлы с
        диска), но ничто не мешает им случайно содержать такую подстроку.

        "<\/" — валидный экранированный слэш в JS-строках (не спецсимвол,
        декодируется в тот же "/"), который ломает поиск тега парсером HTML,
        не меняя значение после разбора JSON.
        """
        return json.dumps(obj, **kwargs).replace("</", r"<\/")

    def _generate_html_report(self, results, test_file, open_elapsed,
                              version_str, ram_vals, cpu_vals,
                              peak_ram, avg_ram, min_ram, peak_cpu):
        """Builds and returns the full HTML performance report as a string."""
        ts_display = datetime.now().strftime("%d.%m.%Y %H:%M")
        os_info = platform.platform()
        try:
            cpu_info = platform.processor() or "N/A"
        except Exception:
            cpu_info = "N/A"
        sys_mem_gb = (round(psutil.virtual_memory().total / (1024 ** 3), 1)
                      if PSUTIL_OK else "N/A")
        file_size_mb = (round(test_file.stat().st_size / (1024 * 1024), 2)
                        if test_file.exists() else "N/A")
        cpu_count_display = psutil.cpu_count() if PSUTIL_OK else "N/A"

        # CPU-показатели psutil не нормализованы по числу ядер (могут быть >100%);
        # норм. значение делит их на cpu_count(), получая шкалу 0–100% как в Task Manager.
        cpu_norm_vals = [r.get("cpu_normalized") for r in results if r.get("cpu_normalized") is not None]
        peak_cpu_norm = max(cpu_norm_vals) if cpu_norm_vals else None

        # Chart data
        labels_json   = self._json_for_script([r["name"] for r in results], ensure_ascii=False)
        times_json    = self._json_for_script([round(r["time"], 3) for r in results])
        ram_json      = self._json_for_script([r.get("ram") for r in results])
        cpu_json      = self._json_for_script([r.get("cpu") for r in results])
        cpu_norm_json = self._json_for_script([r.get("cpu_normalized") for r in results])

        # Stats cards
        def stat_card(title, value, unit="", warn=False):
            color = "#e74c3c" if warn else "#2980b9"
            val_str = f"{value:.1f}" if isinstance(value, float) else (str(value) if value is not None else "—")
            return (f'<div class="card" style="border-left:4px solid {color}">'
                    f'<div class="card-title">{title}</div>'
                    f'<div class="card-value" style="color:{color}">{val_str}{unit}</div></div>')

        # Порог предупреждения считаем по нормализованному CPU — «сырое» значение
        # может законно превышать 100% на многоядерной системе и не годится для warn.
        cpu_warn = peak_cpu_norm is not None and peak_cpu_norm > 80
        cards_html = (stat_card("Пик RAM", peak_ram, " МБ") +
                      stat_card("Средн. RAM", avg_ram, " МБ") +
                      stat_card("Мин. RAM", min_ram, " МБ") +
                      stat_card("Пик CPU (сырое)", peak_cpu, "%") +
                      stat_card("Пик CPU (норм.)", peak_cpu_norm, "%", warn=cpu_warn))

        # Results table rows
        rows_html = ""
        for r in results:
            err_class = "row-error" if r.get("error") else ""
            ram_cell = f"{r['ram']:.1f}" if r.get("ram") is not None else "—"
            cpu_cell = f"{r['cpu']:.1f}" if r.get("cpu") is not None else "—"
            cpu_norm_cell = (f"{r['cpu_normalized']:.1f}"
                              if r.get("cpu_normalized") is not None else "—")
            err_cell = html.escape(r.get("error") or "")
            if r.get("runs") and len(r["runs"]) > 1:
                time_cell = (f"{r['avg']:.3f} "
                             f"<span style='color:#888'>({r['min']:.3f}–{r['max']:.3f})</span>")
            else:
                time_cell = f"{r['time']:.3f}"
            # Операция завершилась быстрее, чем детектор успевает заметить
            # занятость Р7 — цифру нельзя сравнивать между версиями.
            if r.get("below_floor"):
                time_cell += (" <span title='Р7-Офис не был занят дольше порога — "
                              "операция быстрее, чем инструмент умеет измерять' "
                              "style='color:#e67e22;font-weight:bold'>&lt;порога</span>")
            rows_html += (f"<tr class='{err_class}'>"
                          f"<td>{r['name']}</td>"
                          f"<td>{time_cell}</td>"
                          f"<td>{ram_cell}</td>"
                          f"<td>{cpu_cell}</td>"
                          f"<td>{cpu_norm_cell}</td>"
                          f"<td>{err_cell}</td></tr>\n")

        html_content = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>R7-Office Performance Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body{{font-family:Arial,sans-serif;margin:0;padding:20px;background:#f5f6fa;color:#333}}
  h1{{color:#2c3e50}}
  .info-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:8px;margin-bottom:20px}}
  .info-item{{background:#fff;padding:8px 12px;border-radius:6px;box-shadow:0 1px 3px rgba(0,0,0,.1)}}
  .info-label{{font-size:.75em;color:#888;text-transform:uppercase}}
  .info-value{{font-weight:bold;margin-top:2px}}
  .cards{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:24px}}
  .card{{background:#fff;padding:14px 18px;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.12);min-width:150px}}
  .card-title{{font-size:.8em;color:#888}}
  .card-value{{font-size:1.6em;font-weight:bold;margin-top:4px}}
  .cpu-note{{font-size:.85em;color:#555;margin:-6px 0 20px;padding:8px 12px;
    background:#eef3fb;border-radius:6px}}
  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:28px}}
  .chart-box{{background:#fff;padding:16px;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.12)}}
  @media(max-width:700px){{.charts{{grid-template-columns:1fr}}}}
  table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.12)}}
  th{{background:#2c3e50;color:#fff;padding:10px 12px;text-align:left;font-size:.85em}}
  td{{padding:8px 12px;border-bottom:1px solid #eee;font-size:.9em}}
  tr:last-child td{{border-bottom:none}}
  tr.row-error td{{background:#fdecea;color:#c0392b}}
  tr:not(.row-error):hover td{{background:#f0f4ff}}
  .pdf-btn{{position:fixed;top:16px;right:16px;padding:8px 18px;background:#2c3e50;
    color:#fff;border:none;border-radius:6px;font-size:.9em;cursor:pointer;
    box-shadow:0 2px 6px rgba(0,0,0,.25);z-index:1000}}
  .pdf-btn:hover{{background:#34495e}}
  @media print{{
    .pdf-btn{{display:none}}
    body{{background:#fff}}
    canvas,.chart-box,table{{page-break-inside:avoid}}
    h1,h2,h3{{page-break-after:avoid}}
  }}
</style>
</head>
<body>
<button class="pdf-btn" onclick="window.print()">📄 Сохранить как PDF</button>
<h1>Отчёт о производительности R7-Office</h1>

<div class="info-grid">
  <div class="info-item"><div class="info-label">Версия R7-Office</div><div class="info-value">{html.escape(version_str) if version_str else "—"}</div></div>
  <div class="info-item"><div class="info-label">Дата и время</div><div class="info-value">{ts_display}</div></div>
  <div class="info-item"><div class="info-label">Тестовый файл</div><div class="info-value">{html.escape(test_file.name)}</div></div>
  <div class="info-item"><div class="info-label">Размер файла</div><div class="info-value">{file_size_mb} МБ</div></div>
  <div class="info-item"><div class="info-label">ОС</div><div class="info-value">{os_info}</div></div>
  <div class="info-item"><div class="info-label">Процессор</div><div class="info-value">{cpu_info}</div></div>
  <div class="info-item"><div class="info-label">RAM (всего)</div><div class="info-value">{sys_mem_gb} ГБ</div></div>
  <div class="info-item"><div class="info-label">Время открытия файла</div><div class="info-value">{open_elapsed:.2f} сек</div></div>
</div>

<div class="cards">{cards_html}</div>
<p class="cpu-note">ℹ️ CPU показан относительно всех ядер (0–100%). «Сырое» значение — как
в диспетчере задач Windows на вкладке «Подробности» (может превышать 100% на многоядерных
системах), «норм.» — то же значение, делённое на количество логических ядер
({cpu_count_display}).</p>

<div class="charts">
  <div class="chart-box"><canvas id="timeChart"></canvas></div>
  <div class="chart-box"><canvas id="ramChart"></canvas></div>
  <div class="chart-box"><canvas id="cpuChart"></canvas></div>
</div>

<table>
<thead><tr><th>Операция</th><th>Время (сек)</th><th>RAM (МБ)</th><th>CPU (%)</th><th>CPU норм. (%)</th><th>Ошибка</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>

<script>
const labels   = {labels_json};
const times    = {times_json};
const rams     = {ram_json};
const cpus     = {cpu_json};
const cpusNorm = {cpu_norm_json};
const defOpts = (title) => ({{
  responsive: true,
  plugins: {{legend:{{display:false}}, title:{{display:true, text:title}}}},
  scales: {{y:{{beginAtZero:true}}}}
}});
new Chart(document.getElementById('timeChart'), {{
  type:'bar', data:{{labels, datasets:[{{label:'сек',data:times,backgroundColor:'#3498db'}}]}},
  options: defOpts('Время выполнения (сек)')
}});
new Chart(document.getElementById('ramChart'), {{
  type:'line', data:{{labels, datasets:[{{label:'МБ',data:rams,borderColor:'#27ae60',backgroundColor:'rgba(39,174,96,.15)',fill:true,tension:.3}}]}},
  options: defOpts('Потребление RAM (МБ)')
}});
new Chart(document.getElementById('cpuChart'), {{
  type:'line',
  data:{{labels, datasets:[
    {{label:'CPU сырое (%)', data:cpus, borderColor:'#e67e22', backgroundColor:'rgba(230,126,34,.12)', fill:false, tension:.3}},
    {{label:'CPU норм. (%)', data:cpusNorm, borderColor:'#8e44ad', backgroundColor:'rgba(142,68,173,.15)', fill:true, tension:.3}}
  ]}},
  options: {{
    responsive: true,
    plugins: {{legend:{{display:true}}, title:{{display:true, text:'Нагрузка на CPU (%)'}}}},
    scales: {{y:{{beginAtZero:true}}}}
  }}
}});
</script>
</body>
</html>"""
        return html_content

    # ---------------------- Диалог после теста ----------------------

    def _show_post_test_dialog(self, html_path, ts):
        """Shows dialog after test completion: open report, new test, or exit."""
        dlg = tk.Toplevel(self.root)
        dlg.transient(self.root)
        dlg.configure(bg=COLORS["bg"])
        dlg.title("Тест завершён")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.focus_set()

        ttk.Label(dlg, text="Тест завершён!", font=("Arial", 12, "bold")).pack(
            pady=(24, 6), padx=40)
        ttk.Label(dlg, text="Что делать дальше?", foreground=COLORS["text_secondary"]).pack(pady=(0, 20))

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(pady=(0, 24), padx=40)

        def show_report():
            webbrowser.open(str(html_path))
            dlg.destroy()
            if messagebox.askyesno("Сохранить копию", "Сохранить копию HTML-отчёта?"):
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".html",
                    filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
                    initialfile=f"Performance_Report_{ts}.html"
                )
                if save_path:
                    shutil.copy(str(html_path), save_path)
                    self.add_test_log(f"📎 Копия отчёта сохранена: {save_path}")

        def new_test():
            dlg.destroy()
            self._reset_test_state()

        def exit_app():
            dlg.destroy()
            self.root.quit()

        ttk.Button(btn_frame, text="📊 Показать отчёт", command=show_report, width=20
                   ).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🔄 Новый тест", command=new_test, width=14
                   ).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Выход", command=exit_app, width=10
                   ).pack(side=tk.LEFT, padx=5)

        dlg.update_idletasks()
        w = dlg.winfo_reqwidth()
        h = dlg.winfo_reqheight()
        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

    def _reset_test_state(self):
        """Clears the test log and resets the status bar for a new run."""
        self.test_log.delete("1.0", tk.END)
        self.status_var.set("Готов")
        self.add_test_log("🔄 Готов к новому тесту.")

    # ---------------------- Сравнение версий ----------------------

    def _load_comparison_settings(self):
        path = BASE_DIR / "last_comparison_settings.json"
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {"custom_names": {}, "last_selected_files": [], "last_base_version": ""}

    def _save_comparison_settings(self, settings):
        path = BASE_DIR / "last_comparison_settings.json"
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(settings, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def compare_versions(self):
        """Opens dialog to select 2-10 performance JSON files and builds a comparison report."""
        MAX_FILES = 10
        CHART_COLORS = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                        '#1abc9c', '#e67e22', '#c0392b', '#16a085', '#f1c40f']

        settings = self._load_comparison_settings()
        custom_names = settings.get("custom_names", {})
        last_selected = set(settings.get("last_selected_files", []))
        last_base = settings.get("last_base_version", "")

        def scan_files():
            json_files = sorted(
                self.reports_folder.glob("performance_full_*.json"),
                key=lambda fp: fp.stat().st_mtime, reverse=True
            )
            result = []
            for jf in json_files:
                key = str(jf)
                try:
                    with open(jf, encoding="utf-8") as fh:
                        jdata = json.load(fh)
                    version = jdata.get("version") or jf.stem
                    ts_raw = jdata.get("timestamp", "")
                    ts_disp = (f"{ts_raw[6:8]}.{ts_raw[4:6]}.{ts_raw[:4]} "
                               f"{ts_raw[9:11]}:{ts_raw[11:13]}"
                               if len(ts_raw) >= 13 else ts_raw)
                except Exception:
                    jdata = None
                    version = jf.stem
                    ts_disp = ""
                result.append({
                    "path": jf, "key": key, "version": version,
                    "ts": ts_disp, "data": jdata,
                    "display_name": custom_names.get(key, version),
                })
            return result

        initial_meta = scan_files()
        if len(initial_meta) < 2:
            self.add_test_log(
                f"⚠️ Сравнение версий: найдено {len(initial_meta)} файлов "
                f"performance_full_*.json (нужно минимум 2)")
            messagebox.showwarning(
                "Недостаточно данных",
                "Для сравнения нужно минимум 2 файла performance_full_*.json.\n"
                "Запустите тесты для нескольких версий R7-Office."
            )
            return

        # Mutable state shared by all closures
        file_meta_by_key = {}   # key -> meta dict
        sel_vars = {}           # key -> BooleanVar
        combo_keys_ref = []     # ordered list of keys matching combo values

        # ── Dialog ──────────────────────────────────────────────────────────
        try:
            dlg = tk.Toplevel(self.root)
            dlg.transient(self.root)
            dlg.configure(bg=COLORS["bg"])
            dlg.title("Сравнение версий")
            dlg.resizable(True, True)
            dlg.minsize(580, 400)
            dlg.grab_set()

            ttk.Label(dlg, text="Выберите 2–10 файлов для сравнения:",
                      font=("Arial", 10, "bold")).pack(pady=(12, 4), padx=14, anchor=tk.W)

            # ── Scrollable list ──────────────────────────────────────────────────
            list_outer = ttk.LabelFrame(dlg, text="Доступные результаты", padding="4")
            list_outer.pack(fill=tk.BOTH, expand=True, padx=14, pady=4)

            list_canvas = tk.Canvas(list_outer, highlightthickness=0)
            vsb = ttk.Scrollbar(list_outer, orient=tk.VERTICAL, command=list_canvas.yview)
            list_canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side=tk.RIGHT, fill=tk.Y)
            list_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            inner = ttk.Frame(list_canvas)
            inner_id = list_canvas.create_window((0, 0), window=inner, anchor="nw")

            def _on_inner_cfg(e):
                list_canvas.configure(scrollregion=list_canvas.bbox("all"))
            inner.bind("<Configure>", _on_inner_cfg)

            def _on_canvas_cfg(e):
                list_canvas.itemconfig(inner_id, width=e.width)
            list_canvas.bind("<Configure>", _on_canvas_cfg)

            def _on_mwheel(e):
                list_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
            list_canvas.bind_all("<MouseWheel>", _on_mwheel)

            # ── Row builder ─────────────────────────────────────────────────────
            def build_row(meta, idx):
                key = meta["key"]
                var = tk.BooleanVar(value=(key in last_selected))
                sel_vars[key] = var
                color = CHART_COLORS[idx % len(CHART_COLORS)]

                rf = ttk.Frame(inner)
                rf.pack(fill=tk.X, pady=1, padx=2)

                ttk.Checkbutton(rf, variable=var).pack(side=tk.LEFT)

                dot = tk.Canvas(rf, width=14, height=14, highlightthickness=0,
                                bg=dlg.cget("bg"))
                dot.create_oval(2, 2, 12, 12, fill=color, outline="")
                dot.pack(side=tk.LEFT, padx=(2, 4))

                ts_val = meta.get("ts", "")
                name_txt = meta.get("display_name", meta["version"])
                lbl_txt = f"{name_txt}  •  {ts_val}" if ts_val else name_txt
                lbl = ttk.Label(rf, text=lbl_txt, anchor=tk.W)
                lbl.pack(side=tk.LEFT, padx=(0, 6), fill=tk.X, expand=True)

                def make_rename(m, lb):
                    def do_rename():
                        new_name = simpledialog.askstring(
                            "Переименовать", "Новое название:",
                            initialvalue=m.get("display_name", m["version"]),
                            parent=dlg
                        )
                        if new_name and new_name.strip():
                            m["display_name"] = new_name.strip()
                            custom_names[m["key"]] = new_name.strip()
                            ts = m.get("ts", "")
                            lb.config(text=f"{new_name.strip()}  •  {ts}" if ts
                                      else new_name.strip())
                            refresh_base_combo()
                    return do_rename

                def make_delete(m, row_frame):
                    def do_delete():
                        file_meta_by_key.pop(m["key"], None)
                        sel_vars.pop(m["key"], None)
                        custom_names.pop(m["key"], None)
                        row_frame.destroy()
                        refresh_base_combo()
                    return do_delete

                btn_ren = ttk.Button(rf, text="✏️", width=3,
                                     command=make_rename(meta, lbl))
                btn_ren.pack(side=tk.RIGHT, padx=1)
                btn_del = ttk.Button(rf, text="🗑️", width=3,
                                     command=make_delete(meta, rf))
                btn_del.pack(side=tk.RIGHT, padx=1)

                ctx = tk.Menu(dlg, tearoff=0)

                def make_ctx_handler(m, lb, row_frame):
                    def show(e):
                        try:
                            ctx.delete(0, tk.END)
                            ctx.add_command(label="✏️ Переименовать",
                                            command=make_rename(m, lb))
                            ctx.add_command(label="🗑️ Удалить из списка",
                                            command=make_delete(m, row_frame))
                            ctx.add_separator()
                            ctx.add_command(label="📌 Сделать базовой",
                                            command=lambda: _set_base_by_key(m["key"]))
                            ctx.tk_popup(e.x_root, e.y_root)
                        finally:
                            ctx.grab_release()
                    return show

                show_ctx = make_ctx_handler(meta, lbl, rf)
                rf.bind("<Button-3>", show_ctx)
                lbl.bind("<Button-3>", show_ctx)

            # ── Populate initial rows ────────────────────────────────────────────
            for i, m in enumerate(initial_meta[:MAX_FILES]):
                file_meta_by_key[m["key"]] = m
                build_row(m, i)

            # ── Toolbar ─────────────────────────────────────────────────────────
            toolbar = ttk.Frame(dlg)
            toolbar.pack(fill=tk.X, padx=14, pady=(4, 0))

            def add_file():
                if len(file_meta_by_key) >= MAX_FILES:
                    messagebox.showwarning("Лимит",
                                           f"Максимум {MAX_FILES} файлов.", parent=dlg)
                    return
                path_str = filedialog.askopenfilename(
                    parent=dlg,
                    title="Выбрать JSON-файл результатов",
                    filetypes=[("JSON файлы", "*.json"), ("Все файлы", "*.*")],
                    initialdir=str(self.reports_folder)
                )
                if not path_str:
                    return
                from pathlib import Path as _Path
                jf = _Path(path_str)
                key = str(jf)
                if key in file_meta_by_key:
                    messagebox.showinfo("Уже добавлен",
                                        "Этот файл уже есть в списке.", parent=dlg)
                    return
                try:
                    with open(jf, encoding="utf-8") as fh:
                        jdata = json.load(fh)
                    version = jdata.get("version") or jf.stem
                    ts_raw = jdata.get("timestamp", "")
                    ts_disp = (f"{ts_raw[6:8]}.{ts_raw[4:6]}.{ts_raw[:4]} "
                               f"{ts_raw[9:11]}:{ts_raw[11:13]}"
                               if len(ts_raw) >= 13 else ts_raw)
                except Exception as ex:
                    messagebox.showerror("Ошибка",
                                         f"Не удалось прочитать файл:\n{ex}", parent=dlg)
                    return
                meta = {
                    "path": jf, "key": key, "version": version,
                    "ts": ts_disp, "data": jdata,
                    "display_name": custom_names.get(key, version),
                }
                idx = len(file_meta_by_key)
                file_meta_by_key[key] = meta
                build_row(meta, idx)
                refresh_base_combo()

            def refresh_list():
                new_meta = scan_files()
                added = 0
                for m in new_meta:
                    if m["key"] not in file_meta_by_key:
                        if len(file_meta_by_key) >= MAX_FILES:
                            break
                        idx = len(file_meta_by_key)
                        file_meta_by_key[m["key"]] = m
                        build_row(m, idx)
                        added += 1
                if added:
                    refresh_base_combo()
                    messagebox.showinfo("Обновлено",
                                        f"Добавлено новых файлов: {added}", parent=dlg)
                else:
                    messagebox.showinfo("Нет изменений",
                                        "Новых файлов не найдено.", parent=dlg)

            ttk.Button(toolbar, text="➕ Добавить файл",
                       command=add_file).pack(side=tk.LEFT, padx=(0, 6))
            ttk.Button(toolbar, text="🔄 Обновить список",
                       command=refresh_list).pack(side=tk.LEFT)

            ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=14, pady=8)

            # ── Base version selector ────────────────────────────────────────────
            base_frame = ttk.LabelFrame(dlg, text="Базовая версия (для расчёта Δ%)", padding="6")
            base_frame.pack(fill=tk.X, padx=14, pady=4)

            base_var = tk.StringVar()
            base_combo = ttk.Combobox(base_frame, textvariable=base_var,
                                      state="readonly", width=60)
            base_combo.pack(fill=tk.X, padx=4, pady=2)

            def refresh_base_combo():
                prev_key = (combo_keys_ref[base_combo.current()]
                            if combo_keys_ref and 0 <= base_combo.current() < len(combo_keys_ref)
                            else "")
                keys = list(file_meta_by_key.keys())
                combo_keys_ref.clear()
                combo_keys_ref.extend(keys)
                values = []
                for k in keys:
                    m = file_meta_by_key[k]
                    nm = m.get("display_name", m["version"])
                    ts = m.get("ts", "")
                    values.append(f"{nm}  •  {ts}" if ts else nm)
                base_combo["values"] = values
                if prev_key and prev_key in keys:
                    base_combo.current(keys.index(prev_key))
                elif last_base and last_base in keys:
                    base_combo.current(keys.index(last_base))
                elif keys:
                    base_combo.current(0)

            def _set_base_by_key(key):
                if key in combo_keys_ref:
                    base_combo.current(combo_keys_ref.index(key))

            refresh_base_combo()

            # ── Action buttons ───────────────────────────────────────────────────
            btn_frame = ttk.Frame(dlg)
            btn_frame.pack(pady=10, padx=14, fill=tk.X)

            def _cleanup():
                list_canvas.unbind_all("<MouseWheel>")
                dlg.destroy()

            def do_compare():
                selected_keys = [k for k, v in sel_vars.items() if v.get()]
                if len(selected_keys) < 2:
                    messagebox.showwarning("Мало файлов",
                                           "Выберите минимум 2 файла.", parent=dlg)
                    return
                if len(selected_keys) > MAX_FILES:
                    messagebox.showwarning("Много файлов",
                                           f"Выберите не более {MAX_FILES} файлов.", parent=dlg)
                    return
                cidx = base_combo.current()
                if cidx < 0 or cidx >= len(combo_keys_ref):
                    messagebox.showwarning("Базовая версия",
                                           "Выберите базовую версию.", parent=dlg)
                    return
                base_key = combo_keys_ref[cidx]
                if base_key not in selected_keys:
                    messagebox.showwarning(
                        "Базовая версия",
                        "Базовая версия должна быть среди выбранных файлов.", parent=dlg)
                    return

                datasets = []
                for k in selected_keys:
                    m = file_meta_by_key[k]
                    jdata = m.get("data")
                    if jdata is None:
                        try:
                            with open(m["path"], encoding="utf-8") as fh:
                                jdata = json.load(fh)
                        except Exception as ex:
                            messagebox.showerror(
                                "Ошибка",
                                f"Не удалось загрузить {m['path'].name}:\n{ex}",
                                parent=dlg)
                            return
                    datasets.append({
                        "path": str(m["path"]),
                        "version": m.get("display_name", m["version"]),
                        "data": jdata,
                    })

                self._save_comparison_settings({
                    "custom_names": custom_names,
                    "last_selected_files": selected_keys,
                    "last_base_version": base_key,
                })
                _cleanup()
                html = self._generate_comparison_html(
                    datasets, str(file_meta_by_key[base_key]["path"]))
                ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_path = self.reports_folder / f"comparison_{ts_now}.html"
                try:
                    out_path.write_text(html, encoding="utf-8")
                    self.add_test_log(f"📊 Отчёт сравнения сохранён: {out_path.name}")
                    webbrowser.open(str(out_path))
                except Exception as ex:
                    messagebox.showerror("Ошибка", f"Не удалось сохранить отчёт:\n{ex}")

            ttk.Button(btn_frame, text="📊 Сравнить",
                       command=do_compare).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="Отмена",
                       command=_cleanup).pack(side=tk.LEFT)

            dlg.protocol("WM_DELETE_WINDOW", _cleanup)

            dlg.update_idletasks()
            row_h = max(len(file_meta_by_key) * 34 + 20, 80)
            list_canvas.configure(height=min(row_h, 220))
            w = max(600, dlg.winfo_reqwidth())
            h = min(700, max(440, dlg.winfo_reqheight()))
            sx = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
            sy = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
            dlg.geometry(f"{w}x{h}+{sx}+{sy}")
        except Exception as ex:
            self.add_test_log(f"❌ Ошибка при построении окна сравнения версий: {ex}")
            try:
                list_canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass
            try:
                dlg.destroy()
            except Exception:
                pass
            messagebox.showerror("Ошибка", f"Не удалось открыть окно сравнения версий:\n{ex}")

    # ── Страница трендов (этап 2, M5) ───────────────────────────────────────
    TRENDS_CHART_COLORS = ('#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                          '#1abc9c', '#e67e22', '#c0392b', '#16a085', '#f1c40f')

    def show_trends(self):
        """Строит и открывает в браузере страницу трендов по всем
        накопленным performance_full_*.json. Точка входа из UI (кнопка
        «📈 Тренды» рядом с «Сравнить версии»)."""
        runs = self._load_trends_runs()
        if len(runs) < 2:
            messagebox.showinfo(
                "Недостаточно данных",
                f"Найдено {len(runs)} файлов performance_full_*.json "
                f"(нужно минимум 2 для тренда).\nЗапустите тесты несколько раз.")
            return
        html_content = self._generate_trends_html(runs)
        ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = self.reports_folder / f"trends_{ts_now}.html"
        try:
            out_path.write_text(html_content, encoding="utf-8")
            self.add_test_log(f"📈 Страница трендов: {out_path.name}")
            webbrowser.open(str(out_path))
        except Exception as e:
            self.add_test_log(f"⚠️ Ошибка сохранения страницы трендов: {e}")
            messagebox.showerror("Ошибка", f"Не удалось сохранить страницу трендов:\n{e}")

    def _load_trends_runs(self):
        """Читает все performance_full_*.json из reports_folder в
        хронологическом порядке (по времени модификации файла — timestamp
        внутри JSON тот же по построению, mtime надёжнее при ручном
        переименовании файлов).

        Файлы, которые не удалось разобрать (битый JSON, обрезанный прогон),
        пропускаются молча — один повреждённый файл не должен ронять всю
        страницу трендов, накопленную за недели прогонов.

        Returns:
            list[dict]: [{"path", "ts_raw", "ts_disp", "version", "schema",
            "results": {имя_операции: dict-результат}}, ...], отсортировано
            по времени.
        """
        files = sorted(self.reports_folder.glob("performance_full_*.json"),
                       key=lambda p: p.stat().st_mtime)
        runs = []
        for fp in files:
            try:
                with open(fp, encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                continue
            ts_raw = data.get("timestamp", "")
            ts_disp = (f"{ts_raw[6:8]}.{ts_raw[4:6]}.{ts_raw[:4]} "
                      f"{ts_raw[9:11]}:{ts_raw[11:13]}"
                      if len(ts_raw) >= 13 else (ts_raw or fp.stem))
            runs.append({
                "path": fp,
                "ts_raw": ts_raw,
                "ts_disp": ts_disp,
                "version": data.get("version") or fp.stem,
                "schema": data.get("measure_schema", 1),
                "results": {r["name"]: r for r in data.get("results", [])
                           if isinstance(r, dict) and "name" in r},
            })
        return runs

    def _generate_trends_html(self, runs):
        """Строит HTML-страницу трендов из уже загруженных прогонов (см.
        _load_trends_runs) — по одному графику Chart.js на операцию: время
        (results[op]["time"] — медиана для measure_schema=2, среднее для
        версии 1, см. этап 1) на оси Y, дата прогона на оси X, полоса MAD
        вокруг линии (только там, где MAD посчитан — measure_schema=2),
        точки раскрашены по версии Р7.

        Разделено на _load_trends_runs (чтение с диска) + эта функция
        (чистое построение HTML из данных) специально для тестируемости —
        сама генерация не должна требовать реальных файлов на диске.

        Args:
            runs: Список прогонов в формате _load_trends_runs (минимум 2 —
                проверяется вызывающим кодом, show_trends).

        Returns:
            str: готовый HTML.
        """
        op_names = []
        seen = set()
        for run in runs:
            for name in run["results"]:
                if name not in seen:
                    op_names.append(name)
                    seen.add(name)

        versions = []
        for run in runs:
            if run["version"] not in versions:
                versions.append(run["version"])
        version_color = {v: self.TRENDS_CHART_COLORS[i % len(self.TRENDS_CHART_COLORS)]
                         for i, v in enumerate(versions)}

        # Тот же предупреждающий баннер, что и в _generate_comparison_html:
        # measure_schema различает среднее (v1) и медиану (v2) — на одном
        # графике времени они несопоставимы, даже если обе подписаны "time".
        schema_versions = {run["schema"] for run in runs}
        schema_warning_html = ""
        if len(schema_versions) > 1:
            schema_warning_html = (
                '<div style="background:#4a2a1a;border:1px solid #d68910;'
                'border-radius:6px;padding:10px 16px;margin:0 0 16px;color:#f5cba7">'
                '⚠️ В истории смешаны файлы разных версий схемы замера '
                f'({", ".join(str(v) for v in sorted(schema_versions))}) — '
                'до 25.08.2026 «время» считалось как среднее по сырым '
                'CPU-порогам, после — как медиана по нормированным. Излом '
                'линии на границе версий схемы может отражать смену метода '
                'замера, а не реальное изменение производительности.</div>\n'
            )

        legend_items = "".join(
            f'<div class="legend-item"><span class="legend-dot" '
            f'style="background:{version_color[v]}"></span>'
            f'<span>{html.escape(v)}</span></div>\n'
            for v in versions
        )

        charts_html = ""
        charts_js = ""
        for idx, op in enumerate(op_names):
            labels, values, mad_lower, mad_upper, point_colors = [], [], [], [], []
            has_mad = False
            for run in runs:
                r = run["results"].get(op)
                if r is None or r.get("time") is None:
                    continue
                v = r["time"]
                labels.append(run["ts_disp"])
                values.append(round(v, 3))
                mad = r.get("mad")
                if mad is not None:
                    mad_lower.append(round(v - mad, 3))
                    mad_upper.append(round(v + mad, 3))
                    has_mad = True
                else:
                    mad_lower.append(None)
                    mad_upper.append(None)
                point_colors.append(version_color.get(run["version"], "#888"))

            if len(values) < 2:
                continue  # график из одной точки бесполезен для тренда

            canvas_id = f"trend{idx}"
            datasets = []
            if has_mad:
                datasets.append({
                    "label": "MAD −", "data": mad_lower, "borderWidth": 0,
                    "pointRadius": 0, "fill": False, "tension": 0.15,
                    "spanGaps": True,
                })
                datasets.append({
                    "label": "MAD-полоса", "data": mad_upper, "borderWidth": 0,
                    "pointRadius": 0, "backgroundColor": "rgba(52,152,219,.15)",
                    "fill": "-1", "tension": 0.15, "spanGaps": True,
                })
            datasets.append({
                "label": op, "data": values, "borderColor": "#2c3e50",
                "borderWidth": 2, "backgroundColor": point_colors,
                "pointBackgroundColor": point_colors, "pointRadius": 4,
                "pointHoverRadius": 6, "tension": 0.15, "fill": False,
            })

            charts_html += (
                f'<div class="chart-box"><h3>{html.escape(op)}</h3>'
                f'<canvas id="{canvas_id}"></canvas></div>\n'
            )
            charts_js += f"""
new Chart(document.getElementById({json.dumps(canvas_id)}), {{
  type: 'line',
  data: {{ labels: {self._json_for_script(labels)},
           datasets: {self._json_for_script(datasets)} }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{ y: {{ beginAtZero: false,
                     title: {{ display: true, text: 'секунды' }} }},
               x: {{ ticks: {{ maxRotation: 45, minRotation: 45 }} }} }}
  }}
}});
"""

        if not charts_html:
            charts_html = ('<p style="color:#888">Ни одна операция не '
                          'встретилась хотя бы в двух прогонах — трендов '
                          'строить не из чего.</p>')

        ts_display = datetime.now().strftime("%d.%m.%Y %H:%M")
        return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Тренды производительности R7-Office</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body{{font-family:Arial,sans-serif;margin:0;padding:20px;background:#f5f6fa;color:#333}}
  h1{{color:#2c3e50;margin-bottom:2px}}
  h3{{color:#2c3e50;margin:0 0 10px}}
  .subtitle{{color:#888;font-size:.9em;margin-bottom:14px}}
  .legend{{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:20px}}
  .legend-item{{display:flex;align-items:center;gap:6px;font-size:.88em}}
  .legend-dot{{width:13px;height:13px;border-radius:50%;flex-shrink:0}}
  .charts{{display:grid;grid-template-columns:1fr;gap:18px;margin-bottom:28px}}
  .chart-box{{background:#fff;padding:16px;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.12)}}
  .pdf-btn{{position:fixed;top:16px;right:16px;padding:8px 18px;background:#2c3e50;
    color:#fff;border:none;border-radius:6px;font-size:.9em;cursor:pointer;
    box-shadow:0 2px 6px rgba(0,0,0,.25);z-index:1000}}
  .pdf-btn:hover{{background:#34495e}}
  @media print{{
    .pdf-btn{{display:none}}
    body{{background:#fff}}
    canvas,.chart-box{{page-break-inside:avoid}}
    h1,h2,h3{{page-break-after:avoid}}
  }}
</style>
</head>
<body>
<button class="pdf-btn" onclick="window.print()">📄 Сохранить как PDF</button>
<h1>Тренды производительности R7-Office</h1>
<div class="subtitle">Сформировано: {ts_display} &nbsp;|&nbsp; Прогонов: {len(runs)}</div>
{schema_warning_html}<div class="legend">
{legend_items}</div>
<div class="charts">
{charts_html}</div>
<script>
{charts_js}
</script>
</body>
</html>"""

    def _generate_comparison_html(self, datasets, base_path_str):
        """Builds comparison HTML for 2-10 performance datasets.

        Args:
            datasets: list of dicts {path: str, version: str, data: dict}
            base_path_str: path string of the dataset used as baseline
        """
        CHART_COLORS = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                  '#1abc9c', '#e67e22', '#c0392b', '#16a085', '#f1c40f']
        ALPHA  = ["rgba(52,152,219,.15)", "rgba(231,76,60,.15)", "rgba(46,204,113,.15)",
                  "rgba(243,156,18,.15)", "rgba(155,89,182,.15)", "rgba(26,188,156,.15)",
                  "rgba(230,126,34,.15)", "rgba(192,57,43,.15)", "rgba(22,160,133,.15)",
                  "rgba(241,196,15,.15)"]

        # Collect operation names in the order they first appear
        seen, op_names = set(), []
        for ds in datasets:
            for r in ds["data"].get("results", []):
                if r["name"] not in seen:
                    op_names.append(r["name"])
                    seen.add(r["name"])

        for ds in datasets:
            ds["lookup"] = {r["name"]: r for r in ds["data"].get("results", [])}

        base_ds = next(ds for ds in datasets if ds["path"] == base_path_str)
        legend_pos = "right" if len(datasets) >= 6 else "top"

        # Chart datasets
        time_ds, ram_ds, cpu_ds = [], [], []
        for i, ds in enumerate(datasets):
            is_base = ds["path"] == base_path_str
            lbl = ds["version"] + (" (база)" if is_base else "")
            lk  = ds["lookup"]
            time_ds.append({
                "label": lbl,
                "data": [round(lk[op]["time"], 3) if op in lk else None for op in op_names],
                "backgroundColor": CHART_COLORS[i % len(CHART_COLORS)],
                "borderRadius": 3,
            })
            ram_ds.append({
                "label": lbl,
                "data": [lk[op].get("ram") if op in lk else None for op in op_names],
                "borderColor": CHART_COLORS[i % len(CHART_COLORS)],
                "backgroundColor": ALPHA[i % len(ALPHA)],
                "tension": 0.3, "fill": True,
            })
            cpu_ds.append({
                "label": lbl,
                "data": [lk[op].get("cpu") if op in lk else None for op in op_names],
                "borderColor": CHART_COLORS[i % len(CHART_COLORS)],
                "backgroundColor": ALPHA[i % len(ALPHA)],
                "tension": 0.3, "fill": False,
            })

        labels_json  = self._json_for_script(op_names, ensure_ascii=False)
        time_ds_json = self._json_for_script(time_ds,  ensure_ascii=False)
        ram_ds_json  = self._json_for_script(ram_ds,   ensure_ascii=False)
        cpu_ds_json  = self._json_for_script(cpu_ds,   ensure_ascii=False)

        # Table header (two rows)
        th1 = "<tr><th rowspan='2'>Операция</th>"
        th2 = "<tr>"
        for i, ds in enumerate(datasets):
            is_base = ds["path"] == base_path_str
            suffix  = " <em>(база)</em>" if is_base else ""
            color   = CHART_COLORS[i % len(CHART_COLORS)]
            th1 += f'<th colspan="3" style="background:{color}">{html.escape(ds["version"])}{suffix}</th>'
            th2 += (f'<th style="background:{color}">Время (сек)</th>'
                    f'<th style="background:{color}">Δ%</th>'
                    f'<th style="background:{color}">Вердикт (M5)</th>')
        th1 += "</tr>"
        th2 += "</tr>"

        # Table rows
        def delta_td(t, base_t, is_base):
            if is_base or base_t is None:
                return "<td class='delta-base'>—</td>"
            if t is None:
                return "<td>—</td>"
            pct = (t - base_t) / base_t * 100
            if abs(pct) <= 5:
                return f"<td class='delta-same'>{'+'if pct>0 else ''}{pct:.1f}%</td>"
            if pct < 0:
                return f"<td class='delta-better'>{pct:.1f}%</td>"
            return f"<td class='delta-worse'>+{pct:.1f}%</td>"

        # Вердикт M5 (compare_runs): нужны "сырые" повторы (results[...]["runs"])
        # с обеих сторон — их пишет только одиночный прогон вкладки
        # «Производительность» (run_test_with_runs), Batch-режим делает один
        # замер на операцию и "runs" не сохраняет вовсе. Меньше
        # MIN_RUNS_FOR_COMPARISON повторов — критерий Манна-Уитни статистически
        # ненадёжен (см. compare_runs), вердикт не выносится.
        VERDICT_CLASS = {"РЕГРЕССИЯ": "delta-worse", "УСКОРЕНИЕ": "delta-better",
                         "без изменений": "delta-same"}

        def verdict_td(base_r, r, is_base):
            if is_base:
                return "<td class='delta-base'>—</td>"
            base_runs = (base_r or {}).get("runs") or []
            new_runs = (r or {}).get("runs") or []
            if len(base_runs) < MIN_RUNS_FOR_COMPARISON or len(new_runs) < MIN_RUNS_FOR_COMPARISON:
                return (f"<td class='delta-base' title='Нужно минимум "
                       f"{MIN_RUNS_FOR_COMPARISON} повторов на каждую версию — "
                       f"есть {len(base_runs)} и {len(new_runs)}. Данные Batch-режима "
                       f"этого не хранят'>—</td>")
            result = compare_runs(base_runs, new_runs)
            cls = VERDICT_CLASS.get(result["verdict"], "")
            title = (f"p={result['p_value']}, эффект {result['effect_pct']:+.1f}%, "
                    f"n={result['n_base']}/{result['n_new']}")
            return f"<td class='{cls}' title='{html.escape(title)}'>{result['verdict']}</td>"

        table_rows = ""
        for op in op_names:
            base_r = base_ds["lookup"].get(op)
            base_t = base_r["time"] if base_r else None
            row = f"<tr><td>{html.escape(op)}</td>"
            for ds in datasets:
                r = ds["lookup"].get(op)
                t = r["time"] if r else None
                is_base = ds["path"] == base_path_str
                row += f"<td>{'—' if t is None else f'{t:.3f}'}</td>"
                row += delta_td(t, base_t, is_base)
                row += verdict_td(base_r, r, is_base)
            row += "</tr>"
            table_rows += row + "\n"

        # System info cards
        sys_cards = ""
        for i, ds in enumerate(datasets):
            color    = CHART_COLORS[i % len(CHART_COLORS)]
            is_base  = ds["path"] == base_path_str
            sys_info = ds["data"].get("system", {})
            summ     = ds["data"].get("summary", {})
            ts_raw   = ds["data"].get("timestamp", "")
            base_lbl = " <strong>(база)</strong>" if is_base else ""
            pr  = summ.get("peak_ram_mb")
            ar  = summ.get("avg_ram_mb")
            pc  = summ.get("peak_cpu_pct")
            sys_cards += (
                f'<div class="sys-card" style="border-left:4px solid {color}">'
                f'<div class="sys-title" style="color:{color}">'
                f'{html.escape(ds["version"])}{base_lbl}</div>'
                f'<div class="sys-row"><span class="sys-lbl">ОС</span>'
                f'<span>{sys_info.get("os","—")}</span></div>'
                f'<div class="sys-row"><span class="sys-lbl">RAM</span>'
                f'<span>{sys_info.get("ram_total_gb","—")} ГБ</span></div>'
                f'<div class="sys-row"><span class="sys-lbl">Пик RAM</span>'
                f'<span>{"—" if pr is None else f"{pr:.1f} МБ"}</span></div>'
                f'<div class="sys-row"><span class="sys-lbl">Средн. RAM</span>'
                f'<span>{"—" if ar is None else f"{ar:.1f} МБ"}</span></div>'
                f'<div class="sys-row"><span class="sys-lbl">Пик CPU</span>'
                f'<span>{"—" if pc is None else f"{pc:.1f}%"}</span></div>'
                f'<div class="sys-row"><span class="sys-lbl">Дата теста</span>'
                f'<span>{ts_raw}</span></div>'
                f'</div>'
            )

        # Legend
        legend_items = ""
        for i, ds in enumerate(datasets):
            is_base = ds["path"] == base_path_str
            suffix  = " (база)" if is_base else ""
            legend_items += (
                f'<div class="legend-item">'
                f'<span class="legend-dot" style="background:{CHART_COLORS[i%len(CHART_COLORS)]}"></span>'
                f'<span>{html.escape(ds["version"])}{suffix}</span></div>\n'
            )

        base_version = html.escape(base_ds["version"])
        ts_display   = datetime.now().strftime("%d.%m.%Y %H:%M")

        # Файлы без measure_schema — версия 1 (сырые пороги CPU, "time" =
        # среднее); сравнивать их напрямую с версией 2 (нормированные пороги,
        # "time" = медиана) некорректно — это разные величины (см. CLAUDE.md,
        # раздел «Нагрузочный стенд: этап 1»). measure_schema пишется в JSON
        # с 25.08.2026, но раньше нигде не читался при сравнении версий —
        # страница молча строила график по несопоставимым числам.
        schema_versions = {ds["data"].get("measure_schema", 1) for ds in datasets}
        schema_warning_html = ""
        if len(schema_versions) > 1:
            schema_warning_html = (
                '<div style="background:#4a2a1a;border:1px solid #d68910;'
                'border-radius:6px;padding:10px 16px;margin:0 0 16px;color:#f5cba7">'
                '⚠️ В сравнении смешаны файлы разных версий схемы замера '
                f'({", ".join(str(v) for v in sorted(schema_versions))}) — '
                'до 25.08.2026 «время» считалось как среднее по сырым '
                'CPU-порогам, после — как медиана по нормированным. Числа '
                'из разных версий несопоставимы напрямую.</div>\n'
            )

        return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Сравнение версий R7-Office</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body{{font-family:Arial,sans-serif;margin:0;padding:20px;background:#f5f6fa;color:#333}}
  h1{{color:#2c3e50;margin-bottom:2px}}
  h2{{color:#2c3e50;margin-top:28px;margin-bottom:10px}}
  .subtitle{{color:#888;font-size:.9em;margin-bottom:14px}}
  .legend{{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:20px}}
  .legend-item{{display:flex;align-items:center;gap:6px;font-size:.88em}}
  .legend-dot{{width:13px;height:13px;border-radius:50%;flex-shrink:0}}
  .sys-cards{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:8px}}
  .sys-card{{background:#fff;padding:14px 16px;border-radius:8px;
    box-shadow:0 1px 4px rgba(0,0,0,.12);min-width:200px;flex:1}}
  .sys-title{{font-weight:bold;font-size:.9em;margin-bottom:8px}}
  .sys-row{{display:flex;gap:8px;font-size:.82em;margin-bottom:3px;color:#444}}
  .sys-lbl{{color:#888;min-width:90px}}
  .charts{{display:grid;grid-template-columns:1fr;gap:18px;margin-bottom:28px}}
  .chart-box{{background:#fff;padding:16px;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.12)}}
  table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;
    overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.12);margin-bottom:16px}}
  th{{color:#fff;padding:8px 10px;text-align:center;font-size:.8em}}
  th:first-child{{background:#2c3e50;text-align:left}}
  td{{padding:7px 10px;border-bottom:1px solid #eee;font-size:.87em;text-align:center}}
  td:first-child{{text-align:left;font-weight:500}}
  tr:last-child td{{border-bottom:none}}
  tr:hover td{{background:#f0f4ff}}
  .delta-better{{color:#27ae60;font-weight:bold}}
  .delta-worse{{color:#e74c3c;font-weight:bold}}
  .delta-same{{color:#e67e22}}
  .delta-base{{color:#aaa}}
  .legend-note{{font-size:.82em;color:#555;margin-bottom:10px;padding:8px 12px;
    background:#fff;border-radius:6px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  .pdf-btn{{position:fixed;top:16px;right:16px;padding:8px 18px;background:#2c3e50;
    color:#fff;border:none;border-radius:6px;font-size:.9em;cursor:pointer;
    box-shadow:0 2px 6px rgba(0,0,0,.25);z-index:1000}}
  .pdf-btn:hover{{background:#34495e}}
  @media print{{
    .pdf-btn{{display:none}}
    body{{background:#fff}}
    canvas,.chart-box,table{{page-break-inside:avoid}}
    h1,h2,h3{{page-break-after:avoid}}
  }}
</style>
</head>
<body>
<button class="pdf-btn" onclick="window.print()">📄 Сохранить как PDF</button>
<h1>Сравнение версий R7-Office</h1>
<div class="subtitle">Сформировано: {ts_display} &nbsp;|&nbsp; Базовая версия: <strong>{base_version}</strong></div>
{schema_warning_html}<div class="legend">
{legend_items}</div>

<h2>Сведения о системе</h2>
<div class="sys-cards">{sys_cards}</div>

<h2>Графики производительности</h2>
<div class="charts">
  <div class="chart-box"><canvas id="timeChart"></canvas></div>
  <div class="chart-box"><canvas id="ramChart"></canvas></div>
  <div class="chart-box"><canvas id="cpuChart"></canvas></div>
</div>

<h2>Детальное сравнение</h2>
<div class="legend-note">
  <span class="delta-better">Зелёный</span> — быстрее базовой версии &nbsp;&nbsp;
  <span class="delta-worse">Красный</span> — медленнее &nbsp;&nbsp;
  <span class="delta-same">Оранжевый</span> — разница ≤ 5%
</div>
<table>
<thead>{th1}{th2}</thead>
<tbody>
{table_rows}</tbody>
</table>

<script>
const labels = {labels_json};
new Chart(document.getElementById('timeChart'), {{
  type:'bar', data:{{labels, datasets:{time_ds_json}}},
  options:{{responsive:true,
    plugins:{{title:{{display:true,text:'Время выполнения операций (сек)'}},legend:{{position:'{legend_pos}'}}}},
    scales:{{y:{{beginAtZero:true}}}}}}
}});
new Chart(document.getElementById('ramChart'), {{
  type:'line', data:{{labels, datasets:{ram_ds_json}}},
  options:{{responsive:true,
    plugins:{{title:{{display:true,text:'Потребление RAM (МБ)'}},legend:{{position:'{legend_pos}'}}}},
    scales:{{y:{{beginAtZero:false}}}}}}
}});
new Chart(document.getElementById('cpuChart'), {{
  type:'line', data:{{labels, datasets:{cpu_ds_json}}},
  options:{{responsive:true,
    plugins:{{title:{{display:true,text:'Нагрузка на CPU (%)'}},legend:{{position:'{legend_pos}'}}}},
    scales:{{y:{{beginAtZero:true}}}}}}
}});
</script>
</body>
</html>"""

    # ---------------------- Batch-режим ----------------------

    def run_batch_mode(self):
        """Entry point for Batch mode — validates prerequisites then shows config dialog."""
        if self._batch_running:
            messagebox.showwarning("Batch уже выполняется",
                                   "Дождитесь завершения текущего Batch-прогона.")
            return
        if self._perf_running:
            messagebox.showwarning("Выполняется тест производительности",
                                   "Оба режима управляют клавиатурой Р7-Офис и не могут "
                                   "работать одновременно. Дождитесь завершения теста "
                                   "или нажмите «Остановить» на вкладке «Производительность».")
            return
        if not ctypes.windll.shell32.IsUserAnAdmin():
            messagebox.showerror(
                "Ошибка прав",
                "Batch-режим требует прав администратора.\n"
                "Перезапустите программу от имени администратора."
            )
            return
        if not PYAUTOGUI_OK or not pyperclip or not EXCEL_OK or not WIN32_OK:
            missing = []
            if not PYAUTOGUI_OK: missing.append("pyautogui")
            if not pyperclip:    missing.append("pyperclip")
            if not EXCEL_OK:     missing.append("openpyxl")
            if not WIN32_OK:     missing.append("pywin32")
            messagebox.showerror("Ошибка",
                                 f"Отсутствуют библиотеки: {', '.join(missing)}\n"
                                 "Установите: pip install " + " ".join(missing))
            return
        files = (list(self.distributives_folder.glob("*.msi")) +
                 list(self.distributives_folder.glob("*.exe")))
        files.sort(key=lambda f: self._extract_version(f.stem) or f.name)
        if not files:
            messagebox.showwarning("Нет дистрибутивов",
                                   "В папке Distributives не найдено .msi/.exe файлов.")
            return
        self._show_batch_config_dialog(files)

    def _show_batch_config_dialog(self, files):
        """Shows batch configuration dialog: version checkboxes, test file, options."""
        dlg = tk.Toplevel(self.root)
        dlg.transient(self.root)
        dlg.configure(bg=COLORS["bg"])
        dlg.title("Batch-режим")
        dlg.resizable(False, False)
        dlg.grab_set()

        ttk.Label(dlg, text=f"Найдено дистрибутивов: {len(files)}",
                  font=("Arial", 10, "bold")).pack(pady=(14, 4), padx=16, anchor=tk.W)

        # ── Список версий ─────────────────────────────────────────────────────
        # Прокручиваемый список вместо обычного pack() — при resizable(False, False)
        # и десятке+ дистрибутивов список раньше выталкивал кнопки «Запустить»/
        # «Отмена» за нижнюю границу экрана без какой-либо возможности прокрутки.
        ver_frame = ttk.LabelFrame(dlg, text="Выберите версии для тестирования", padding="8")
        ver_frame.pack(fill=tk.BOTH, padx=16, pady=4)

        MAX_LIST_HEIGHT = 220
        ver_canvas = tk.Canvas(ver_frame, borderwidth=0, highlightthickness=0,
                               bg=COLORS["bg"])
        ver_vsb = ttk.Scrollbar(ver_frame, orient=tk.VERTICAL, command=ver_canvas.yview)
        ver_canvas.configure(yscrollcommand=ver_vsb.set)
        ver_inner = ttk.Frame(ver_canvas)
        ver_inner_id = ver_canvas.create_window((0, 0), window=ver_inner, anchor="nw")

        def _ver_on_inner_cfg(_e):
            ver_canvas.configure(scrollregion=ver_canvas.bbox("all"))
        def _ver_on_canvas_cfg(e):
            ver_canvas.itemconfig(ver_inner_id, width=e.width)
        ver_inner.bind("<Configure>", _ver_on_inner_cfg)
        ver_canvas.bind("<Configure>", _ver_on_canvas_cfg)

        def _ver_on_mwheel(e):
            ver_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        ver_canvas.bind("<Enter>", lambda _e: ver_canvas.bind_all("<MouseWheel>", _ver_on_mwheel))
        ver_canvas.bind("<Leave>", lambda _e: ver_canvas.unbind_all("<MouseWheel>"))

        ver_vars = {}
        for f in files:
            var = tk.BooleanVar(value=True)
            ver_vars[f] = var
            ttk.Checkbutton(ver_inner, text=f.name, variable=var).pack(anchor=tk.W, pady=1)

        dlg.update_idletasks()
        content_h = min(MAX_LIST_HEIGHT, max(ver_inner.winfo_reqheight(), 24))
        ver_canvas.configure(height=content_h)
        ver_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ver_vsb.pack(side=tk.RIGHT, fill=tk.Y)

        mini = ttk.Frame(dlg)
        mini.pack(fill=tk.X, padx=16, pady=(0, 4))
        ttk.Button(mini, text="☑ Все", width=7,
                   command=lambda: [v.set(True) for v in ver_vars.values()]).pack(side=tk.LEFT)
        ttk.Button(mini, text="☐ Снять", width=7,
                   command=lambda: [v.set(False) for v in ver_vars.values()]).pack(
                       side=tk.LEFT, padx=3)

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=16, pady=8)

        # ── Тестовый файл ─────────────────────────────────────────────────────
        file_frame = ttk.LabelFrame(dlg, text="Тестовый файл", padding="8")
        file_frame.pack(fill=tk.X, padx=16, pady=4)

        test_file_var = tk.StringVar()
        for sd in [self.test_files_folder, BASE_DIR, Path.home() / "Downloads", Path.home() / "Загрузки"]:
            if not sd.exists():
                continue
            for pat in ["файл-для-теста-Р7-офис-50К*.xlsx", "*50К*.xlsx"]:
                for found in sd.glob(pat):
                    test_file_var.set(str(found))
                    break
            if test_file_var.get():
                break

        file_row = ttk.Frame(file_frame)
        file_row.pack(fill=tk.X)
        ttk.Label(file_row, text="Файл:").pack(side=tk.LEFT)
        ttk.Entry(file_row, textvariable=test_file_var, width=38).pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        def browse_test_file():
            path = filedialog.askopenfilename(
                parent=dlg, title="Выберите тестовый файл",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
            if path:
                test_file_var.set(path)

        ttk.Button(file_row, text="Обзор", command=browse_test_file).pack(side=tk.LEFT)

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=16, pady=8)

        # ── Опции ─────────────────────────────────────────────────────────────
        opt_frame = ttk.LabelFrame(dlg, text="Параметры", padding="8")
        opt_frame.pack(fill=tk.X, padx=16, pady=4)

        stop_on_error_var = tk.BooleanVar(value=True)
        cleanup_var       = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_frame, text="Останавливаться при первой ошибке",
                        variable=stop_on_error_var).pack(anchor=tk.W)
        ttk.Checkbutton(opt_frame, text="Удалять временные файлы кеша после каждого теста",
                        variable=cleanup_var).pack(anchor=tk.W, pady=(4, 0))

        # ── Кнопки ────────────────────────────────────────────────────────────
        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(pady=12, padx=16, fill=tk.X)

        def on_start():
            selected = [f for f, v in ver_vars.items() if v.get()]
            if not selected:
                messagebox.showwarning("Нет выбора",
                                       "Выберите хотя бы одну версию.", parent=dlg)
                return
            tf = test_file_var.get().strip()
            if not tf or not Path(tf).exists():
                messagebox.showwarning("Файл не найден",
                                       "Укажите существующий тестовый файл.", parent=dlg)
                return
            dlg.destroy()
            self._start_batch_run(selected, Path(tf),
                                  stop_on_error_var.get(), cleanup_var.get())

        ttk.Button(btn_frame, text="▶ Запустить", command=on_start).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Отмена", command=dlg.destroy).pack(side=tk.LEFT)

        dlg.update_idletasks()
        dlg.minsize(460, dlg.winfo_reqheight())

    def _start_batch_run(self, versions, test_file, stop_on_error, cleanup):
        """Creates the progress window and launches the batch worker thread."""
        prog = tk.Toplevel(self.root)
        prog.transient(self.root)
        prog.configure(bg=COLORS["bg"])
        prog.title("Batch-режим: выполнение")
        prog.geometry("680x540")
        prog.resizable(True, True)

        # ── Шапка прогресса ───────────────────────────────────────────────────
        top = ttk.Frame(prog, padding="10")
        top.pack(fill=tk.X)

        lbl_current = ttk.Label(top, text="Подготовка...", font=("Arial", 10, "bold"))
        lbl_current.pack(anchor=tk.W)

        progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(top, variable=progress_var,
                        maximum=len(versions), mode="determinate").pack(
                            fill=tk.X, pady=(4, 0))

        # ── Список версий с иконками ──────────────────────────────────────────
        ver_list_frame = ttk.LabelFrame(prog, text="Версии", padding="6")
        ver_list_frame.pack(fill=tk.X, padx=10, pady=4)

        ver_labels = {}
        for f in versions:
            var = tk.StringVar(value=f"⏳ {f.name}")
            ttk.Label(ver_list_frame, textvariable=var, anchor=tk.W).pack(anchor=tk.W, pady=1)
            ver_labels[f] = var

        # ── Лог ───────────────────────────────────────────────────────────────
        log_frame = ttk.LabelFrame(prog, text="Лог", padding="4")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        log_text = tk.Text(log_frame, font=("Consolas", 9), wrap=tk.WORD,
                          bg=COLORS["log_bg"], fg=COLORS["text"],
                          insertbackground=COLORS["text"],
                          borderwidth=0, highlightthickness=0)
        log_scroll = ttk.Scrollbar(log_frame, command=log_text.yview)
        log_text.configure(yscrollcommand=log_scroll.set)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ── Управление ────────────────────────────────────────────────────────
        ctrl = ttk.Frame(prog, padding="6")
        ctrl.pack(fill=tk.X)

        stop_event  = threading.Event()
        pause_event = threading.Event()
        paused = [False]

        def toggle_pause():
            if paused[0]:
                paused[0] = False
                pause_event.clear()
                btn_pause.config(text="⏸ Пауза")
            else:
                paused[0] = True
                pause_event.set()
                btn_pause.config(text="▶ Продолжить")

        def request_stop():
            stop_event.set()
            pause_event.clear()
            btn_stop.config(state=tk.DISABLED)
            _log("⏹ Запрошена остановка...")

        btn_pause = ttk.Button(ctrl, text="⏸ Пауза", command=toggle_pause)
        btn_stop  = ttk.Button(ctrl, text="⏹ Остановить", command=request_stop)
        btn_pause.pack(side=tk.LEFT, padx=5)
        btn_stop.pack(side=tk.LEFT, padx=5)

        # ── UI-callback-и ──────────────────────────────────────────────────────
        def _log(msg):
            def _do():
                try:
                    ts = datetime.now().strftime("%H:%M:%S")
                    log_text.insert(tk.END, f"[{ts}] {msg}\n")
                    log_text.see(tk.END)
                    self.add_test_log(msg)
                except tk.TclError:
                    pass
            try:
                prog.after(0, _do)
            except tk.TclError:
                pass

        def _set_current(text):
            try:
                prog.after(0, lambda: lbl_current.config(text=text))
            except tk.TclError:
                pass

        def _set_ver_status(f, text):
            def _do():
                try:
                    if f in ver_labels:
                        ver_labels[f].set(text)
                except tk.TclError:
                    pass
            try:
                prog.after(0, _do)
            except tk.TclError:
                pass

        def _set_progress(n):
            try:
                prog.after(0, lambda: progress_var.set(n))
            except tk.TclError:
                pass

        def _on_done(batch_results, errors):
            def _do():
                try:
                    btn_pause.config(state=tk.DISABLED)
                    btn_stop.config(state=tk.DISABLED)
                    ok = sum(1 for r in batch_results if r.get("success"))
                    _log(f"✅ Batch-режим завершён. Успешно: {ok}, Ошибок: {errors}")
                    self.status_var.set(f"Batch завершён: {ok}/{len(batch_results)} успешно")
                    self.detect_current_version()
                    if batch_results:
                        html = self._generate_batch_summary_html(batch_results)
                        ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
                        out_path = self.reports_folder / f"batch_summary_{ts_now}.html"
                        try:
                            out_path.write_text(html, encoding="utf-8")
                            _log(f"📊 Сводный отчёт: {out_path.name}")
                            webbrowser.open(str(out_path))
                        except Exception as e:
                            _log(f"⚠️ Ошибка сохранения отчёта: {e}")
                except tk.TclError:
                    pass
            try:
                prog.after(0, _do)
            except tk.TclError:
                pass

        self._batch_running = True

        def _batch_thread():
            try:
                self._batch_worker(versions, test_file, stop_on_error, cleanup,
                                   _log, _set_current, _set_ver_status, _set_progress,
                                   _on_done, stop_event, pause_event)
            finally:
                self._batch_running = False

        threading.Thread(target=_batch_thread, daemon=True).start()

    def _batch_worker(self, versions, test_file, stop_on_error, cleanup,
                      log_cb, current_cb, ver_status_cb, progress_cb,
                      done_cb, stop_event, pause_event):
        """Batch worker thread: install each version, run tests, collect results."""
        batch_results = []
        errors = 0
        log_cb(f"🚀 Запуск Batch-режима: найдено {len(versions)} версий")

        for idx, dist_file in enumerate(versions):
            if stop_event.is_set():
                log_cb("⏹ Остановлено пользователем.")
                break

            ver_name = self._extract_version(dist_file.stem) or dist_file.stem
            current_cb(f"Текущая версия: {ver_name} ({idx + 1} из {len(versions)})")
            ver_status_cb(dist_file, f"🔄 {dist_file.name}: выполняется...")
            log_cb(f"--- Версия {idx + 1}/{len(versions)}: {dist_file.name} ---")

            result = {
                "file":            dist_file.name,
                "version":         ver_name,
                "success":         False,
                "error":           None,
                "open_elapsed":    None,
                "vlookup_elapsed": None,
                "peak_ram":        None,
                "avg_ram":         None,
                "peak_cpu":        None,
            }

            try:
                log_cb("🗑️ Удаление текущей версии...")
                if not self.uninstall_current_version():
                    raise RuntimeError("Удаление текущей версии не завершилось успешно")
                time.sleep(2)

                log_cb(f"📥 Установка {dist_file.name}...")
                if not self.install_version(dist_file):
                    raise RuntimeError("Установка не завершилась успешно (таймаут или код ошибки)")
                self.detect_current_version()
                ver_display = (self.current_version_info.get("name")
                               if self.current_version_info else ver_name)
                log_cb(f"✅ Установлена: {ver_display}")

                if pause_event.is_set():
                    log_cb("⏸ Пауза...")
                    pause_event.wait()
                    log_cb("▶ Продолжение...")
                if stop_event.is_set():
                    break

                if cleanup:
                    cleared = self._clear_r7_cache()
                    if cleared:
                        log_cb(f"🧹 Очищено {cleared} объектов кеша")

                test_result = self._batch_run_single_version(
                    test_file, ver_display, log_cb, stop_event, pause_event)

                if test_result:
                    result.update(test_result)
                    result["success"] = True
                    ot  = result.get("open_elapsed") or 0
                    vt  = result.get("vlookup_elapsed")
                    vts = f", ВПР {vt:.2f} сек" if vt else ""
                    log_cb(f"✅ {ver_name}: открытие {ot:.2f} сек{vts}")
                    ver_status_cb(dist_file,
                                  f"✅ {dist_file.name}: {ot:.1f} сек"
                                  + (f" / ВПР {vt:.1f} сек" if vt else ""))
                else:
                    raise RuntimeError("Тест не вернул результатов")

            except Exception as e:
                errors += 1
                result["error"] = str(e)
                log_cb(f"❌ {ver_name}: ошибка — {e}")
                ver_status_cb(dist_file, f"❌ {dist_file.name}: ошибка")
                if stop_on_error:
                    log_cb("🛑 Остановка (включена опция 'стоп при ошибке')")
                    batch_results.append(result)
                    break

            batch_results.append(result)
            progress_cb(idx + 1)

            if pause_event.is_set():
                log_cb("⏸ Пауза между версиями...")
                pause_event.wait()
                log_cb("▶ Продолжение...")

        done_cb(batch_results, errors)

    def _batch_run_single_version(self, test_file, version_label, log_cb,
                                  stop_event, pause_event):
        """Runs the full 12-operation stress test for the currently installed version.

        Returns a result dict with timing/resource data, or None on critical failure.
        """
        r7_path = self._find_r7_path()
        if not r7_path:
            log_cb("❌ Р7-Офис не найден.")
            return None

        # ── Оконные вспомогательные функции ──────────────────────────────────
        if WIN32_OK:
            import win32gui as _wg
            import win32con as _wc

        def _find_hwnd():
            if not WIN32_OK:
                return None
            wins = []
            def _cb(h, _):
                if _wg.IsWindowVisible(h):
                    t = _wg.GetWindowText(h)
                    if "Р7-Офис" in t or test_file.stem[:12] in t:
                        wins.append(h)
            _wg.EnumWindows(_cb, wins)
            return wins[0] if wins else None

        def _focus():
            hwnd = _find_hwnd()
            if hwnd:
                try:
                    _wg.SetForegroundWindow(hwnd)
                    time.sleep(0.2)
                    return True
                except Exception:
                    pass
            return not WIN32_OK

        def _maximize():
            hwnd = _find_hwnd()
            if hwnd and WIN32_OK:
                _wg.ShowWindow(hwnd, _wc.SW_MAXIMIZE)
                time.sleep(0.5)

        def _close_update_dlg(search_timeout=0):
            self._close_update_dialog_if_exists(log_cb=log_cb,
                                                search_timeout=search_timeout)

        # Зеркало safe_hotkey/safe_press из _spreadsheet_worker: без interval,
        # паузы — только явные, через _pace (вычитаются из замера).
        KEY_PACE  = self.OP_KEY_PACE
        MENU_PACE = self.OP_MENU_PACE

        def _hk(*keys):
            pyautogui.hotkey(*keys)

        def _pr(key, n=1, pace=0.0):
            for _ in range(n):
                pyautogui.press(key)
                if pace:
                    self._pace(pace)

        # ── Открытие Р7-Офис ──────────────────────────────────────────────────
        log_cb(f"▶ Запуск Р7-Офис: {test_file.name}")
        # Порт проверяется ДО старта секундомера — см. комментарий в
        # _spreadsheet_worker (зеркалим сюда, как требует правило репозитория
        # про синхронность мест паузы между Batch и вкладкой «Производительность»).
        debug_args = self._prepare_webdriver_launch(log_cb=log_cb, filename_hint=test_file.name)
        open_start = time.time()
        subprocess.Popen([r7_path, str(test_file), *debug_args], shell=True)

        deadline = time.time() + 60
        while time.time() < deadline:
            if _find_hwnd():
                break
            time.sleep(0.5)
        else:
            log_cb("❌ Окно Р7-Офис не появилось.")
            return None

        # Подготовку окна засекаем отдельно и вычитаем — как в _spreadsheet_worker,
        # иначе она попадает в замер открытия файла.
        _setup_start = time.time()
        _maximize()
        _focus()
        _close_update_dlg(search_timeout=0)
        _setup_elapsed = time.time() - _setup_start

        # Фоновый мониторинг окна обновления на весь период теста
        _upd_stop = threading.Event()
        threading.Thread(
            target=self._monitor_update_dialog,
            args=(_upd_stop,),
            kwargs={"log_cb": log_cb},
            daemon=True,
        ).start()
        log_cb("🔍 Запущен мониторинг окна обновления (проверка каждые 2 сек)")

        try:
            data_ready   = self._wait_until_r7_ready(_find_hwnd, timeout=120, log_cb=log_cb)
            open_elapsed = time.time() - open_start - _setup_elapsed
            log_cb(f"✅ Файл открыт за {open_elapsed:.2f} сек"
                   + ("" if data_ready else " (таймаут — возможна частичная загрузка)"))
            _focus()

            # Зеркало _spreadsheet_worker — подключение к CDP, базовый
            # DOM-снимок ДО первой операции (issue #9, см.
            # _capture_cdp_ui_baseline) и разовая диагностика api редактора.
            self._cdp_ensure_connected(log_cb=log_cb)
            self._capture_cdp_ui_baseline(log_cb=log_cb)
            self._cdp_log_api_info(log_cb=log_cb)

            # ── Мониторинг ресурсов ───────────────────────────────────────────────
            self._r7_pids = None
            self._x2t_logged_pids = set()  # сбросить дедуп x2t перед новым тестом
            r7_procs = self._get_r7_processes(log_cb=log_cb)

            sample0 = self._sample_r7_resources(r7_procs)
            results = [{
                "name": "Открытие файла", "time": open_elapsed, "error": None,
                "ram":            sample0["ram_mb"]       if sample0 else None,
                "cpu":            sample0["cpu_raw_pct"]   if sample0 else None,
                "cpu_normalized": sample0["cpu_norm_pct"]  if sample0 else None,
                "threads":        sample0["threads"]       if sample0 else None,
                "uptime_sec":     sample0["uptime_sec"]    if sample0 else None,
            }]

            def measure(name, func):
                nonlocal r7_procs
                if stop_event.is_set():
                    return
                if pause_event.is_set():
                    log_cb("⏸ Пауза...")
                    pause_event.wait()
                    log_cb("▶ Продолжение...")
                log_cb(f"⏳ {name}...")
                _focus()
                # Как в run_test_with_runs: секундомер останавливается по признаку
                # «Р7 освободился», а собственные паузы вычитаются.
                self._paced_total = 0.0
                self._op_start_grace = None
                self._op_max_wait = None
                self._op_via_cdp = False
                self._cdp_api_ms = 0.0
                t0  = time.time()
                err = None
                try:
                    func()
                except Exception as e:
                    err = str(e)
                done_ts, status = self._wait_operation_done(_find_hwnd, log_cb=log_cb)
                if status == "timeout":
                    elapsed = time.time() - t0 - self._paced_total
                else:
                    elapsed = max(0.0, done_ts - t0 - self._paced_total)
                # Зеркало run_test_with_runs: добиваем модалку «Вставить ячейки»
                # и доводим отложенную проверку CDP-операции после закрытия
                # замера, чтобы паузы и round-trip не съедали результат.
                self._flush_pending_modal_confirm(log_cb=log_cb)
                self._flush_pending_cdp_verify(log_cb=log_cb)
                time.sleep(0.5)
                self._r7_pids = None
                r7_procs = self._get_r7_processes(log_cb=log_cb)
                sample = self._sample_r7_resources(r7_procs)
                self._log_resources(sample, log_cb=log_cb)
                # Зеркало run_test_with_runs: на CDP-пути below_floor означает
                # «api отработал синхронно», а не «измерить не смогли».
                _mark = {"below_floor": (" (api-вызов, Р7 не стал занятым)"
                                         if self._op_via_cdp
                                         else " (ниже порога измерения)"),
                         "timeout": " (Р7 не освободился)"}.get(status, "")
                # api_ms — субмиллисекундное разрешение (см. _cdp_sequence),
                # рядом с elapsed (settle_ms), а не вместо него.
                api_ms = round(self._cdp_api_ms, 3) if self._op_via_cdp else None
                _api_note = f" [api: {api_ms:.2f} мс]" if api_ms is not None else ""
                log_cb(f"   ✅ {name}: {elapsed:.3f} сек{_api_note}{_mark}"
                       + (f" (ошибка: {err})" if err else ""))
                results.append({
                    "name": name, "time": elapsed, "error": err,
                    "ram":            sample["ram_mb"]      if sample else None,
                    "cpu":            sample["cpu_raw_pct"]  if sample else None,
                    "cpu_normalized": sample["cpu_norm_pct"] if sample else None,
                    "threads":        sample["threads"]      if sample else None,
                    "uptime_sec":     sample["uptime_sec"]    if sample else None,
                    "below_floor":    status == "below_floor",
                    "api_ms":         api_ms,
                })

            # ── Тест-функции (зеркало _spreadsheet_worker) ────────────────────────
            # Все паузы — через _pace, чтобы вычитаться из замера. Значения и места
            # пауз должны совпадать с одиночным тестом, иначе Batch и вкладка
            # «Производительность» дадут несравнимые цифры.
            def paste_big():
                if self._cdp_paste_big(log_cb=log_cb, key_pace=KEY_PACE):
                    return
                _hk('shift', 'f11')
                self._pace(KEY_PACE)
                _hk('ctrl', 'v')

            def add_col_hk():
                if self._cdp_add_column(log_cb=log_cb, key_pace=KEY_PACE):
                    return
                _hk('ctrl', 'pageup')
                self._pace(KEY_PACE)
                pyautogui.press('right')
                _hk('ctrl', 'shift', '=')

            def add_col_menu():
                # Как и в _spreadsheet_worker: на CDP-пути «меню Вставка» и
                # «горячие клавиши» — один и тот же вызов asc_insertCells.
                if self._cdp_add_column(log_cb=log_cb, key_pace=KEY_PACE):
                    return
                _hk('ctrl', 'pageup')
                self._pace(KEY_PACE)
                pyautogui.press('right')
                _hk('alt', 'i')
                self._pace(MENU_PACE)
                _pr('c')

            def paste_hk(cell_count, paste_offset):
                if self._cdp_copy_paste(cell_count, paste_offset,
                                        log_cb=log_cb, key_pace=KEY_PACE):
                    return
                _hk('ctrl', 'home')
                for _ in range(cell_count - 1):
                    pyautogui.hotkey('shift', 'right')
                _hk('ctrl', 'c')
                self._pace(KEY_PACE)
                pyautogui.press('right', presses=paste_offset)
                _hk('ctrl', 'v')

            def paste_pkm(cell_count, paste_offset):
                # Зеркало copy_paste_context: вставка ячеек со сдвигом вниз
                # через asc_insertCells, без контекстного меню и модалки.
                if self._cdp_copy_paste(cell_count, paste_offset, shift="down",
                                        log_cb=log_cb, key_pace=KEY_PACE):
                    return
                _hk('ctrl', 'home')
                for _ in range(cell_count - 1):
                    pyautogui.hotkey('shift', 'right')
                pyautogui.click(button='right')
                self._pace(MENU_PACE)
                # Зеркало copy_paste_context: разовый дамп состава меню.
                self._cdp_dump_ui("контекстное меню ячейки (копирование)",
                                  log_cb=log_cb, charge_pace=True)
                if not self._cdp_click_context_item(("копировать", "copy"),
                                                    log_cb=log_cb):
                    _pr('down', 2, pace=MENU_PACE)
                    _pr('enter')
                self._pace(MENU_PACE)
                pyautogui.press('right', presses=paste_offset)
                pyautogui.click(button='right')
                self._pace(MENU_PACE)
                self._cdp_dump_ui("контекстное меню ячейки (вставка)",
                                  log_cb=log_cb, charge_pace=True)
                if not self._cdp_click_context_item(
                        ("вставить скопированные ячейки", "вставить ячейки",
                         "insert copied cells", "insert cells"), log_cb=log_cb):
                    _pr('down', 3, pace=MENU_PACE)
                    _pr('enter')
                # Модалка «Вставить ячейки» — зеркало copy_paste_context()
                # из _spreadsheet_worker (см. _confirm_modal_enter)
                self._confirm_modal_enter()

            def vlookup():
                _hk('ctrl', 'pagedown')
                self._pace(KEY_PACE)
                _hk('ctrl', 'home')
                pyperclip.copy('=VLOOKUP(A2;Лист1!A:B;2;FALSE)')
                _hk('ctrl', 'v')
                self._pace(KEY_PACE)
                _pr('enter')
                _hk('ctrl', 'shift', 'down')
                _hk('ctrl', 'd')

            def del_col():
                _hk('ctrl', 'home')
                pyautogui.press('right')
                pyautogui.press('delete')

            def save_as_pdf():
                tmp_pdf = str(Path(os.environ.get("TEMP", ".")) /
                              f"temp_export_x2t_{int(time.time())}.pdf")
                self._op_start_grace = self.OP_PDF_GRACE_SEC

                _hk('ctrl', 'shift', 's')
                _t_dlg = time.time()
                if not self._wait_for_window_title(("сохранить как", "save as"), timeout=3.0):
                    self._paced_total += time.time() - _t_dlg
                    log_cb("   ⚠️ Ctrl+Shift+S не открыл диалог, пробуем меню Файл")
                    _hk('alt', 'f')
                    self._pace(MENU_PACE)
                    _pr('down', 3, pace=MENU_PACE)
                    _pr('enter')
                    self._wait_for_window_title(("сохранить как", "save as"), timeout=3.0)

                pyperclip.copy(tmp_pdf)
                _hk('ctrl', 'a')
                _hk('ctrl', 'v')
                self._pace(KEY_PACE)
                _pr('enter')
                # Прежний _focus() здесь добавлял 0.2 сек внутрь замера. Фокус и так
                # восстанавливается в начале следующего measure().

            def select_all():
                # Зеркало select_all() из _spreadsheet_worker: укороченный
                # предохранитель, иначе Ctrl+A на большом файле занимает Р7
                # десятками секунд и выглядит как зависание.
                self._op_max_wait = self.OP_SELECT_ALL_MAX_SEC
                if self._cdp_select_all(log_cb=log_cb):
                    return
                _hk('ctrl', 'a')

            def copy_all():
                if self._cdp_copy(log_cb=log_cb):
                    return
                _hk('ctrl', 'c')

            def add_sheet():
                if self._cdp_add_sheet(log_cb=log_cb):
                    return
                _hk('shift', 'f11')

            # ── Выполнение тестов ─────────────────────────────────────────────────
            measure("Выделение всех ячеек (Ctrl+A)",      select_all)
            measure("Копирование всех ячеек (Ctrl+C)",     copy_all)
            measure("Вставка большого массива (Ctrl+V)",    paste_big)
            measure("Добавление нового листа",              add_sheet)
            measure("Добавление столбца (горячие клавиши)", add_col_hk)
            measure("Добавление столбца (меню Вставка)",    add_col_menu)
            measure("Вставка 1 ячейки (горячие клавиши)",   lambda: paste_hk(1, 10))
            measure("Вставка 5 ячеек (горячие клавиши)",    lambda: paste_hk(5, 15))
            measure("Вставка 1 ячейки (ПКМ)",               lambda: paste_pkm(1, 10))
            measure("Вставка 5 ячеек (ПКМ)",                lambda: paste_pkm(5, 15))
            measure("Функция ВПР (50K строк)",              vlookup)
            measure("Удаление столбца (Del)",               del_col)
            measure("Сохранение в PDF (конвертация x2t)",   save_as_pdf)
            self._cleanup_x2t_temp_pdfs(log_cb=log_cb)

            # ── Статистика ────────────────────────────────────────────────────────
            ram_vals      = [r["ram"] for r in results if r.get("ram") is not None]
            cpu_vals      = [r["cpu"] for r in results if r.get("cpu") is not None]
            cpu_norm_vals = [r["cpu_normalized"] for r in results if r.get("cpu_normalized") is not None]
            peak_ram = max(ram_vals) if ram_vals else None
            avg_ram  = round(sum(ram_vals) / len(ram_vals), 1) if ram_vals else None
            peak_cpu = max(cpu_vals) if cpu_vals else None
            peak_cpu_norm = max(cpu_norm_vals) if cpu_norm_vals else None
            avg_cpu_norm  = round(sum(cpu_norm_vals) / len(cpu_norm_vals), 1) if cpu_norm_vals else None

            # ── Закрытие Р7-Офис ──────────────────────────────────────────────────
            _upd_stop.set()
            log_cb("🔍 Мониторинг окна обновления остановлен")
            log_cb("🔚 Закрытие Р7-Офис...")
            self._close_r7_gracefully(_find_hwnd(), log_cb=log_cb)

            # ── Сохранение JSON ───────────────────────────────────────────────────
            ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
            json_path = self.reports_folder / f"performance_full_{ts_now}.json"
            try:
                with open(json_path, "w", encoding="utf-8") as jf:
                    json.dump({
                        "timestamp": ts_now,
                        "measure_schema": MEASURE_SCHEMA_VERSION,
                        "version":   version_label,
                        "test_file": str(test_file),
                        "system": self._build_system_info(),
                        "summary": {
                            "peak_ram_mb": peak_ram, "avg_ram_mb": avg_ram,
                            "min_ram_mb":  min(ram_vals) if ram_vals else None,
                            "peak_cpu_pct": peak_cpu,
                            "peak_cpu_normalized_pct": peak_cpu_norm,
                            "avg_cpu_normalized_pct": avg_cpu_norm,
                        },
                        "results": results,
                    }, jf, indent=2, ensure_ascii=False)
                log_cb(f"📄 JSON сохранён: {json_path.name}")
            except Exception as e:
                log_cb(f"⚠️ Ошибка сохранения JSON: {e}")

            vpr_r = next((r for r in results if r["name"] == "Функция ВПР (50K строк)"), None)
            return {
                "open_elapsed":     open_elapsed,
                "vlookup_elapsed":  vpr_r["time"] if vpr_r else None,
                "peak_ram":         peak_ram,
                "avg_ram":          avg_ram,
                "peak_cpu":         peak_cpu,
                "peak_cpu_normalized": peak_cpu_norm,
                "results":          results,
                "json_path":        str(json_path),
            }
        finally:
            _upd_stop.set()
            self._close_webdriver_connector()

    def _generate_batch_summary_html(self, batch_results):
        """Builds summary HTML report for all batch results."""
        ts_display = datetime.now().strftime("%d.%m.%Y %H:%M")
        n          = len(batch_results)
        versions   = [r["version"] for r in batch_results]

        open_times = [r.get("open_elapsed")    for r in batch_results]
        vpr_times  = [r.get("vlookup_elapsed") for r in batch_results]
        peak_rams  = [r.get("peak_ram")        for r in batch_results]
        peak_cpus  = [r.get("peak_cpu")        for r in batch_results]

        def _idx_best(vals):
            valid = [(v, i) for i, v in enumerate(vals) if v is not None]
            return min(valid, key=lambda x: x[0])[1] if valid else -1

        def _idx_worst(vals):
            valid = [(v, i) for i, v in enumerate(vals) if v is not None]
            return max(valid, key=lambda x: x[0])[1] if valid else -1

        bi_open, wi_open = _idx_best(open_times), _idx_worst(open_times)
        bi_vpr,  wi_vpr  = _idx_best(vpr_times),  _idx_worst(vpr_times)
        bi_ram,  wi_ram  = _idx_best(peak_rams),  _idx_worst(peak_rams)
        bi_cpu,  wi_cpu  = _idx_best(peak_cpus),  _idx_worst(peak_cpus)

        def _cell(val, i, bi, wi, fmt=".2f"):
            if val is None:
                return "<td>—</td>"
            style = (' style="background:#d4edda;font-weight:bold;color:#155724"' if i == bi
                     else ' style="background:#f8d7da;color:#721c24"' if i == wi else "")
            return f"<td{style}>{val:{fmt}}</td>"

        rows_html = ""
        for i, r in enumerate(batch_results):
            status  = "✅ OK" if r.get("success") else f"❌ {html.escape(str(r.get('error',''))[:50])}"
            rows_html += (
                f"<tr><td>{html.escape(str(r['version']))}</td>"
                + _cell(r.get("open_elapsed"),    i, bi_open, wi_open, ".2f")
                + _cell(r.get("vlookup_elapsed"), i, bi_vpr,  wi_vpr,  ".2f")
                + _cell(r.get("peak_ram"),        i, bi_ram,  wi_ram,  ".0f")
                + _cell(r.get("peak_cpu"),        i, bi_cpu,  wi_cpu,  ".0f")
                + f"<td>{status}</td></tr>\n"
            )

        labels_json = self._json_for_script(versions, ensure_ascii=False)
        open_json   = self._json_for_script(open_times)
        vpr_json    = self._json_for_script(vpr_times)
        ram_json    = self._json_for_script(peak_rams)

        return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Batch-отчёт R7-Office</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body{{font-family:Arial,sans-serif;margin:0;padding:20px;background:#f5f6fa;color:#333}}
  h1{{color:#2c3e50;margin-bottom:2px}}
  h2{{color:#2c3e50;margin-top:28px;margin-bottom:10px}}
  .subtitle{{color:#888;font-size:.9em;margin-bottom:16px}}
  .legend-note{{font-size:.82em;color:#555;margin-bottom:12px;padding:8px 12px;
    background:#fff;border-radius:6px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  .charts{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:28px}}
  @media(max-width:700px){{.charts{{grid-template-columns:1fr}}}}
  .chart-box{{background:#fff;padding:16px;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.12)}}
  table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;
    overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.12)}}
  th{{background:#2c3e50;color:#fff;padding:9px 12px;text-align:left;font-size:.85em}}
  td{{padding:8px 12px;border-bottom:1px solid #eee;font-size:.88em;text-align:center}}
  td:first-child{{text-align:left;font-weight:500}}
  tr:last-child td{{border-bottom:none}}
  tr:hover td{{background:#f8f9ff}}
  .pdf-btn{{position:fixed;top:16px;right:16px;padding:8px 18px;background:#2c3e50;
    color:#fff;border:none;border-radius:6px;font-size:.9em;cursor:pointer;
    box-shadow:0 2px 6px rgba(0,0,0,.25);z-index:1000}}
  .pdf-btn:hover{{background:#34495e}}
  @media print{{
    .pdf-btn{{display:none}}
    body{{background:#fff}}
    canvas,.chart-box,table{{page-break-inside:avoid}}
    h1,h2,h3{{page-break-after:avoid}}
  }}
</style>
</head>
<body>
<button class="pdf-btn" onclick="window.print()">📄 Сохранить как PDF</button>
<h1>Batch-отчёт R7-Office</h1>
<div class="subtitle">Сформировано: {ts_display} &nbsp;|&nbsp; Протестировано версий: {n}</div>
<div class="legend-note">
  <span style="background:#d4edda;color:#155724;font-weight:bold;padding:2px 6px;border-radius:3px">Зелёный</span>
  — лучший результат в столбце &nbsp;&nbsp;
  <span style="background:#f8d7da;color:#721c24;padding:2px 6px;border-radius:3px">Красный</span>
  — худший
</div>

<h2>Сводная таблица</h2>
<table>
<thead><tr>
  <th>Версия</th><th>Открытие (сек)</th><th>ВПР (сек)</th>
  <th>Пик RAM (МБ)</th><th>Пик CPU (%)</th><th>Статус</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>

<h2>Графики</h2>
<div class="charts">
  <div class="chart-box"><canvas id="openChart"></canvas></div>
  <div class="chart-box"><canvas id="vprChart"></canvas></div>
  <div class="chart-box"><canvas id="ramChart"></canvas></div>
</div>

<script>
const labels = {labels_json};
const defOpts = t => ({{
  responsive:true,
  plugins:{{legend:{{display:false}},title:{{display:true,text:t}}}},
  scales:{{y:{{beginAtZero:false}}}}
}});
new Chart(document.getElementById('openChart'),{{type:'bar',
  data:{{labels,datasets:[{{label:'сек',data:{open_json},backgroundColor:'#3498db',borderRadius:4}}]}},
  options:defOpts('Открытие файла (сек)')}});
new Chart(document.getElementById('vprChart'),{{type:'bar',
  data:{{labels,datasets:[{{label:'сек',data:{vpr_json},backgroundColor:'#27ae60',borderRadius:4}}]}},
  options:defOpts('Функция ВПР (сек)')}});
new Chart(document.getElementById('ramChart'),{{type:'bar',
  data:{{labels,datasets:[{{label:'МБ',data:{ram_json},backgroundColor:'#e67e22',borderRadius:4}}]}},
  options:defOpts('Пик RAM (МБ)')}});
</script>
</body>
</html>"""

    # --- Хранилище последних параметров тестового файла ---
    _LAST_PARAMS_FILE = "last_test_params.json"

    def _load_last_params(self):
        """Returns dict with last used rows/cols/filename, or defaults."""
        path = BASE_DIR / self._LAST_PARAMS_FILE
        try:
            if path.exists():
                data = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    return data
        except Exception:
            pass
        return {"rows": 50000, "cols": 50, "filename": "test_data_50000x50.xlsx"}

    def _save_last_params(self, rows, cols, filename):
        """Persists rows/cols/filename to last_test_params.json."""
        path = BASE_DIR / self._LAST_PARAMS_FILE
        try:
            path.write_text(
                json.dumps({"rows": rows, "cols": cols, "filename": filename},
                           indent=2, ensure_ascii=False),
                encoding="utf-8"
            )
        except Exception:
            pass

    def compare_file_sizes(self):
        """Opens the test-file generation dialog with 4 separate action buttons."""
        last = self._load_last_params()

        dlg = tk.Toplevel(self.root)
        dlg.transient(self.root)
        dlg.configure(bg=COLORS["bg"])
        dlg.title("Генерация тестового файла")
        dlg.resizable(False, False)
        dlg.grab_set()

        PAD = {"padx": 16, "pady": 5}

        # ── Строки ──────────────────────────────────────────────────────────
        ttk.Label(dlg, text="Количество строк:").grid(
            row=0, column=0, sticky=tk.W, **PAD)
        rows_var = tk.StringVar(value=str(last.get("rows", 50000)))
        rows_entry = ttk.Entry(dlg, textvariable=rows_var, width=14)
        rows_entry.grid(row=0, column=1, sticky=tk.W, **PAD)
        ttk.Label(dlg, text="(1 000 – 1 000 000)", foreground=COLORS["text_secondary"]).grid(
            row=0, column=2, sticky=tk.W, padx=(0, 16))

        # ── Столбцы ─────────────────────────────────────────────────────────
        ttk.Label(dlg, text="Количество столбцов:").grid(
            row=1, column=0, sticky=tk.W, **PAD)
        cols_var = tk.StringVar(value=str(last.get("cols", 50)))
        cols_entry = ttk.Entry(dlg, textvariable=cols_var, width=14)
        cols_entry.grid(row=1, column=1, sticky=tk.W, **PAD)
        ttk.Label(dlg, text="(1 – 100)", foreground=COLORS["text_secondary"]).grid(
            row=1, column=2, sticky=tk.W, padx=(0, 16))

        ttk.Separator(dlg, orient=tk.HORIZONTAL).grid(
            row=2, column=0, columnspan=3, sticky=tk.EW, padx=16, pady=8)

        # ── Перезаписать ─────────────────────────────────────────────────────
        overwrite_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(dlg, text="Перезаписать если существует",
                        variable=overwrite_var).grid(
            row=3, column=0, columnspan=3, sticky=tk.W, padx=16, pady=2)

        # ── Имя файла ────────────────────────────────────────────────────────
        ttk.Label(dlg, text="Имя файла:").grid(row=4, column=0, sticky=tk.W, **PAD)
        filename_var = tk.StringVar(value=last.get("filename", "test_data_50000x50.xlsx"))
        filename_entry = ttk.Entry(dlg, textvariable=filename_var, width=36)
        filename_entry.grid(row=4, column=1, columnspan=2, sticky=tk.EW,
                            padx=(0, 16), pady=5)

        # Авто-имя при смене размеров; сбрасывается при ручном редактировании
        _auto_name = [True]
        _ext_path  = [None]   # полный путь из filedialog

        def _on_dim_change(*_):
            if _auto_name[0]:
                try:
                    filename_var.set(
                        f"test_data_{int(rows_var.get())}x{int(cols_var.get())}.xlsx")
                    _ext_path[0] = None
                except ValueError:
                    pass

        def _on_filename_edit(*_):
            try:
                expected = (
                    f"test_data_{int(rows_var.get())}x{int(cols_var.get())}.xlsx")
            except ValueError:
                expected = ""
            _auto_name[0] = (filename_var.get() == expected)
            _ext_path[0]  = None

        rows_var.trace_add("write", _on_dim_change)
        cols_var.trace_add("write", _on_dim_change)
        filename_var.trace_add("write", _on_filename_edit)

        ttk.Separator(dlg, orient=tk.HORIZONTAL).grid(
            row=5, column=0, columnspan=3, sticky=tk.EW, padx=16, pady=8)

        # ── Кнопки 2×2 ───────────────────────────────────────────────────────
        bf = ttk.Frame(dlg)
        bf.grid(row=6, column=0, columnspan=3, sticky=tk.EW, padx=16)
        bf.columnconfigure(0, weight=1)
        bf.columnconfigure(1, weight=1)

        btn_create = ttk.Button(bf, text="1. Создать файл")
        btn_choose = ttk.Button(bf, text="2. Выбрать файл")
        btn_test   = ttk.Button(bf, text="3. Протестировать")
        btn_cancel = ttk.Button(bf, text="4. Отмена", command=dlg.destroy)

        btn_create.grid(row=0, column=0, sticky=tk.EW, padx=(0, 3), pady=(0, 5))
        btn_choose.grid(row=0, column=1, sticky=tk.EW, padx=(3, 0), pady=(0, 5))
        btn_test  .grid(row=1, column=0, sticky=tk.EW, padx=(0, 3))
        btn_cancel.grid(row=1, column=1, sticky=tk.EW, padx=(3, 0))

        # ── Статус ───────────────────────────────────────────────────────────
        status_var = tk.StringVar(value="Статус: Готов")
        status_lbl = ttk.Label(dlg, textvariable=status_var, anchor=tk.W,
                               foreground=COLORS["text_secondary"])
        status_lbl.grid(row=7, column=0, columnspan=3, sticky=tk.EW,
                        padx=16, pady=(10, 14))

        # ── Вспомогательные функции ──────────────────────────────────────────
        _action_btns = [btn_create, btn_test]

        def _set_status(text, color=COLORS["text_secondary"]):
            def _do():
                try:
                    status_var.set(f"Статус: {text}")
                    status_lbl.config(foreground=color)
                except tk.TclError:
                    pass
            try:
                dlg.after(0, _do)
            except tk.TclError:
                pass

        def _lock():
            def _do():
                try:
                    for b in _action_btns:
                        b.config(state="disabled")
                except tk.TclError:
                    pass
            try:
                dlg.after(0, _do)
            except tk.TclError:
                pass

        def _unlock():
            def _do():
                try:
                    for b in _action_btns:
                        b.config(state="normal")
                except tk.TclError:
                    pass
            try:
                dlg.after(0, _do)
            except tk.TclError:
                pass

        def _validate_dims():
            try:
                r = int(rows_var.get())
                assert 1_000 <= r <= 1_000_000
            except (ValueError, AssertionError):
                messagebox.showwarning(
                    "Ошибка", "Строки: от 1 000 до 1 000 000.", parent=dlg)
                rows_entry.focus_set()
                return None, None
            try:
                c = int(cols_var.get())
                assert 1 <= c <= 100
            except (ValueError, AssertionError):
                messagebox.showwarning(
                    "Ошибка", "Столбцы: от 1 до 100.", parent=dlg)
                cols_entry.focus_set()
                return None, None
            return r, c

        def _resolve_path():
            if _ext_path[0]:
                return Path(_ext_path[0])
            fname = filename_var.get().strip()
            if not fname:
                return None
            if not fname.endswith(".xlsx"):
                fname += ".xlsx"
            return self.test_files_folder / fname

        # ── Кнопка 1: только создать файл ────────────────────────────────────
        def on_create():
            r, c = _validate_dims()
            if r is None:
                return
            if _ext_path[0]:
                messagebox.showwarning(
                    "Внимание",
                    "Файл выбран через диалог — кнопка «Создать файл» работает\n"
                    "только с именем в поле «Имя файла».\n"
                    "Введите имя файла вручную или очистите поле.",
                    parent=dlg)
                return
            fname = filename_var.get().strip()
            if not fname or not re.fullmatch(r"[A-Za-z0-9_.]+", fname):
                messagebox.showwarning(
                    "Ошибка",
                    "Имя файла: только латиница, цифры, '_' и '.'.",
                    parent=dlg)
                filename_entry.focus_set()
                return
            if not fname.endswith(".xlsx"):
                fname += ".xlsx"
                filename_var.set(fname)
            file_path = self.test_files_folder / fname
            if file_path.exists() and not overwrite_var.get():
                _set_status(f"⚠️ Файл уже существует: {fname}", "#e67e22")
                self.add_test_log(f"⚠️ Файл уже существует: {file_path}")
                return
            self._save_last_params(r, c, fname)
            _lock()
            _set_status("⏳ Создание файла...", "#2980b9")

            def _worker():
                try:
                    self._generate_custom_test_file(r, c, file_path)
                    self.add_test_log(
                        f"📊 Создан тестовый файл: {fname} ({r} строк, {c} столбцов)")
                    _set_status(f"✅ Файл создан: {fname}", "#27ae60")
                except Exception as e:
                    self.add_test_log(f"❌ Ошибка создания файла: {e}")
                    _set_status(f"❌ Ошибка: {e}", "#e74c3c")
                finally:
                    _unlock()

            threading.Thread(target=_worker, daemon=True).start()

        # ── Кнопка 2: выбрать любой xlsx ─────────────────────────────────────
        def on_choose():
            path = filedialog.askopenfilename(
                parent=dlg,
                title="Выбрать xlsx-файл для тестирования",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if path:
                _ext_path[0]  = path
                _auto_name[0] = False
                filename_var.set(path)
                self.add_test_log(f"📁 Выбран файл: {path}")
                _set_status(f"📁 Выбран файл: {Path(path).name}", "#2980b9")

        # ── Кнопка 3: только тестирование ────────────────────────────────────
        def on_test():
            file_path = _resolve_path()
            if not file_path:
                _set_status("❌ Укажите имя или путь к файлу", "#e74c3c")
                return
            if not file_path.exists():
                msg = f"❌ Файл не найден: {file_path.name}"
                _set_status(msg, "#e74c3c")
                self.add_test_log(msg)
                return
            try:
                r, c = int(rows_var.get()), int(cols_var.get())
            except ValueError:
                r, c = 0, 0
            self._save_last_params(r, c, file_path.name)
            _lock()
            _set_status("⏳ Тестирование...", "#2980b9")

            def _done(success):
                _set_status(
                    "✅ Тест завершён" if success else "❌ Тест завершён с ошибкой",
                    "#27ae60" if success else "#e74c3c")
                _unlock()

            threading.Thread(
                target=self._worker_run_test,
                args=(file_path, r, c, _done),
                daemon=True
            ).start()

        btn_create.config(command=on_create)
        btn_choose.config(command=on_choose)
        btn_test  .config(command=on_test)

        dlg.columnconfigure(1, weight=1)
        rows_entry.focus_set()
        dlg.wait_window()

    def _worker_run_test(self, file_path, rows, cols, done_cb):
        """Worker: kills stale R7 instances, clears cache, opens file, runs VPR, shows report."""
        success = False
        try:
            # ----- 1. Завершаем старые процессы Р7 ---------------------------------------
            killed = self._kill_r7_processes_for_test()
            if killed:
                self.add_test_log(f"🔄 Завершено {killed} процессов Р7-Офис")
                time.sleep(2)
            else:
                self.add_test_log("ℹ️ Активных процессов Р7-Офис не найдено")

            # ----- 2. Очистка кеша -------------------------------------------------------
            cleared = self._clear_r7_cache()
            if cleared:
                self.add_test_log(f"🧹 Очищено {cleared} временных объектов Р7 из %TEMP%")

            # ----- 3. Реальное количество строк ------------------------------------------
            real_rows = self._get_xlsx_row_count(file_path)
            if real_rows is not None:
                self.add_test_log(f"📊 Реальное количество строк в файле: {real_rows:,}")
            else:
                real_rows = rows

            # ----- 4. Поиск пути к Р7-Офис -----------------------------------------------
            r7_path = self._find_r7_path()
            if not r7_path:
                self.add_test_log("❌ Р7-Офис не найден.")
                return

            # ----- 5. Запуск и ожидание окна --------------------------------------------
            self.add_test_log(f"⏳ Запуск теста на файле {file_path.name}")
            debug_args = self._prepare_webdriver_launch(filename_hint=file_path.name)
            open_start = time.time()
            subprocess.Popen([r7_path, str(file_path), *debug_args], shell=True)

            def _find_hwnd():
                found = [None]
                if WIN32_OK:
                    stem = file_path.stem[:12]
                    def _cb(h, _):
                        t = win32gui.GetWindowText(h)
                        if stem in t or "Р7-Офис" in t:
                            found[0] = h
                    win32gui.EnumWindows(_cb, None)
                return found[0]

            deadline = time.time() + 60
            hwnd = None
            while time.time() < deadline:
                hwnd = _find_hwnd()
                if hwnd:
                    break
                time.sleep(0.5)

            if not hwnd:
                self.add_test_log("⚠️ Окно Р7 не найдено, продолжаем без фокуса")

            # ----- 6. Фокус и разворот ---------------------------------------------------
            # Засекаем отдельно и вычитаем: подготовка окна не относится к
            # скорости открытия файла.
            _setup_start = time.time()
            if WIN32_OK and hwnd:
                try:
                    win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
                    win32gui.SetForegroundWindow(hwnd)
                    time.sleep(0.3)
                except Exception:
                    pass
            _setup_elapsed = time.time() - _setup_start

            # ----- 7. Динамическое ожидание загрузки -------------------------------------
            data_ready   = self._wait_until_r7_ready(_find_hwnd, timeout=120)
            open_elapsed = time.time() - open_start - _setup_elapsed
            self.add_test_log(
                f"✅ Файл открыт за {open_elapsed:.2f} сек "
                f"({'данные загружены' if data_ready else 'таймаут — возможна частичная загрузка'})"
            )

            # ----- 8. ВПР-бенчмарк на всех строках --------------------------------------
            vlookup_elapsed = None
            vlookup_error   = None
            vlookup_rows    = 0

            if PYAUTOGUI_OK and pyperclip:
                try:
                    # Переходим в C2: первая свободная колонка после ID и Name
                    pyautogui.hotkey('ctrl', 'Home')
                    self._pace(self.OP_KEY_PACE)
                    pyautogui.press('right')   # → B1
                    pyautogui.press('right')   # → C1
                    pyautogui.press('down')    # → C2

                    # Вставляем формулу ВПР в C2
                    pyperclip.copy('=VLOOKUP(A2,A:B,2,FALSE)')
                    pyautogui.hotkey('ctrl', 'v')
                    self._pace(self.OP_KEY_PACE)

                    # Замер, как в остальных тестах: секундомер останавливается,
                    # когда Р7 освободился, минус собственные паузы.
                    self._paced_total = 0.0
                    vstart = time.time()
                    pyautogui.press('enter')

                    # Возвращаемся в C2 и заполняем формулой весь столбец
                    pyautogui.hotkey('ctrl', 'Home')
                    pyautogui.press('right')
                    pyautogui.press('right')
                    pyautogui.press('down')                # C2
                    pyautogui.hotkey('ctrl', 'shift', 'down')  # выделяем до конца данных
                    pyautogui.hotkey('ctrl', 'd')              # заполняем вниз

                    _done_ts, _status = self._wait_operation_done(_find_hwnd)
                    _end = _done_ts if _done_ts is not None else time.time()
                    vlookup_elapsed = round(
                        max(0.0, _end - vstart - self._paced_total), 3)
                    if _status == "below_floor":
                        self.add_test_log(
                            "⚠️ ВПР завершился быстрее порога измерения — "
                            "результат ненадёжен")
                    vlookup_rows    = real_rows
                    self.add_test_log(
                        f"✅ ВПР по {real_rows:,} строкам завершён за {vlookup_elapsed:.2f} сек"
                    )
                except Exception as e:
                    vlookup_error = str(e)
                    self.add_test_log(f"⚠️ Ошибка ВПР: {e}")
            else:
                self.add_test_log("⚠️ pyautogui/pyperclip недоступны — ВПР пропущен")

            # ----- 9. Закрытие Р7 --------------------------------------------------------
            self._close_r7_gracefully(hwnd)

            # ----- 10. Отчёт -------------------------------------------------------------
            file_size_mb = (round(file_path.stat().st_size / (1024 ** 2), 2)
                            if file_path.exists() else None)
            self._show_custom_test_report({
                "filename":        file_path.name,
                "rows":            rows,
                "cols":            cols,
                "real_rows":       real_rows,
                "vlookup_rows":    vlookup_rows,
                "file_size_mb":    file_size_mb,
                "open_elapsed":    round(open_elapsed, 3),
                "vlookup_elapsed": vlookup_elapsed,
                "vlookup_error":   vlookup_error,
                "cache_cleared":   cleared > 0,
                "data_ready":      data_ready,
                "timestamp":       datetime.now().strftime("%d.%m.%Y %H:%M"),
            })
            success = True
        except Exception as e:
            self.add_test_log(f"❌ Ошибка тестирования: {e}")
        finally:
            self._close_webdriver_connector()
            done_cb(success)

    def _kill_r7_processes_for_test(self):
        """Kills all R7-Office processes. Returns count killed."""
        if not PSUTIL_OK:
            return 0
        search = ("editors_helper", "desktopeditors", "r7officemain", "r7office")
        killed = 0
        try:
            for proc in psutil.process_iter(["name", "pid"]):
                try:
                    name = (proc.info.get("name") or "").lower()
                    if any(s in name for s in search):
                        proc.kill()
                        killed += 1
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except Exception:
            pass
        return killed

    def _clear_r7_cache(self):
        """Removes R7-Office temp items from %%TEMP%%. Returns count removed."""
        cleared = 0
        temp_dir = Path(os.environ.get("TEMP", ""))
        if not temp_dir.exists():
            return 0
        # "R7*" не нужен отдельно от "r7*": glob на Windows регистронезависим,
        # так что оба паттерна и так матчат одни и те же файлы — второй
        # проход просто не находит ничего (первый уже всё удалил).
        for pat in ("r7*", "editors*"):
            for item in temp_dir.glob(pat):
                try:
                    if item.is_dir():
                        shutil.rmtree(item, ignore_errors=True)
                    else:
                        item.unlink(missing_ok=True)
                    cleared += 1
                except Exception:
                    pass
        return cleared

    def _get_xlsx_row_count(self, path):
        """Returns data row count (excluding header row) via openpyxl read-only, or None."""
        if not EXCEL_OK:
            return None
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(str(path), read_only=True, data_only=True)
            count = max(0, (wb.active.max_row or 1) - 1)
            wb.close()
            return count
        except Exception as e:
            self.add_test_log(f"⚠️ Не удалось прочитать количество строк: {e}")
            return None

    def _show_custom_test_report(self, result):
        """Builds and opens an HTML report for a single custom-file benchmark."""
        # filename может прийти из диалога "Выбрать файл" — это произвольный
        # путь на диске пользователя, не сгенерированное этим инструментом имя.
        fname        = html.escape(result["filename"])
        cols         = result.get("cols") or 0
        real_rows    = result.get("real_rows") or result.get("rows") or 0
        vlookup_rows = result.get("vlookup_rows") or 0
        size_mb      = f"{result['file_size_mb']:.2f}" if result.get("file_size_mb") else "—"
        open_t       = f"{result['open_elapsed']:.3f}"
        vlook_t      = (f"{result['vlookup_elapsed']:.3f}"
                        if result.get("vlookup_elapsed") is not None else "—")
        vlook_err    = html.escape(result.get("vlookup_error") or "")
        ts           = result["timestamp"]
        cache_ok     = result.get("cache_cleared", False)
        data_ready   = result.get("data_ready", None)

        bar_data   = self._json_for_script([
            result["open_elapsed"],
            result["vlookup_elapsed"] if result.get("vlookup_elapsed") is not None else 0,
        ])
        bar_labels = self._json_for_script(
            ["Открытие файла", f"ВПР ({vlookup_rows:,} строк)".replace(",", " ")])

        # Баннер предупреждения если данные могут быть не загружены
        warn_html = ""
        if data_ready is False:
            warn_html = (
                '<div class="warn-banner">⚠️ Данные могут быть загружены не полностью '
                '(сработал таймаут ожидания). Результаты могут быть занижены.</div>'
            )

        # Бейдж очистки кеша
        cache_badge = (
            '<span class="badge badge-ok">🧹 Кеш очищен перед тестом</span>'
            if cache_ok else
            '<span class="badge badge-warn">⚠️ psutil недоступен — кеш не очищался</span>'
        )

        # Статус загрузки данных
        if data_ready is True:
            load_status = '<td class="ok">✅ Данные загружены</td>'
        elif data_ready is False:
            load_status = '<td class="err">⏰ Таймаут</td>'
        else:
            load_status = '<td>—</td>'

        html_content = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Тест: {fname}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body{{font-family:Arial,sans-serif;padding:24px;background:#f5f6fa;color:#333}}
  h1{{color:#2c3e50;margin-bottom:4px}}
  .subtitle{{color:#888;font-size:.9em;margin-bottom:14px}}
  .warn-banner{{background:#fff3cd;border:1px solid #ffc107;border-radius:6px;
    padding:10px 16px;margin-bottom:14px;color:#856404;font-size:.9em}}
  .badge{{display:inline-block;padding:4px 12px;border-radius:12px;
    font-size:.8em;font-weight:bold;margin-bottom:16px}}
  .badge-ok{{background:#d4edda;color:#155724}}
  .badge-warn{{background:#fff3cd;color:#856404}}
  .info-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));
    gap:10px;margin-bottom:20px}}
  .info-item{{background:#fff;padding:10px 14px;border-radius:8px;
    box-shadow:0 1px 3px rgba(0,0,0,.1)}}
  .info-label{{font-size:.75em;color:#888;text-transform:uppercase}}
  .info-value{{font-weight:bold;font-size:1.05em;margin-top:3px}}
  .chart-box{{background:#fff;padding:20px;border-radius:8px;
    box-shadow:0 1px 4px rgba(0,0,0,.12);max-width:560px;margin-bottom:24px}}
  table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;
    box-shadow:0 1px 4px rgba(0,0,0,.12);overflow:hidden}}
  th{{background:#2c3e50;color:#fff;padding:10px 14px;text-align:left;font-size:.85em}}
  td{{padding:9px 14px;border-bottom:1px solid #eee;font-size:.9em}}
  .ok{{color:#27ae60;font-weight:bold}} .err{{color:#e74c3c;font-weight:bold}}
</style>
</head>
<body>
<h1>Тест производительности: {fname}</h1>
<div class="subtitle">{ts}</div>
{warn_html}
{cache_badge}

<div class="info-grid">
  <div class="info-item">
    <div class="info-label">Файл</div>
    <div class="info-value" style="font-size:.9em;word-break:break-all">{fname}</div>
  </div>
  <div class="info-item">
    <div class="info-label">Строк в файле</div>
    <div class="info-value">{real_rows:,}</div>
  </div>
  <div class="info-item">
    <div class="info-label">Столбцов</div>
    <div class="info-value">{cols}</div>
  </div>
  <div class="info-item">
    <div class="info-label">Размер файла</div>
    <div class="info-value">{size_mb} МБ</div>
  </div>
  <div class="info-item">
    <div class="info-label">Открытие файла</div>
    <div class="info-value">{open_t} сек</div>
  </div>
  <div class="info-item">
    <div class="info-label">ВПР ({vlookup_rows:,} строк)</div>
    <div class="info-value {'err' if vlook_err else 'ok'}">{vlook_t} {'⚠ ' + vlook_err if vlook_err else 'сек'}</div>
  </div>
</div>

<div class="chart-box">
  <canvas id="barChart"></canvas>
</div>

<table>
<thead>
  <tr><th>Операция</th><th>Строк</th><th>Время (сек)</th><th>Статус</th></tr>
</thead>
<tbody>
<tr>
  <td>Открытие файла</td>
  <td>{real_rows:,}</td>
  <td>{open_t}</td>
  {load_status}
</tr>
<tr>
  <td>ВПР (VLOOKUP)</td>
  <td>{vlookup_rows:,}</td>
  <td>{vlook_t}</td>
  <td class="{'err' if vlook_err else 'ok'}">{'⚠️ ' + vlook_err if vlook_err else '✅ OK'}</td>
</tr>
</tbody>
</table>

<script>
new Chart(document.getElementById('barChart'), {{
  type: 'bar',
  data: {{
    labels: {bar_labels},
    datasets: [{{
      label: 'Время (сек)',
      data: {bar_data},
      backgroundColor: ['#3498db', '#27ae60'],
      borderRadius: 4,
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{
      legend: {{display: false}},
      title: {{display: true, text: 'Время выполнения операций (сек)'}}
    }},
    scales: {{y: {{beginAtZero: true}}}}
  }}
}});
</script>
</body>
</html>"""

        ts_file  = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = self.reports_folder / f"custom_test_{ts_file}.html"
        try:
            out_path.write_text(html_content, encoding="utf-8")
            self.add_test_log(f"📊 Отчёт готов: {out_path.name}")
            webbrowser.open(str(out_path))
        except Exception as e:
            self.add_test_log(f"⚠️ Ошибка записи отчёта: {e}")

    # ── Генератор фикстур с профилями нагрузки (этап 2, M2) ────────────────
    #
    # ПОЧЕМУ ПРОФИЛИ, А НЕ ПРОСТО «БОЛЬШЕ СТРОК»: «жирный текст» (вставка
    # большого объёма символов) нагружает путь ввода и парсер текста, но не
    # трогает то, что в реальности чаще всего медленно — таблицу стилей,
    # движок пересчёта формул, парсинг .xlsx при открытии. Один и тот же
    # объём данных по-разному нагружает Р7 в зависимости от ЧЕГО он состоит
    # (см. отчёт по нагрузочному тестированию, 25.08.2026, раздел 06).
    #
    # ПОЧЕМУ БЕЗ ОДНОРОДНОСТИ: сгенерированные значения намеренно РАЗНЫЕ
    # (rnd.randint / f-строка с индексом), а не повторяющаяся константа —
    # однородный текст даёт быстрый путь дедупликации в sharedStrings и
    # завышает результат по сравнению с реальным документом (та же ловушка,
    # что и у «жирного текста», см. отчёт).
    FIXTURE_PROFILES = ("flat", "formula", "styled", "mixed")

    def _generate_fixture(self, path, rows=100_000, profile="flat", seed=42, cols=6):
        """Генерирует .xlsx для нагрузочного тестирования по одному из
        FIXTURE_PROFILES.

        Args:
            path: Куда сохранить файл.
            rows: Число строк данных (без строки заголовка).
            profile: "flat" (числа/текст — парсер и аллокация ячеек),
                "formula" (цепочка формул — движок пересчёта),
                "styled" (уникальный стиль почти на каждой строке —
                таблица стилей), "mixed" (всё сразу).
            seed: Сид генератора случайных чисел. Одно и то же значение
                (по умолчанию 42) даёт БУКВАЛЬНО одинаковый файл на
                повторном вызове — обязательное условие для сравнения
                версий Р7 на одинаковой нагрузке, а не на случайно разных
                файлах одного размера.
            cols: Число столбцов данных. Используется только профилем
                "flat" — у "formula"/"styled"/"mixed" набор колонок
                фиксирован их структурой (цепочка формул/стиль/оба сразу).

        Returns:
            Path: тот же path — для цепочки вызовов.

        Raises:
            RuntimeError: openpyxl не установлен.
            ValueError: profile не входит в FIXTURE_PROFILES.
        """
        if not EXCEL_OK:
            raise RuntimeError("openpyxl не установлен")
        if profile not in self.FIXTURE_PROFILES:
            raise ValueError(
                f"неизвестный профиль фикстуры: {profile!r} "
                f"(допустимо: {', '.join(self.FIXTURE_PROFILES)})")

        rnd = random.Random(seed)
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Данные")

        {
            "flat": self._fixture_fill_flat,
            "formula": self._fixture_fill_formula,
            "styled": self._fixture_fill_styled,
            "mixed": self._fixture_fill_mixed,
        }[profile](ws, rows, cols, rnd)

        wb.save(str(path))
        return path

    @staticmethod
    def _fixture_fill_flat(ws, rows, cols, rnd):
        """flat: числа + короткий разнородный текст. Нагружает парсер
        значений и аллокацию ячеек — не движок пересчёта и не стили."""
        ws.append(["ID", "Name", "Qty"] + [f"Col{c}" for c in range(4, cols + 1)])
        for i in range(1, rows + 1):
            row = [i, f"Позиция {i} {rnd.randint(1, 999999)}", rnd.randint(1, 10_000)]
            for _ in range(4, cols + 1):
                row.append(round(rnd.random() * 1000, 2))
            ws.append(row)

    @staticmethod
    def _fixture_fill_formula(ws, rows, cols, rnd):
        """formula: колонка "Chain" каждой строки ссылается на "Chain"
        предыдущей — граф зависимостей длиной rows, который движок
        пересчёта не может распараллелить (в отличие от rows независимых
        формул).

        Ссылка на предыдущую строку — C{i}, НЕ C{i-1}: заголовок занимает
        строку листа 1, данные index i лежат в строке i+1, поэтому «Chain»
        предыдущего index (i-1) — это строка листа i, а не i-1. C{i-1} для
        i=2 указал бы на C1 — строку заголовка, а не на данные (ловилось
        тестом test_formula_profile_chains_to_previous_row)."""
        ws.append(["ID", "Base", "Chain", "Qty"])
        ws.append([1, rnd.randint(1, 1000), rnd.randint(1, 1000), rnd.randint(1, 10_000)])
        for i in range(2, rows + 1):
            ws.append([i, rnd.randint(1, 1000),
                      f"=C{i}*1.02+{rnd.randint(1, 50)}",
                      rnd.randint(1, 10_000)])

    _FIXTURE_STYLE_PALETTE = ("FFECE0", "E0F0FF", "E8FFE0", "FFF6D5", "F0E0FF")

    @classmethod
    def _fixture_fill_styled(cls, ws, rows, cols, rnd):
        """styled: у почти каждой строки свой шрифт/заливка — раздувает
        таблицу стилей (styles.xml), а не таблицу данных. Именно это НЕ
        нагружает «жирный текст» и обычные flat-фикстуры."""
        ws.append(["ID", "Name", "Status"])
        for i in range(1, rows + 1):
            color = cls._FIXTURE_STYLE_PALETTE[i % len(cls._FIXTURE_STYLE_PALETTE)]
            name_cell = WriteOnlyCell(ws, value=f"Позиция {i}")
            name_cell.font = Font(bold=(i % 3 == 0), size=9 + (i % 4))
            name_cell.fill = PatternFill(start_color=color, end_color=color,
                                         fill_type="solid")
            ws.append([i, name_cell, rnd.choice(["ок", "ждём", "отказ"])])

    @classmethod
    def _fixture_fill_mixed(cls, ws, rows, cols, rnd):
        """mixed: числа + формулы + стили через строку — ближе всего к
        реальному документу, где узкое место заранее не известно.

        Ссылка D{i} (не D{i-1}) по той же причине, что и в formula-профиле:
        заголовок в строке листа 1 сдвигает данные index i в строку i+1."""
        ws.append(["ID", "Name", "Qty", "Formula", "Status"])
        ws.append([1, "Позиция 1", rnd.randint(1, 10_000), rnd.randint(1, 10_000), "ок"])
        for i in range(2, rows + 1):
            name_cell = WriteOnlyCell(ws, value=f"Позиция {i}")
            if i % 5 == 0:
                color = cls._FIXTURE_STYLE_PALETTE[i % len(cls._FIXTURE_STYLE_PALETTE)]
                name_cell.font = Font(bold=True)
                name_cell.fill = PatternFill(start_color=color, end_color=color,
                                             fill_type="solid")
            ws.append([i, name_cell, rnd.randint(1, 10_000),
                      f"=D{i}*2+{rnd.randint(1, 20)}",
                      rnd.choice(["ок", "ждём", "отказ"])])

    def _generate_custom_test_file(self, rows, cols, path):
        """Creates an xlsx file with rows×cols of test data using openpyxl.

        write_only=True: обычный Workbook() держит все объекты ячеек в
        памяти до save(). При заявленном максимуме 1 000 000 строк × 100
        столбцов это 100 млн объектов ячеек одновременно. В write_only-режиме
        openpyxl пишет каждую добавленную строку сразу в поток архива и не
        накапливает их — единственное отличие в API: лист создаётся через
        wb.create_sheet(), а не берётся готовым через wb.active (write_only
        workbook стартует без единого листа).
        """
        if not EXCEL_OK:
            raise RuntimeError("openpyxl не установлен")
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Лист1")
        # Заголовки
        header = ["ID", "Name"] + [f"Col{i}" for i in range(3, cols + 1)]
        ws.append(header)
        # Данные
        for i in range(1, rows + 1):
            row = [i, f"Item_{i:05d}"]
            for c in range(3, cols + 1):
                row.append(i * c)
            ws.append(row)
        wb.save(str(path))

    # ---------------------- Хеш-суммы дистрибутивов ----------------------
    def check_hashes(self):
        """Entry point for hash verification — creates progress window then spawns worker thread."""
        files = (list(self.distributives_folder.glob("*.msi")) +
                 list(self.distributives_folder.glob("*.exe")))
        if not files:
            messagebox.showwarning("Нет файлов", "В папке Distributives нет файлов для проверки.")
            return

        prog_win = tk.Toplevel(self.root)
        prog_win.transient(self.root)
        prog_win.configure(bg=COLORS["bg"])
        prog_win.title("Вычисление хеш-сумм...")
        prog_win.geometry("440x120")
        prog_win.resizable(False, False)
        prog_win.grab_set()

        lbl_file = ttk.Label(prog_win, text="Подготовка...", wraplength=410, anchor=tk.W)
        lbl_file.pack(pady=(14, 4), padx=15, fill=tk.X)

        progressbar = ttk.Progressbar(prog_win, maximum=len(files), mode="determinate")
        progressbar.pack(fill=tk.X, padx=15)

        lbl_count = ttk.Label(prog_win, text=f"0 / {len(files)}")
        lbl_count.pack(pady=4)

        threading.Thread(
            target=self._hash_worker,
            args=(files, prog_win, progressbar, lbl_file, lbl_count),
            daemon=True,
        ).start()

    def _hash_worker(self, files, prog_win, progressbar, lbl_file, lbl_count):
        """Computes MD5/SHA256 for each file in a background thread, then shows results.

        Args:
            files: List of Path objects to hash.
            prog_win: Progress Toplevel window (destroyed when done).
            progressbar: ttk.Progressbar widget to update.
            lbl_file: Label showing the current filename.
            lbl_count: Label showing N / total progress.
        """
        hashes_json = self.distributives_folder / "hashes.json"
        reference = {}
        if hashes_json.exists():
            try:
                with open(hashes_json, encoding="utf-8") as f:
                    reference = json.load(f)
            except Exception as e:
                self.add_test_log(f"⚠️ Ошибка загрузки hashes.json: {e}")

        def _update_progress(filename, idx):
            lbl_file.config(text=f"Обработка: {filename}")
            progressbar.config(value=idx)
            lbl_count.config(text=f"{idx + 1} / {len(files)}")

        results = []
        for i, path in enumerate(files):
            self.root.after(0, lambda fn=path.name, idx=i: _update_progress(fn, idx))
            try:
                size_mb = path.stat().st_size / (1024 * 1024)
                md5h = hashlib.md5()
                sha256h = hashlib.sha256()
                # 1 МБ вместо прежних 8 КБ — дистрибутивы весят сотни МБ/ГБ,
                # и мелкий чанк умножает накладные расходы на системные вызовы.
                with open(path, "rb") as f:
                    while True:
                        chunk = f.read(1024 * 1024)
                        if not chunk:
                            break
                        md5h.update(chunk)
                        sha256h.update(chunk)
                md5_val = md5h.hexdigest()
                sha256_val = sha256h.hexdigest()

                ref = reference.get(path.name, {})
                if not ref:
                    status, tag = "⚠️ Нет эталона", "no_ref"
                elif (ref.get("md5", "").lower() == md5_val and
                      ref.get("sha256", "").lower() == sha256_val):
                    status, tag = "✅ Совпадает", "ok"
                else:
                    status, tag = "❌ Не совпадает", "fail"

                results.append({
                    "name": path.name,
                    "size": f"{size_mb:.2f}",
                    "md5": md5_val,
                    "sha256": sha256_val,
                    "status": status,
                    "tag": tag,
                })
                self.add_test_log(f"🔐 {path.name}: {status}")

            except Exception as e:
                results.append({
                    "name": path.name,
                    "size": "—",
                    "md5": "ОШИБКА",
                    "sha256": str(e),
                    "status": "❌ Ошибка чтения",
                    "tag": "fail",
                })
                self.add_test_log(f"❌ {path.name}: ошибка чтения — {e}")

        ok_count   = sum(1 for r in results if r["tag"] == "ok")
        fail_count = sum(1 for r in results if r["tag"] == "fail")
        self.add_test_log(
            f"🔐 Проверка завершена: {len(results)} файлов  "
            f"✅ {ok_count} совпадают  ❌ {fail_count} не совпадают"
        )
        self.root.after(0, prog_win.destroy)
        self.root.after(0, lambda: self._show_hash_results(results))

    def _show_hash_results(self, results):
        """Opens a Treeview window with hash results.

        Supports: copy-on-double-click, reference editing/deletion via button and
        context menu, status refresh without re-scanning, and CSV export.

        Args:
            results: List of dicts with keys name, size, md5, sha256, status, tag.
                     Dicts are mutated in place when references are saved/deleted.
        """
        hashes_path = self.distributives_folder / "hashes.json"

        win = tk.Toplevel(self.root)
        win.transient(self.root)
        win.configure(bg=COLORS["bg"])
        win.title("Хеш-суммы дистрибутивов")
        win.geometry("1120x480")
        win.resizable(True, True)

        # ── Treeview ──────────────────────────────────────────────────────────
        columns = ("name", "size", "md5", "sha256", "status")
        tree = ttk.Treeview(win, columns=columns, show="headings", selectmode="browse")

        tree.heading("name",   text="Имя файла")
        tree.heading("size",   text="Размер (МБ)")
        tree.heading("md5",    text="MD5")
        tree.heading("sha256", text="SHA256")
        tree.heading("status", text="Статус")

        tree.column("name",   width=260, anchor=tk.W,      stretch=True)
        tree.column("size",   width=90,  anchor=tk.CENTER, stretch=False)
        tree.column("md5",    width=245, anchor=tk.W,      stretch=False)
        tree.column("sha256", width=370, anchor=tk.W,      stretch=False)
        tree.column("status", width=130, anchor=tk.CENTER, stretch=False)

        tree.tag_configure("ok",     background="#2E4A3A", foreground=COLORS["text"])
        tree.tag_configure("no_ref", background="#4A4326", foreground=COLORS["text"])
        tree.tag_configure("fail",   background="#4A2E2E", foreground=COLORS["text"])

        sb_y = ttk.Scrollbar(win, orient=tk.VERTICAL,   command=tree.yview)
        sb_x = ttk.Scrollbar(win, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        tree.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")

        win.rowconfigure(0, weight=1)
        win.columnconfigure(0, weight=1)

        # row_id → result dict, built while populating the tree
        id_to_result = {}
        for r in results:
            iid = tree.insert("", tk.END,
                              values=(r["name"], r["size"], r["md5"], r["sha256"], r["status"]),
                              tags=(r["tag"],))
            id_to_result[iid] = r

        # ── hashes.json helpers ───────────────────────────────────────────────
        def load_reference():
            """Returns current hashes.json content or an empty dict."""
            if hashes_path.exists():
                try:
                    with open(hashes_path, encoding="utf-8") as f:
                        return json.load(f)
                except Exception:
                    return {}
            return {}

        def save_reference(ref):
            """Persists the reference dict to hashes.json (creates if absent).

            Args:
                ref: Dict mapping filename → {md5, sha256}.
            """
            hashes_path.parent.mkdir(parents=True, exist_ok=True)
            with open(hashes_path, "w", encoding="utf-8") as f:
                json.dump(ref, f, indent=2, ensure_ascii=False)

        def recompute_status(row_data, ref):
            """Returns (status_str, tag) for row_data against current reference.

            Files with read errors keep their error status regardless of the reference.

            Args:
                row_data: Result dict for one file.
                ref: Current hashes.json dict.

            Returns:
                Tuple[str, str]: Human-readable status and Treeview tag name.
            """
            if row_data["md5"] in ("ОШИБКА", "—"):
                return row_data["status"], row_data["tag"]
            entry = ref.get(row_data["name"], {})
            if not entry:
                return "⚠️ Нет эталона", "no_ref"
            if (entry.get("md5", "").lower() == row_data["md5"].lower() and
                    entry.get("sha256", "").lower() == row_data["sha256"].lower()):
                return "✅ Совпадает", "ok"
            return "❌ Не совпадает", "fail"

        def refresh_row(iid, row_data):
            """Redraws one Treeview row from the (already updated) row_data dict."""
            tree.item(iid, values=(
                row_data["name"], row_data["size"],
                row_data["md5"], row_data["sha256"], row_data["status"],
            ), tags=(row_data["tag"],))

        # ── Selection helper ──────────────────────────────────────────────────
        def get_selected():
            """Returns (iid, row_data) for the selected row, or warns and returns (None, None)."""
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Нет выбора", "Выберите файл в таблице.", parent=win)
                return None, None
            iid = sel[0]
            return iid, id_to_result[iid]

        # ── Edit dialog ───────────────────────────────────────────────────────
        def open_edit_dialog(iid, row_data):
            """Opens the reference-hash editing dialog for a single file.

            Args:
                iid: Treeview item id for the file.
                row_data: Mutable result dict for the file.
            """
            if row_data["md5"] in ("ОШИБКА", "—"):
                messagebox.showwarning(
                    "Недоступно",
                    "Нельзя добавить эталон для файла с ошибкой чтения.",
                    parent=win,
                )
                return

            ref = load_reference()
            current = ref.get(row_data["name"], {})

            dlg = tk.Toplevel(win)
            dlg.transient(win)
            dlg.title(f"Редактирование эталона: {row_data['name']}")
            dlg.geometry("520x185")
            dlg.resizable(False, False)
            dlg.grab_set()

            ttk.Label(dlg, text="MD5 (32 hex-символа):").grid(
                row=0, column=0, sticky=tk.W, padx=12, pady=(16, 5))
            md5_var = tk.StringVar(value=current.get("md5", row_data["md5"]))
            md5_entry = ttk.Entry(dlg, textvariable=md5_var, width=46, font=("Consolas", 10))
            md5_entry.grid(row=0, column=1, padx=(0, 12), pady=(16, 5), sticky=tk.EW)

            ttk.Label(dlg, text="SHA256 (64 hex-символа):").grid(
                row=1, column=0, sticky=tk.W, padx=12, pady=5)
            sha256_var = tk.StringVar(value=current.get("sha256", row_data["sha256"]))
            sha256_entry = ttk.Entry(dlg, textvariable=sha256_var, width=46, font=("Consolas", 10))
            sha256_entry.grid(row=1, column=1, padx=(0, 12), pady=5, sticky=tk.EW)

            dlg.columnconfigure(1, weight=1)

            def validate_hex(value, expected_len, label):
                """Returns (cleaned_str, error_msg_or_None)."""
                s = value.strip().lower()
                if len(s) != expected_len:
                    return None, f"{label}: длина должна быть {expected_len} символов (введено {len(s)})"
                if not all(c in "0123456789abcdef" for c in s):
                    return None, f"{label}: допустимы только символы 0–9 и a–f"
                return s, None

            def on_save():
                md5_clean, err = validate_hex(md5_var.get(), 32, "MD5")
                if err:
                    messagebox.showerror("Ошибка ввода", err, parent=dlg)
                    return
                sha256_clean, err = validate_hex(sha256_var.get(), 64, "SHA256")
                if err:
                    messagebox.showerror("Ошибка ввода", err, parent=dlg)
                    return

                ref = load_reference()
                ref[row_data["name"]] = {"md5": md5_clean, "sha256": sha256_clean}
                try:
                    save_reference(ref)
                except Exception as e:
                    messagebox.showerror("Ошибка записи",
                                         f"Не удалось сохранить hashes.json:\n{e}", parent=dlg)
                    return

                new_status, new_tag = recompute_status(row_data, ref)
                row_data["status"] = new_status
                row_data["tag"]    = new_tag
                refresh_row(iid, row_data)

                self.add_test_log(f"✏️ Добавлен эталон для {row_data['name']}")
                messagebox.showinfo("Готово", "Эталон сохранён", parent=dlg)
                dlg.destroy()

            btn_row = ttk.Frame(dlg)
            btn_row.grid(row=2, column=0, columnspan=2, pady=14)
            ttk.Button(btn_row, text="Сохранить", command=on_save).pack(side=tk.LEFT, padx=10)
            ttk.Button(btn_row, text="Отмена",    command=dlg.destroy).pack(side=tk.LEFT)

            md5_entry.focus_set()
            dlg.bind("<Return>", lambda _: on_save())
            dlg.bind("<Escape>", lambda _: dlg.destroy())

        # ── Delete reference ──────────────────────────────────────────────────
        def delete_reference(iid, row_data):
            """Removes the reference entry for this file from hashes.json.

            Args:
                iid: Treeview item id.
                row_data: Mutable result dict for the file.
            """
            ref = load_reference()
            if row_data["name"] not in ref:
                messagebox.showinfo("Нет эталона",
                                    f"Для файла «{row_data['name']}» эталон не задан.",
                                    parent=win)
                return
            if not messagebox.askyesno("Подтверждение",
                                       f"Удалить эталон для:\n{row_data['name']}?",
                                       parent=win):
                return
            del ref[row_data["name"]]
            try:
                save_reference(ref)
            except Exception as e:
                messagebox.showerror("Ошибка записи",
                                     f"Не удалось сохранить hashes.json:\n{e}", parent=win)
                return

            new_status, new_tag = recompute_status(row_data, ref)
            row_data["status"] = new_status
            row_data["tag"]    = new_tag
            refresh_row(iid, row_data)

            self.add_test_log(f"🗑️ Удалён эталон для {row_data['name']}")
            messagebox.showinfo("Готово", "Эталон удалён", parent=win)

        # ── Context menu ──────────────────────────────────────────────────────
        ctx_menu = tk.Menu(win, tearoff=0)

        def show_context_menu(event):
            iid = tree.identify_row(event.y)
            if not iid:
                return
            tree.selection_set(iid)
            ctx_menu.post(event.x_root, event.y_root)

        def ctx_edit():
            iid, row_data = get_selected()
            if iid:
                open_edit_dialog(iid, row_data)

        def ctx_delete():
            iid, row_data = get_selected()
            if iid:
                delete_reference(iid, row_data)

        def ctx_copy(col_idx, label):
            iid, row_data = get_selected()
            if not iid:
                return
            value = tree.item(iid, "values")[col_idx]
            if value in ("ОШИБКА", "—", ""):
                return
            win.clipboard_clear()
            win.clipboard_append(value)
            messagebox.showinfo("Скопировано",
                                f"{label} скопирован в буфер обмена:\n{value}", parent=win)

        ctx_menu.add_command(label="✏️ Добавить/редактировать эталон", command=ctx_edit)
        ctx_menu.add_command(label="🗑️ Удалить эталон",                command=ctx_delete)
        ctx_menu.add_separator()
        ctx_menu.add_command(label="Скопировать MD5",    command=lambda: ctx_copy(2, "MD5"))
        ctx_menu.add_command(label="Скопировать SHA256", command=lambda: ctx_copy(3, "SHA256"))

        tree.bind("<Button-3>", show_context_menu)
        win.bind("<Button-1>", lambda e: ctx_menu.unpost())

        # ── Double-click copies MD5 / SHA256 ─────────────────────────────────
        HASH_COLS = {"#3": ("MD5", 2), "#4": ("SHA256", 3)}

        def on_double_click(event):
            col   = tree.identify_column(event.x)
            iid   = tree.identify_row(event.y)
            if not iid or col not in HASH_COLS:
                return
            label, idx = HASH_COLS[col]
            value = tree.item(iid, "values")[idx]
            if value in ("ОШИБКА", "—", ""):
                return
            win.clipboard_clear()
            win.clipboard_append(value)
            messagebox.showinfo("Скопировано",
                                f"{label} скопирован в буфер обмена:\n{value}", parent=win)

        tree.bind("<Double-1>", on_double_click)

        # ── Bottom bar ────────────────────────────────────────────────────────
        bottom = ttk.Frame(win)
        bottom.grid(row=2, column=0, columnspan=2, sticky="ew", pady=6, padx=8)

        hint = ttk.Label(bottom,
                         text="ПКМ или двойной клик по MD5/SHA256 — дополнительные действия",
                         foreground=COLORS["text_secondary"])
        hint.pack(side=tk.LEFT)

        def on_edit_btn():
            iid, row_data = get_selected()
            if iid:
                open_edit_dialog(iid, row_data)

        def on_delete_btn():
            iid, row_data = get_selected()
            if iid:
                delete_reference(iid, row_data)

        def save_csv():
            path = filedialog.asksaveasfilename(
                parent=win,
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"hash_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            )
            if not path:
                return
            try:
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(["Имя файла", "Размер (МБ)", "MD5", "SHA256", "Статус"])
                    for r in results:
                        writer.writerow([r["name"], r["size"], r["md5"], r["sha256"], r["status"]])
                messagebox.showinfo("Сохранено", f"Отчёт сохранён:\n{path}", parent=win)
                self.add_test_log(f"💾 Отчёт хешей сохранён: {path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{e}", parent=win)

        ttk.Button(bottom, text="💾 Сохранить отчёт (CSV)", command=save_csv).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bottom, text="🗑️ Удалить эталон",        command=on_delete_btn).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bottom, text="✏️ Добавить/редактировать эталон", command=on_edit_btn).pack(side=tk.RIGHT, padx=4)

    # ---------------------- Поиск пути Р7 ----------------------
    def _monitor_update_dialog(self, stop_event, log_cb=None, interval=2):
        """Background thread: checks for the update dialog every `interval` seconds.

        Args:
            stop_event: threading.Event — set it to stop the loop.
            log_cb: Callable for log output. Defaults to self.add_test_log.
            interval: Seconds between checks.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        while not stop_event.is_set():
            self._close_update_dialog_if_exists(log_cb=log_cb, search_timeout=1)
            stop_event.wait(timeout=interval)

    def _win_title_contains(self, *substrings):
        """Returns True if any visible top-level window's title contains
        any of the given substrings (case-insensitive).

        Args:
            *substrings: One or more strings to search for.

        Returns:
            bool
        """
        if not WIN32_OK:
            return False
        import win32gui
        found = []
        needles = [s.lower() for s in substrings]
        def _cb(h, _):
            if win32gui.IsWindowVisible(h):
                t = win32gui.GetWindowText(h).lower()
                if any(n in t for n in needles):
                    found.append(h)
        win32gui.EnumWindows(_cb, found)
        return bool(found)

    def _cleanup_x2t_temp_pdfs(self, log_cb=None):
        """Removes leftover temp_export_x2t_*.pdf files from %TEMP%.

        Safe to call even if save_as_pdf never ran or failed mid-save —
        glob simply matches nothing in that case.

        Args:
            log_cb: Callable for error logging; defaults to self.add_test_log.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        try:
            temp_dir = Path(os.environ.get("TEMP", "."))
            for leftover in temp_dir.glob("temp_export_x2t_*.pdf"):
                leftover.unlink(missing_ok=True)
        except Exception as e:
            log_cb(f"⚠️ Не удалось удалить временный PDF: {e}")

    def _click_priority_button(self, hwnd, keyword_priority, log_cb=None):
        """Ищет среди дочерних окон hwnd кнопку, текст которой содержит одно
        из ключевых слов (по приоритету — первое совпавшее слово выигрывает),
        и кликает по ней через Win32-сообщения (BM_CLICK, с запасным путём
        через WM_LBUTTONDOWN/UP) — без pyautogui и без зависимости от фокуса.

        Вынесено из _close_update_dialog_if_exists, чтобы тот же код кликал
        по кнопке диалога «Сохранить изменения?» при закрытии Р7-Офис.

        Args:
            hwnd: Родительское окно, чьи дочерние окна перебираются.
            keyword_priority: Кортеж подстрок (без учёта регистра) в порядке
                приоритета.
            log_cb: Функция логирования; по умолчанию self.add_test_log.

        Returns:
            tuple[bool, str | None]: (кликнули ли, текст найденной кнопки).
        """
        if log_cb is None:
            log_cb = self.add_test_log
        import win32gui
        import win32con

        children = []
        def _collect(h, _):
            try:
                children.append((h, win32gui.GetWindowText(h), win32gui.GetClassName(h)))
            except Exception:
                pass
        try:
            win32gui.EnumChildWindows(hwnd, _collect, None)
        except Exception:
            pass

        for keyword in keyword_priority:
            for h, text, cls in children:
                if keyword in text.lower():
                    log_cb(f"   Найдена кнопка: «{text}», нажимаю...")
                    try:
                        win32gui.SendMessage(h, win32con.BM_CLICK, 0, 0)
                        return True, text
                    except Exception:
                        pass
                    try:
                        win32gui.PostMessage(h, win32con.WM_LBUTTONDOWN,
                                             win32con.MK_LBUTTON, 0)
                        time.sleep(0.05)
                        win32gui.PostMessage(h, win32con.WM_LBUTTONUP, 0, 0)
                        return True, text
                    except Exception:
                        pass
        if log_cb is not None and children:
            log_cb("   ⚠️ Кнопки для закрытия не найдены. Дочерние окна для диагностики:")
            for h, text, cls in children:
                if text or cls:
                    log_cb(f"      hwnd={h}  class={cls!r}  text={text!r}")
        return False, None

    def _terminate_r7_processes(self, log_cb=None):
        """Принудительно завершает все процессы Р7-Офис: terminate(), затем
        kill() для тех, что не откликнулись за 3 сек.

        Крайняя мера closing-последовательности — раньше её не было вовсе:
        если диалог «Сохранить изменения?» не распознавался слепой
        Alt+F4→Right→Enter, процесс Р7 оставался висеть до ручного
        вмешательства, держа файл заблокированным весь оставшийся прогон.

        Returns:
            bool: True, если процессов не осталось (включая случай, когда их
            не было изначально).
        """
        if log_cb is None:
            log_cb = self.add_test_log
        if not PSUTIL_OK:
            return True
        procs = self._get_r7_processes(log_cb=lambda _m: None)
        if not procs:
            return True

        # Родителя — первым. Наблюдалось 25.08.2026: главный процесс
        # (editors.exe) пересоздаёт убитые дочерние рендереры быстрее, чем
        # цикл успевает пройти по списку, и завершение уходит в бесконечный
        # respawn. Пока имя родителя вообще не попадало в маску поиска, это
        # выглядело как «процессы не убиваются»; теперь он в списке, но
        # порядок всё равно важен — сначала тот, кто порождает.
        procs = sorted(procs, key=lambda pr: self._R7_TERMINATE_ORDER.get(
            (self._safe_proc_name(pr) or ""), 99))

        for p in procs:
            try:
                p.terminate()
            except Exception:
                pass
        try:
            _gone, alive = psutil.wait_procs(procs, timeout=3)
        except Exception:
            alive = procs
        for p in alive:
            try:
                p.kill()
            except Exception:
                pass
        if alive:
            log_cb(f"🔪 Принудительно завершено процессов Р7-Офис: {len(alive)}")

        # Проверяем, а не рапортуем. Прежняя версия возвращала True всегда —
        # включая случай, когда процессы пережили и terminate, и kill: вызывающий
        # код считал Р7 закрытым, а тот держал файл заблокированным.
        self._r7_pids = None
        left = self._get_r7_processes(log_cb=lambda _m: None)
        if left:
            log_cb(f"⚠️ Процессы Р7-Офис пережили завершение: {len(left)} "
                   f"(PID: {', '.join(str(p.pid) for p in left)})")
            return False
        return True

    # Порядок завершения: сначала порождающие процессы, потом порождаемые.
    _R7_TERMINATE_ORDER = {"desktopeditors.exe": 0, "editors.exe": 1,
                           "editors_helper.exe": 2}

    @staticmethod
    def _safe_proc_name(proc):
        """Имя процесса в нижнем регистре, либо None — psutil.Process.name()
        поднимает исключение на процессе, умершем между сканом и вызовом."""
        try:
            return (proc.name() or "").lower()
        except Exception:
            return None

    def _close_r7_gracefully(self, hwnd, log_cb=None, timeout=10):
        """Закрывает окно Р7-Офис, адресованное конкретным hwnd.

        Раньше закрытие было слепым: Alt+F4 → Right → Enter уходили тому
        окну, что в этот момент имело фокус, — а фокус мог перехватить
        монитор диалога обновления или случайный клик. Здесь WM_CLOSE
        отправляется напрямую целевому hwnd через Win32-сообщение, клавиатура
        не участвует.

        Если Р7-Офис показывает диалог «Сохранить изменения?», он ищется
        среди top-level окон, принадлежащих тому же процессу (owner PID через
        GetWindowThreadProcessId — не зависит от текста заголовка, который
        отличается между версиями/локалями), и закрывается через
        _click_priority_button — тем же надёжным путём, что и диалог
        обновления, а не вслепую по стрелке и Enter.

        Если окно не исчезло за timeout секунд — например, диалог не
        распознан, — процесс Р7 завершается принудительно через
        _terminate_r7_processes. Раньше в этом случае программа просто
        продолжала бы работу с зависшим Р7 на фоне.

        Args:
            hwnd: Дескриптор закрываемого окна Р7-Офис. None означает, что
                окно не было найдено заранее — сразу переходим к
                принудительному завершению процесса.
            log_cb: Функция логирования; по умолчанию self.add_test_log.
            timeout: Сколько секунд ждать штатного закрытия после WM_CLOSE.

        Returns:
            bool: True — окно закрыто штатно; False — потребовалось
            принудительное завершение процесса (само завершение
            произошло в любом случае).
        """
        if log_cb is None:
            log_cb = self.add_test_log

        if not (WIN32_OK and hwnd):
            log_cb("⚠️ Окно Р7-Офис не найдено — завершаем процесс напрямую")
            self._terminate_r7_processes(log_cb)
            return False

        import win32gui
        import win32con
        import win32process

        # Диалог сохранения — не диалог обновления: узнаём его не по тексту
        # заголовка (тот отличается между версиями и локалями), а по тому,
        # что это НОВОЕ top-level окно того же процесса, появившееся уже
        # после WM_CLOSE.
        SAVE_DIALOG_BUTTONS = ('не сохранять', "don't save", 'нет', 'no')

        try:
            _, owner_pid = win32process.GetWindowThreadProcessId(hwnd)
        except Exception:
            owner_pid = None

        def _sibling_windows():
            wins = []
            def _enum(h, _):
                if h == hwnd or not win32gui.IsWindowVisible(h):
                    return
                try:
                    _, pid = win32process.GetWindowThreadProcessId(h)
                except Exception:
                    return
                if pid == owner_pid:
                    wins.append(h)
            win32gui.EnumWindows(_enum, None)
            return wins

        # Модальный файловый диалог блокирует закрытие наглухо: пока открыт
        # «Сохранить как», WM_CLOSE главному окну не делает ничего, и весь
        # timeout уходит впустую, а потом kill. Такой диалог остаётся, если
        # тест «Сохранение в PDF» не довёл экспорт до конца (не приняли путь,
        # выскочил вопрос о перезаписи). Снимаем его ДО WM_CLOSE.
        self._cancel_blocking_dialogs(owner_pid, log_cb)

        close_started = time.time()
        try:
            win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
        except Exception:
            log_cb("⚠️ Не удалось отправить WM_CLOSE — завершаем процесс напрямую")
            self._terminate_r7_processes(log_cb)
            return False

        # CDP мог ещё ни разу не понадобиться в этом прогоне: коннектор
        # создаётся при запуске Р7, а connect() зовётся лениво из
        # _wait_for_bold_button_cdp, и если триггер готовности ни разу не
        # сработал — соединения нет. connect() идемпотентен, так что для уже
        # подключённого это no-op; таймаут короткий, чтобы не тормозить выход.
        if self._webdriver_connector is not None:
            try:
                self._webdriver_connector.connect(timeout=1.0)
            except Exception:
                pass

        deadline = time.time() + timeout
        dismissed = False
        diag_dumped = False
        cdp_tries = 0
        last_cdp_try = 0.0
        cdp_clicked = False   # только чтобы не повторять строку в логе
        while time.time() < deadline:
            if not win32gui.IsWindow(hwnd):
                log_cb(f"🔚 Р7-Офис закрыт штатно за {time.time() - close_started:.1f} сек")
                return True

            if not dismissed:
                # Путь 1 — отдельное окно-диалог того же процесса. Работает,
                # только если сборка Р7 рисует его классическими Win32-виджетами.
                if owner_pid:
                    for w in _sibling_windows():
                        if not diag_dumped:
                            # Сам факт «окно-диалог есть, но кнопку в нём не
                            # нашли» ниже не логируется: _click_priority_button
                            # печатает дамп только когда дочерние окна ЕСТЬ, а у
                            # диалога Qt их нет вовсе (Qt рисует кнопки сам, не
                            # заводя HWND). Поэтому заголовок и класс окна пишем
                            # здесь — именно они отличают Qt-диалог от
                            # HTML-модалки, у которой окна нет совсем.
                            try:
                                log_cb(f"   Окно-кандидат на диалог сохранения: "
                                       f"hwnd={w} class={win32gui.GetClassName(w)!r} "
                                       f"title={win32gui.GetWindowText(w)!r}")
                            except Exception:
                                pass
                        clicked, text = self._click_priority_button(
                            w, SAVE_DIALOG_BUTTONS,
                            # Раньше сюда передавался глушитель `lambda _m: None`,
                            # и дамп дочерних окон — единственная диагностика,
                            # объясняющая, почему кнопка не нашлась, — молча
                            # выбрасывался. Пишем его, но один раз за закрытие,
                            # чтобы не залить лог на каждой итерации цикла.
                            log_cb=(log_cb if not diag_dumped else (lambda _m: None)))
                        # Флаг взводим только когда окно реально осмотрели.
                        # Если сейчас siblings пусты, а диалог появится на
                        # следующей итерации — его диагностику терять нельзя.
                        diag_dumped = True
                        if clicked:
                            log_cb(f"   Диалог сохранения закрыт кнопкой «{text}»")
                            dismissed = True
                            break

                # Путь 2 — модалка внутри окна редактора (HTML в CEF). Отдельного
                # окна ОС у неё нет, поэтому путь 1 её не находит вообще: hwnd
                # остаётся жив, siblings пусты, и до этой правки цикл просто
                # крутился весь timeout и уходил в kill — ровно тот симптом,
                # с которого начали («не закрылся за 10 сек»).
                # Опрашиваем не чаще CDP_RETRY_SEC: каждый вызов — round-trip по
                # websocket, а при оборванном соединении ещё и строка в логе;
                # на шаге цикла в 0.2 с это залило бы лог полусотней сообщений.
                if not dismissed and (time.time() - last_cdp_try) >= self.CLOSE_CDP_RETRY_SEC:
                    last_cdp_try = time.time()
                    cdp_tries += 1
                    res = self._cdp_dismiss_save_dialog()
                    if res and not cdp_clicked:
                        # Намеренно НЕ ставим dismissed=True: JS сообщает «клик
                        # прошёл», а не «модалка закрылась». Если попали не по той
                        # кнопке (например, по видимому элементу в фоновом
                        # документе), латч навсегда отключил бы и Win32-путь, и
                        # повторные попытки — и закрытие гарантированно свелось бы
                        # к kill. Признак успеха тут ровно один: окно исчезло, его
                        # проверяет IsWindow в начале цикла.
                        cdp_clicked = True
                        log_cb(f"   Нажата кнопка модалки сохранения через CDP: «{res}»")

            time.sleep(0.2)

        # Принудительное завершение — не аварийный путь, а штатный запасной:
        # для бенчмарка терять несохранённые правки тестового файла безопаснее,
        # чем вслепую нажать «Сохранить» и перезаписать эталон.
        log_cb(f"⚠️ Р7-Офис не закрылся за {timeout} сек — завершаем процесс принудительно")
        log_cb(f"   (Win32-кнопка: {'нажата' if dismissed else 'не найдена'}; "
               f"CDP: коннектор {'есть' if self._webdriver_connector else 'нет'}, "
               f"попыток {cdp_tries}, клик {'был' if cdp_clicked else 'не прошёл'})")
        self._terminate_r7_processes(log_cb)
        return False

    # Заголовки модальных окон, которые блокируют закрытие Р7 и которые
    # безопасно отменять: файловый диалог экспорта и вопрос о перезаписи.
    # Только составные фразы — по голому «сохранить» под маску попал бы и сам
    # вопрос «Сохранить изменения?», у которого отмена означает «не закрывать».
    BLOCKING_DIALOG_TITLES = (
        "сохранить как", "save as",
        "подтверждение сохранения", "confirm save as",
        "подтвердите перезапись", "confirm overwrite",
    )
    CANCEL_BUTTONS = ("отмена", "cancel", "отменить")

    def _cancel_blocking_dialogs(self, owner_pid, log_cb=None, max_rounds=3):
        """Отменяет модальные диалоги, из-за которых Р7 не реагирует на WM_CLOSE.

        Речь прежде всего о «Сохранить как»: он остаётся открытым, если тест
        «Сохранение в PDF» не довёл экспорт до конца. Пока он на экране,
        WM_CLOSE главному окну не делает ничего — весь timeout закрытия уходит
        впустую и заканчивается принудительным завершением процесса.

        В отличие от диалога «Сохранить изменения?», здесь безопасно жать
        именно «Отмена»: отмена экспорта в PDF ничего не портит (файл
        временный), тогда как отмена вопроса о сохранении означала бы «не
        закрывать Р7».

        Диалог ищется среди видимых top-level окон процесса Р7 по заголовку —
        в отличие от кнопок, заголовок у такого окна есть и через win32gui
        читается (это настоящее окно ОС, а не HTML-модалка внутри редактора).
        Сначала пробуем кнопку «Отмена», затем WM_CLOSE самому диалогу: для
        файлового диалога это эквивалентно отмене.

        Args:
            owner_pid: PID процесса, чьи окна проверяем. None — проверяем все
                видимые окна, принадлежащие любому процессу Р7.
            log_cb: Функция логирования; по умолчанию self.add_test_log.
            max_rounds: Сколько раз повторить проход — за отменой одного
                диалога может открыться следующий (перезапись → сам «Сохранить
                как»).

        Returns:
            int: сколько диалогов было закрыто.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        if not WIN32_OK:
            return 0

        import win32gui
        import win32con
        import win32process

        allowed_pids = {owner_pid} if owner_pid else None
        if allowed_pids is None and PSUTIL_OK:
            allowed_pids = {p.pid for p in self._get_r7_processes(log_cb=lambda _m: None)}

        closed = 0
        for _ in range(max_rounds):
            targets = []

            def _enum(h, _):
                if not win32gui.IsWindowVisible(h):
                    return
                try:
                    title = win32gui.GetWindowText(h).lower()
                except Exception:
                    return
                if not any(t in title for t in self.BLOCKING_DIALOG_TITLES):
                    return
                if allowed_pids:
                    try:
                        _, pid = win32process.GetWindowThreadProcessId(h)
                    except Exception:
                        return
                    if pid not in allowed_pids:
                        return      # чужой «Сохранить как» — не наш, не трогаем
                targets.append(h)

            try:
                win32gui.EnumWindows(_enum, None)
            except Exception:
                break

            if not targets:
                break

            for h in targets:
                try:
                    title = win32gui.GetWindowText(h)
                except Exception:
                    title = "?"
                clicked, btn = self._click_priority_button(
                    h, self.CANCEL_BUTTONS, log_cb=lambda _m: None)
                if clicked:
                    log_cb(f"   🚪 Блокирующий диалог «{title}» отменён кнопкой «{btn}»")
                else:
                    try:
                        win32gui.PostMessage(h, win32con.WM_CLOSE, 0, 0)
                        log_cb(f"   🚪 Блокирующий диалог «{title}» закрыт через WM_CLOSE")
                    except Exception as e:
                        log_cb(f"   ⚠️ Не удалось закрыть диалог «{title}»: {e}")
                        continue
                closed += 1
            time.sleep(0.3)     # дать диалогу исчезнуть перед следующим проходом

        return closed

    def _capture_cdp_ui_baseline(self, log_cb=None):
        """Снимает базовый DOM-снимок сразу после открытия файла, ДО первой
        операции — чтобы последующие _cdp_dump_ui показывали только элементы,
        появившиеся из-за конкретного действия (right-click и т.п.), а не
        постоянно смонтированные части интерфейса.

        Появилась из-за issue #9: разные вызовы _cdp_dump_ui (в разных
        точках теста — открытие контекстного меню на разных ячейках, разное
        клиповое состояние) возвращали ОДИН И ТОТ ЖЕ список из 12 пунктов на
        неизменных экранных координатах (x=1393 у всех, независимо от того,
        над какой ячейкой был right-click). Это не мог быть настоящий
        всплывающий попап — координаты попапа менялись бы вместе с курсором.
        На деле список совпал с overflow-меню тулбара (кнопка «Условное
        форматирование» в том же дампе — это именно тулбарная кнопка, не
        пункт меню): MENU_SEL цепляет его, потому что он тоже помечен как
        `.dropdown-menu` и остаётся "видимым" по всем текущим критериям
        (offsetParent/rect/computed style), просто свёрнут не через
        display:none. Настоящий контекстный попап, судя по всему, живёт вне
        обходимого DOM (нативный CEF-оверлей) — сравнение с базовым снимком
        не заставит появиться то, чего нет в DOM, зато надёжно уберёт из
        дампа шум вроде этого overflow-меню, если он относится к статичной
        части интерфейса, снятой ДО right-click.

        Безопасно вызывать даже без CDP (коннектор не создан/не подключён —
        self._cdp_ui_baseline останется None, _cdp_dump_ui в этом случае
        просто не диффит и печатает как раньше).

        Сбрасывает self._cdp_dump_seen (дедуп «один дамп на ключ за
        сессию» в _cdp_dump_ui). Без сброса дедуп-множество живёт на весь
        срок жизни self.R7Testovarka, а не на один запуск Р7: при повторном
        прогоне теста в той же сессии GUI (или в Batch по нескольким
        версиям) каждый вызов после первого молча схлопывался бы в
        `if key in seen: return`, а этот метод исправно тратил бы CDP
        round-trip на снимок, который _cdp_dump_ui заведомо не покажет —
        итог: пустая трата времени на каждый запуск, кроме первого в
        сессии. Сброс здесь — на каждый launch Р7 DOM у него всё равно
        свежий, показывать дамп заново для нового запуска корректно.

        Args:
            log_cb: Функция логирования; по умолчанию self.add_test_log.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        self._cdp_ui_baseline = None
        self._cdp_dump_seen = set()
        connector = self._webdriver_connector
        if connector is None or not getattr(connector, "connected", False):
            return
        try:
            self._cdp_ui_baseline = connector.dump_visible_ui()
        except Exception as e:
            self._cdp_ui_baseline = None
            log_cb(f"⚠️ Базовый DOM-снимок не снят ({type(e).__name__}: {e}) — "
                   f"дальнейшие дампы меню покажут все видимые элементы без вычитания")

    # Пункт из дампа считается «уже был в базовом снимке» только если совпал
    # и ключ (тег/id/класс/текст), И положение на экране — с допуском.
    # Строгое совпадение только по ключу пряталось бы за один и тот же
    # generic-маркап: у Р7 и overflow-меню тулбара, и реальный пункт
    # контекстного меню могут быть `<li id="" class="">` с одинаковым
    # текстом (см. issue #9 — ровно так дамп раньше путал два разных
    # элемента). Допуск нужен ровно настолько, чтобы пережить субпиксельный
    # дрожащий рендер одного и того же попапа между снимками, а не чтобы
    # маскировать элемент, реально сидящий в другом месте экрана.
    CDP_ITEM_POSITION_TOLERANCE_PX = 30

    @staticmethod
    def _cdp_item_key(item):
        """Ключ сравнения по содержимому одного элемента DOM-дампа
        (без координат — те сравниваются отдельно, см.
        CDP_ITEM_POSITION_TOLERANCE_PX и _cdp_item_in_baseline)."""
        return (item.get("tag"), item.get("id"), item.get("cls"), item.get("text"))

    @classmethod
    def _cdp_item_in_baseline(cls, item, baseline_by_key):
        """True, если item — тот же элемент, что уже был в базовом снимке:
        совпадает и содержимое (_cdp_item_key), и положение на экране (с
        допуском CDP_ITEM_POSITION_TOLERANCE_PX)."""
        candidates = baseline_by_key.get(cls._cdp_item_key(item))
        if not candidates:
            return False
        ix, iy = item.get("x", 0), item.get("y", 0)
        tol = cls.CDP_ITEM_POSITION_TOLERANCE_PX
        for b in candidates:
            if abs(b.get("x", 0) - ix) <= tol and abs(b.get("y", 0) - iy) <= tol:
                return True
        return False

    def _cdp_dump_ui(self, label, log_cb=None, once_key=None, charge_pace=False):
        """Пишет в лог видимые кнопки и пункты меню, как их видит DOM.

        Диагностика для слепых мест автоматизации: контекстное меню и модалки
        обходятся стрелками вслепую (`down` N раз + Enter), и стоит меню
        обзавестись лишним пунктом, как нажимается не то. По win32gui эти
        подписи недоступны в принципе (Qt+CEF не заводит дочерних HWND), а
        через CDP — видны.

        Если ранее вызывался _capture_cdp_ui_baseline (self._cdp_ui_baseline
        не None) — печатает только элементы, которых не было в базовом
        снимке: постоянно смонтированные части интерфейса (тулбар, его
        overflow-меню) иначе перепечатывались бы на каждый вызов и маскируют
        собой то единственное, что реально интересно — что появилось именно
        из-за этого right-click/модалки (см. issue #9). Без базового снимка
        печатает всё видимое, как раньше.

        Печатает не чаще одного раза на ключ за сессию приложения: дамп нужен,
        чтобы один раз увидеть структуру меню, а не чтобы залить лог на каждом
        прогоне теста (тесты гоняются по 3 прогона, а меню между ними не
        меняется).

        Вызывать вне окна замера либо с charge_pace=True — round-trip по
        websocket иначе попадёт в результат.

        Args:
            label: Человекочитаемая пометка, в какой момент снят дамп.
            log_cb: Функция логирования; по умолчанию self.add_test_log.
            once_key: Ключ дедупликации; по умолчанию сам label.
            charge_pace: Отнести собственную длительность в _paced_total, чтобы
                вычесть её из замера. Ставить только там, где Р7 в этот момент
                гарантированно простаивает (раскрытое меню, открытая модалка) —
                иначе вычтем время, которое Р7 работал.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        connector = self._webdriver_connector
        if connector is None or not getattr(connector, "connected", False):
            return
        key = once_key or label
        seen = getattr(self, "_cdp_dump_seen", None)
        if seen is None:
            seen = self._cdp_dump_seen = set()
        if key in seen:
            return
        seen.add(key)
        _t0 = time.time()
        try:
            items = connector.dump_visible_ui()
        except Exception:
            items = None
        if charge_pace:
            self._paced_total += time.time() - _t0
        if items is None:
            return
        if not items:
            return

        baseline = getattr(self, "_cdp_ui_baseline", None)
        total_visible = len(items)
        if baseline is not None:
            baseline_by_key = {}
            for it in baseline:
                baseline_by_key.setdefault(self._cdp_item_key(it), []).append(it)
            items = [it for it in items if not self._cdp_item_in_baseline(it, baseline_by_key)]
            if not items:
                log_cb(f"   🔬 DOM-дамп ({label}): новых элементов нет "
                       f"(все {total_visible} видимых уже были в базовом снимке до операций)")
                return
            log_cb(f"   🔬 DOM-дамп ({label}): новых элементов {len(items)} "
                   f"из {total_visible} видимых (базовый снимок вычтен)")
        else:
            log_cb(f"   🔬 DOM-дамп ({label}): видимых элементов {total_visible}")
        for it in items[:25]:
            try:
                # x/y/depth (frame) — координаты элемента на экране и глубина
                # вложенности iframe, где он найден. Нужны, чтобы отличить
                # реальный раскрытый попап (координаты рядом с курсором) от
                # статичного элемента где-то в стороне DOM (см. issue #9:
                # дамп раз за разом находил один и тот же список независимо
                # от контекста — координаты покажут, действительно ли это
                # один и тот же неподвижный узел).
                log_cb(f"      • {it.get('text','')!r} "
                       f"<{it.get('tag','')} id={it.get('id','')!r} class={it.get('cls','')!r}> "
                       f"@({it.get('x','?')},{it.get('y','?')}) frame={it.get('depth','?')}")
            except Exception:
                pass

    def _cdp_dismiss_save_dialog(self):
        """Пробует нажать «Не сохранять» в HTML-модалке выхода через CDP.

        Тонкий адаптер над R7WebDriverConnector.dismiss_save_dialog(): гасит
        любые исключения и приводит ответ к тексту нажатой кнопки. Молча
        возвращает None, если CDP в этом запуске недоступен (Р7 стартован без
        debug-флага, нет requests/websocket-client и т.п.) — тогда закрытие
        идёт обычным путём, как и до появления коннектора.

        Returns:
            str | None: текст нажатой кнопки, либо None.
        """
        connector = self._webdriver_connector
        if connector is None:
            return None
        try:
            res = connector.dismiss_save_dialog()
        except Exception:
            return None
        if isinstance(res, dict) and res.get("clicked"):
            return res.get("text") or "не сохранять"
        return None

    # ── Операции теста через CDP ─────────────────────────────────────────
    #
    # Тест-функции обоих воркеров сначала пробуют выполнить операцию через
    # внутренний api редактора по CDP (r7_webdriver_connector: asc_EditSelectAll,
    # asc_Copy/asc_Paste, asc_addWorksheet, asc_insertCells, asc_findCell) и
    # только если это не получилось — шлют клавиши через pyautogui, как раньше.
    #
    # Почему это лучше клавиатуры: клавиши уходят в то окно, которое сейчас в
    # фокусе, требуют развёрнутого окна Р7 и слепой навигации по меню (счётчик
    # `down` уезжает от любого лишнего пункта — см. issue #9), а результат
    # операции ничем не подтверждается. Вызов api идёт мимо фокуса и оконного
    # менеджера и возвращает состояние документа до и после — поэтому каждая
    # операция ниже ещё и проверяется.
    #
    # ГЛАВНОЕ ПРАВИЛО ПОСЛЕДОВАТЕЛЬНОСТЕЙ: изменяющий документ шаг (mutated в
    # ответе JS — вставка, новый лист, insertCells) должен быть В КОНЦЕ. Откат
    # на pyautogui повторяет операцию целиком, и если документ уже изменён
    # предыдущим шагом, правка применится дважды. _cdp_sequence на такой случай
    # откат запрещает (см. mutated_already), но полагаться на это как на
    # штатный путь нельзя — цифра замера в этот момент уже испорчена.
    CDP_OPS_ENABLED = True      # общий выключатель: False → всё идёт клавишами,
                                # как до перевода тестов на CDP (нужно, чтобы
                                # сравнить цифры двух путей на одной сборке)
    CDP_OP_TIMEOUT_SEC = 10.0        # обычная операция
    CDP_LONG_OP_TIMEOUT_SEC = 30.0   # Ctrl+A, вставка, insertCells на большом
                                     # файле: Runtime.evaluate возвращается
                                     # только когда JS отработал, а это и есть
                                     # время самой операции

    def _cdp_ops_connector(self):
        """Коннектор, готовый выполнять операции, либо None.

        Returns:
            R7WebDriverConnector | None
        """
        if not self.CDP_OPS_ENABLED:
            return None
        connector = self._webdriver_connector
        if connector is None or not getattr(connector, "connected", False):
            return None
        return connector

    def _cdp_step(self, caption, fn, log_cb, timeout=None):
        """Выполняет один вызов коннектора и классифицирует результат.

        Args:
            caption: Подпись шага для лога (обычно имя asc_-метода).
            fn: callable(connector, timeout) -> dict | None.
            log_cb: Функция логирования.
            timeout: Таймаут ожидания ответа CDP, сек.

        Returns:
            tuple[str, dict | None]: статус и ответ JS. Статусы:
              "ok"          — выполнено, можно идти дальше;
              "failed"      — не выполнено И документ не тронут, безопасно
                              повторить операцию клавишами;
              "unknown"     — ответа нет или сбой уже после изменения
                              документа: повторять клавишами НЕЛЬЗЯ;
              "unavailable" — CDP в этом запуске недоступен.
        """
        connector = self._cdp_ops_connector()
        if connector is None:
            return ("unavailable", None)
        try:
            res = fn(connector, timeout)
        except Exception as e:
            log_cb(f"   ⚠️ CDP «{caption}»: исключение — {type(e).__name__}: {e}")
            return ("failed", None)

        if res is None:
            # None от evaluate() означает «неизвестно», а не «не выполнено».
            # Если соединение живо — почти наверняка вызов не уложился в
            # таймаут сокета, а JS всё это время работал: операция ушла в Р7,
            # и дублировать её клавишами нельзя.
            if getattr(connector, "connected", False):
                waited = f"за {timeout:.0f} с" if timeout else "в отведённое время"
                log_cb(f"   ⚠️ CDP «{caption}»: ответ не пришёл {waited}, "
                       f"соединение живо — считаю операцию отправленной")
                return ("unknown", None)
            log_cb(f"   ⚠️ CDP «{caption}»: соединение потеряно — откат на клавиши")
            return ("failed", None)

        if not isinstance(res, dict):
            log_cb(f"   ⚠️ CDP «{caption}»: неожиданный ответ {res!r} — откат на клавиши")
            return ("failed", None)

        if res.get("ok"):
            return ("ok", res)

        reason = res.get("reason") or "?"
        detail = res.get("error")
        suffix = f": {detail}" if detail else ""
        if res.get("mutated"):
            log_cb(f"   ⚠️ CDP «{caption}»: сбой уже ПОСЛЕ изменения документа "
                   f"({reason}{suffix}) — повтор клавишами отменён, иначе правка "
                   f"применилась бы дважды")
            return ("unknown", res)
        log_cb(f"   ⚠️ CDP «{caption}»: не выполнено ({reason}{suffix}) — откат на клавиши")
        return ("failed", res)

    def _cdp_sequence(self, label, steps, checker=None, log_cb=None):
        """Выполняет операцию как цепочку вызовов api и ставит её на проверку.

        Args:
            label: Название операции для лога.
            steps: Список кортежей (подпись, fn, timeout, pace_before), где
                fn — callable(connector, timeout) -> dict | None, а
                pace_before — пауза перед шагом (через _pace, вычитается из
                замера). Изменяющий документ шаг обязан быть последним —
                см. комментарий к блоку выше.
            checker: callable(before, after) -> (bool, str) — чем подтверждать
                результат; None — операция не проверяется (например, копирование
                в буфер: в документе оно ничего не меняет).
            log_cb: Функция логирования; по умолчанию self.add_test_log.

        Побочный эффект: суммирует api_ms всех успешных шагов в
        self._cdp_api_ms (сбрасывается вызывающим кодом перед замером, рядом
        с _paced_total и _op_via_cdp — см. run_test_with_runs/measure). Это
        синхронное время внутри рендерера (performance.now(), <1 мс
        разрешение), не зависящее от опроса CPU детектором простоя
        (_wait_operation_done, окно 0.20 с) — тот на операциях короче своего
        окна усреднения даёт разброс до 20× между прогонами одного файла.
        Складываются шаги вместе (не берётся последний): последовательность
        вроде «выделить → скопировать → выделить → вставить» — это несколько
        отдельных вызовов api, и «синхронное время операции» — сумма всех, а
        не только последнего.

        Returns:
            bool: True — операция выполнена (или отправлена) через CDP, вызывающий
            код НЕ должен повторять её клавишами. False — через CDP ничего не
            произошло, нужен обычный pyautogui-путь.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        if self._cdp_ops_connector() is None:
            return False

        mutated_already = False
        last_payload = None
        for caption, fn, timeout, pace_before in steps:
            if pace_before:
                self._pace(pace_before)
            status, payload = self._cdp_step(caption, fn, log_cb, timeout)
            if status == "ok":
                last_payload = payload
                ms = payload.get("api_ms")
                if isinstance(ms, (int, float)):
                    self._cdp_api_ms += ms
                if payload.get("mutated"):
                    mutated_already = True
                continue
            if status == "unknown":
                log_cb(f"   ⚠️ CDP «{label}»: цепочка прервана на шаге «{caption}», "
                       f"клавишами не повторяю — цифра этого прогона недостоверна")
                self._op_via_cdp = True
                return True
            # failed / unavailable
            if mutated_already:
                log_cb(f"   ⚠️ CDP «{label}»: шаг «{caption}» не выполнен, но документ "
                       f"уже изменён предыдущим шагом — откат на клавиши отменён")
                self._op_via_cdp = True
                return True
            return False

        self._op_via_cdp = True
        self._cdp_verify_or_defer(label, last_payload, checker, log_cb)
        return True

    def _cdp_verify_or_defer(self, label, payload, checker, log_cb):
        """Проверяет результат операции по снимкам состояния из ответа JS.

        Снимок «после» снят внутри той же страницы сразу за вызовом api, то
        есть бесплатно (без ещё одного round-trip). Но часть операций Р7
        доводит асинхронно — вставка из системного буфера, например, уходит в
        нативный код и к моменту возврата asc_Paste документ ещё не изменён.
        Поэтому неподтвердившаяся проверка не считается провалом сразу, а
        откладывается до _flush_pending_cdp_verify — тот перечитывает
        состояние уже ПОСЛЕ закрытия окна замера, и его round-trip в цифру
        не попадает.

        Args:
            label: Название операции.
            payload: Ответ JS последнего шага (с полями before/after).
            checker: callable(before, after) -> (bool, str), либо None.
            log_cb: Функция логирования.
        """
        self._pending_cdp_verify = None
        # Сумма api_ms всех шагов последовательности (_cdp_sequence уже
        # накопила её к этому моменту) — синхронное время внутри рендерера,
        # не зависящее от опроса CPU детектором простоя. См. docstring
        # _cdp_sequence.
        api_ms_note = (f" [api: {self._cdp_api_ms:.2f} мс]"
                       if self._cdp_api_ms > 0 else "")
        if not isinstance(payload, dict):
            log_cb(f"   🧩 CDP «{label}»: выполнено{api_ms_note}")
            return
        method = payload.get("method") or "api"
        if checker is None:
            log_cb(f"   🧩 CDP «{label}»: выполнено ({method}){api_ms_note}")
            return
        before, after = payload.get("before"), payload.get("after")
        ok, detail = checker(before, after)
        if ok:
            log_cb(f"   🧩 CDP «{label}»: выполнено ({method}){api_ms_note}, "
                   f"проверено — {detail}")
            return
        log_cb(f"   🧩 CDP «{label}»: выполнено ({method}){api_ms_note}, "
               f"сразу не подтвердилось ({detail}) — перепроверю после замера")
        self._pending_cdp_verify = (label, before, checker)

    def _flush_pending_cdp_verify(self, log_cb=None):
        """Доводит отложенную проверку CDP-операции — уже ВНЕ окна замера.

        Вызывать в обоих воркерах сразу после _wait_operation_done, рядом с
        _flush_pending_modal_confirm (правило зеркалирования из CLAUDE.md).
        Round-trip по websocket здесь стоит миллисекунды, но в замер он
        попадать не должен — отсюда и отложенность.

        Args:
            log_cb: Функция логирования; по умолчанию self.add_test_log.
        """
        pending = getattr(self, "_pending_cdp_verify", None)
        if not pending:
            return
        self._pending_cdp_verify = None
        if log_cb is None:
            log_cb = self.add_test_log
        label, before, checker = pending
        connector = self._cdp_ops_connector()
        if connector is None:
            log_cb(f"   ⚠️ CDP-проверка «{label}»: соединение недоступно, "
                   f"результат операции не подтверждён")
            return
        try:
            after = connector.document_state(timeout=self.CDP_OP_TIMEOUT_SEC)
        except Exception as e:
            log_cb(f"   ⚠️ CDP-проверка «{label}»: {type(e).__name__}: {e}")
            return
        if after is None:
            log_cb(f"   ⚠️ CDP-проверка «{label}»: состояние документа прочитать "
                   f"не удалось")
            return
        ok, detail = checker(before, after)
        if ok:
            log_cb(f"   ✅ CDP-проверка «{label}»: {detail}")
        else:
            log_cb(f"   ⚠️ CDP-проверка «{label}»: не подтверждено — {detail}")

    # ── Чем подтверждается результат операции ────────────────────────────
    # Все проверки терпимы к None: снимок состояния собирается из отдельных
    # asc_-геттеров, и любой из них в чужой сборке может не существовать —
    # тогда поле приходит None, и «проверить не удалось» честнее, чем упасть.

    # Точное число строк листа зависит от сборки, поэтому проверяем не
    # совпадение с константой, а порядок величины: столько строк вручную не
    # выделяют.
    CDP_WHOLE_SHEET_MIN_ROWS = 10000

    @classmethod
    def _cdp_check_whole_sheet_selected(cls, before, after):
        """Ctrl+A: выделение стало полным листом.

        Форм записи у полного выделения несколько, и они зависят от того, как
        Р7 свернул диапазон. На живой сборке (2026.2.2.x, прогон 2026-08-25)
        asc_getActiveRangeStr после asc_EditSelectAll отдаёт «1:1048576» —
        диапазон СТРОК, без букв столбцов. Проверка только по форме
        «A1:XFD1048576» этого не узнавала и честно выполненную операцию
        помечала как неподтверждённую.
        """
        sel = (after or {}).get("selection")
        if not isinstance(sel, str):
            return (False, "выделение прочитать не удалось")
        s = sel.strip().upper()

        # «1:1048576» — все строки листа
        m = re.match(r"^(\d+):(\d+)$", s)
        if m and int(m.group(2)) - int(m.group(1)) + 1 >= cls.CDP_WHOLE_SHEET_MIN_ROWS:
            return (True, f"выделено {sel} (все строки листа)")

        # «A1:XFD1048576» — тот же лист, записанный ячейками
        m = re.match(r"^A1:[A-Z]+(\d+)$", s)
        if m and int(m.group(1)) >= cls.CDP_WHOLE_SHEET_MIN_ROWS:
            return (True, f"выделено {sel}")

        # «A:XFD» — все столбцы листа
        if re.match(r"^A:[A-Z]{2,3}$", s):
            return (True, f"выделено {sel} (все столбцы листа)")

        return (False, f"выделено {sel!r} — это не весь лист")

    @staticmethod
    def _cdp_check_selection_is(expected):
        """Фабрика проверки «выделен ровно этот диапазон»."""
        def _check(before, after):
            sel = (after or {}).get("selection")
            if not isinstance(sel, str):
                return (False, "выделение прочитать не удалось")
            if sel.strip().upper() == expected.upper():
                return (True, f"выделено {sel}")
            return (False, f"выделено {sel!r}, ожидалось {expected!r}")
        return _check

    @staticmethod
    def _cdp_check_sheet_added(before, after):
        """Новый лист: число листов выросло ровно на один."""
        b = (before or {}).get("sheets")
        a = (after or {}).get("sheets")
        if not isinstance(b, int) or not isinstance(a, int):
            return (False, "число листов прочитать не удалось")
        if a == b + 1:
            return (True, f"листов было {b}, стало {a}")
        return (False, f"листов было {b}, стало {a}")

    @staticmethod
    def _cdp_check_document_changed(before, after):
        """Документ изменился: сдвинулась история правок.

        Универсальный признак для операций, у которых нет дешёвого прямого
        подтверждения (вставка, insertCells): любая принятая правка создаёт
        точку в History, а до первой правки Can_Undo() == false.
        """
        b, a = (before or {}), (after or {})
        for key, human in (("historyIndex", "позиция в истории"),
                           ("historyPoints", "число точек истории")):
            bv, av = b.get(key), a.get(key)
            if isinstance(bv, int) and isinstance(av, int) and av != bv:
                return (True, f"{human}: {bv} → {av}")
        if not b.get("canUndo") and a.get("canUndo"):
            return (True, "появилась возможность отменить правку")
        if b.get("historyIndex") is None and b.get("historyPoints") is None:
            return (False, "историю правок прочитать не удалось")
        return (False, "история правок не сдвинулась")

    # ── Сами операции ────────────────────────────────────────────────────
    # Каждая возвращает True, если делать что-то клавишами уже не нужно.

    # Сколько ждать подключения к CDP перед первой операцией. Больше, чем
    # BOLD_BUTTON_CDP_CONNECT_TIMEOUT_SEC (0.5 с): там короткий таймаут защищал
    # замер «Открытие файла», в который попадало ожидание, а здесь файл уже
    # открыт и торопиться некуда — зато от этого подключения зависит, пойдут
    # тесты через api или клавишами.
    CDP_CONNECT_TIMEOUT_SEC = 2.0

    def _cdp_ensure_connected(self, log_cb=None):
        """Подключает коннектор текущего запуска, если он ещё не подключён.

        Нужно потому, что connect() зовётся лениво: коннектор создаётся при
        запуске Р7 (_prepare_webdriver_launch), а подключается впервые внутри
        _wait_for_bold_button_cdp — и только если триггер готовности вообще
        дошёл до CDP-ветки. Открылся файл быстро, признаки готовности совпали
        раньше — соединения нет, и все операции молча ушли бы на клавиши.
        connect() идемпотентен: для уже подключённого это no-op.

        Args:
            log_cb: Функция логирования; по умолчанию self.add_test_log.

        Returns:
            bool: True, если соединение есть.
        """
        connector = self._webdriver_connector
        if connector is None:
            return False
        if getattr(connector, "connected", False):
            return True
        try:
            return bool(connector.connect(timeout=self.CDP_CONNECT_TIMEOUT_SEC))
        except Exception as e:
            (log_cb or self.add_test_log)(
                f"⚠️ CDP: подключиться не удалось ({type(e).__name__}: {e})")
            return False

    def _cdp_log_api_info(self, log_cb=None):
        """Один раз за запуск Р7 пишет в лог, найден ли внутренний api и какие
        методы у него есть.

        Без этой строки «CDP-операция не сработала» неотличимо от «api не
        нашёлся вовсе»: первое чинится в JS, второе означает, что сборка Р7
        держит api под другим именем и весь перевод тестов на CDP в этом
        запуске просто не работает.

        Args:
            log_cb: Функция логирования; по умолчанию self.add_test_log.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        self._cdp_ensure_connected(log_cb)
        connector = self._cdp_ops_connector()
        if connector is None:
            log_cb("🧩 CDP-операции недоступны в этом запуске — тесты пойдут "
                   "клавишами (pyautogui)")
            return
        try:
            info = connector.api_info(timeout=self.CDP_OP_TIMEOUT_SEC)
        except Exception as e:
            log_cb(f"⚠️ CDP: опрос api редактора не удался ({type(e).__name__}: {e})")
            return
        if not isinstance(info, dict):
            log_cb("⚠️ CDP: ответ на опрос api не получен — операции пойдут клавишами")
            return
        if not info.get("found"):
            log_cb("⚠️ CDP: внутренний api редактора (Asc.spreadsheet_api) не найден "
                   "в DOM — все операции пойдут клавишами")
            return
        state = info.get("state") or {}
        log_cb(f"🧩 CDP: api редактора найден (iframe глубины {info.get('frame')}), "
               f"листов {state.get('sheets')}, выделено {state.get('selection')!r}")
        missing = sorted(n for n, present in (info.get("methods") or {}).items()
                         if not present)
        if missing:
            log_cb(f"   ⚠️ у api нет методов: {', '.join(missing)} — "
                   f"соответствующие тесты пойдут клавишами")

    def _cdp_select_all(self, log_cb=None):
        """Ctrl+A → asc_EditSelectAll."""
        return self._cdp_sequence(
            "Выделение всех ячеек",
            [("asc_EditSelectAll", lambda c, t: c.select_all(timeout=t),
              self.CDP_LONG_OP_TIMEOUT_SEC, 0)],
            self._cdp_check_whole_sheet_selected, log_cb)

    def _cdp_copy(self, log_cb=None):
        """Ctrl+C → asc_Copy.

        Проверять нечем: документ не меняется, а содержимое системного буфера
        из DOM не прочитать. Подтверждением служит сам ответ asc_Copy (ok=false,
        если метод вернул false) — дальше вставка либо сработает, либо нет, и
        это увидит проверка вставки.
        """
        return self._cdp_sequence(
            "Копирование выделения",
            [("asc_Copy", lambda c, t: c.copy(timeout=t),
              self.CDP_LONG_OP_TIMEOUT_SEC, 0)],
            None, log_cb)

    def _cdp_add_sheet(self, log_cb=None):
        """Shift+F11 → asc_addWorksheet."""
        return self._cdp_sequence(
            "Добавление нового листа",
            [("asc_addWorksheet", lambda c, t: c.add_sheet(timeout=t),
              self.CDP_OP_TIMEOUT_SEC, 0)],
            self._cdp_check_sheet_added, log_cb)

    def _cdp_paste_big(self, log_cb=None, key_pace=None):
        """Shift+F11 + Ctrl+V → asc_addWorksheet + asc_Paste.

        Пауза между шагами оставлена такой же, как в клавиатурной версии, и
        так же вычитается из замера — чтобы цифры двух путей оставались
        сравнимыми.
        """
        if key_pace is None:
            key_pace = self.OP_KEY_PACE
        return self._cdp_sequence(
            "Вставка большого массива",
            [("asc_addWorksheet", lambda c, t: c.add_sheet(timeout=t),
              self.CDP_OP_TIMEOUT_SEC, 0),
             ("asc_Paste", lambda c, t: c.paste(timeout=t),
              self.CDP_LONG_OP_TIMEOUT_SEC, key_pace)],
            self._cdp_check_document_changed, log_cb)

    def _cdp_add_column(self, log_cb=None, key_pace=None):
        """Добавление столбца → asc_insertCells(InsertColumns).

        Клавиатурная версия сначала уходит на предыдущий лист (Ctrl+PageUp) и
        сдвигает курсор вправо — здесь то же самое, но детерминированно:
        переход на лист левее активного и выделение ровно B1 (клавиатурный
        `press('right')` сдвигал курсор от того места, где он оказался после
        предыдущего теста, то есть от разного).

        Оба теста — «(горячие клавиши)» и «(меню Вставка)» — на этом пути
        выполняются одним и тем же вызовом api: меню как такового здесь нет.
        Разница между ними остаётся только на pyautogui-пути.
        """
        if key_pace is None:
            key_pace = self.OP_KEY_PACE
        return self._cdp_sequence(
            "Добавление столбца",
            [("asc_showWorksheet (лист левее)",
              lambda c, t: c.show_sheet(-1, relative=True, timeout=t),
              self.CDP_OP_TIMEOUT_SEC, 0),
             ("asc_findCell(B1)", lambda c, t: c.select_range("B1", timeout=t),
              self.CDP_OP_TIMEOUT_SEC, key_pace),
             ("asc_insertCells(InsertColumns)", lambda c, t: c.insert_column(timeout=t),
              self.CDP_LONG_OP_TIMEOUT_SEC, 0)],
            self._cdp_check_document_changed, log_cb)

    def _cdp_copy_paste(self, cell_count, paste_offset, shift=None,
                        log_cb=None, key_pace=None):
        """Копирование N ячеек и вставка со смещением — оба варианта тестов
        «Вставка N ячеек».

        Args:
            cell_count: Сколько ячеек копируем (диапазон A1:<буква>1).
            paste_offset: На сколько столбцов вправо уходим перед вставкой —
                ровно столько раз клавиатурная версия жмёт «вправо» от A1.
            shift: None — обычная вставка (вариант «горячие клавиши»);
                "down"/"right" — вставка ячеек со сдвигом, то есть то, что на
                клавиатурном пути делает пункт контекстного меню «Вставить
                ячейки» и следующая за ним модалка выбора сдвига (вариант
                «ПКМ»). Модалки на этом пути не возникает — подтверждать
                Enter'ом нечего.
            log_cb: Функция логирования.
            key_pace: Пауза после копирования; по умолчанию OP_KEY_PACE — та
                же, что и в клавиатурной версии.

        Returns:
            bool: True — операция ушла в Р7 через CDP.
        """
        if key_pace is None:
            key_pace = self.OP_KEY_PACE
        src = f"A1:{_col_letter(max(1, cell_count))}1"
        dst = f"{_col_letter(paste_offset + 1)}1"
        if shift is None:
            label = f"Вставка {cell_count} ячеек (буфер)"
            last = ("asc_Paste", lambda c, t: c.paste(timeout=t),
                    self.CDP_LONG_OP_TIMEOUT_SEC, 0)
        else:
            label = f"Вставка {cell_count} ячеек (со сдвигом {shift})"
            last = (f"asc_insertCells({shift})",
                    lambda c, t: c.insert_cells(shift, timeout=t),
                    self.CDP_LONG_OP_TIMEOUT_SEC, 0)
        return self._cdp_sequence(
            label,
            [(f"asc_findCell({src})", lambda c, t: c.select_range(src, timeout=t),
              self.CDP_OP_TIMEOUT_SEC, 0),
             ("asc_Copy", lambda c, t: c.copy(timeout=t),
              self.CDP_LONG_OP_TIMEOUT_SEC, 0),
             (f"asc_findCell({dst})", lambda c, t: c.select_range(dst, timeout=t),
              self.CDP_OP_TIMEOUT_SEC, key_pace),
             last],
            self._cdp_check_document_changed, log_cb)

    def _cdp_click_context_item(self, wanted, log_cb=None, charge_pace=True):
        """Пробует нажать пункт раскрытого контекстного меню по его подписи.

        Нужен на pyautogui-пути: меню там обходится стрелками вслепую (`down`
        N раз + Enter), и лишний пункт уводит счётчик — ровно та проблема, из-за
        которой заведён issue #9. Если меню нарисовано в DOM, точное попадание
        по подписи надёжнее счёта стрелок.

        По issue #9 контекстное меню Р7, судя по дампам, рисуется нативным
        оверлеем CEF и в обходимом DOM не появляется — поэтому неудача здесь
        штатная, вызывающий код молча продолжает стрелками.

        Args:
            wanted: Подписи (подстроки, регистр не важен) в порядке приоритета.
            log_cb: Функция логирования; по умолчанию self.add_test_log.
            charge_pace: Отнести длительность round-trip в _paced_total. По
                умолчанию True: метод вызывается при раскрытом меню, когда Р7
                гарантированно простаивает.

        Returns:
            bool: True, если пункт найден и нажат.
        """
        if log_cb is None:
            log_cb = self.add_test_log
        connector = self._cdp_ops_connector()
        if connector is None:
            return False
        t0 = time.time()
        try:
            # Базовый снимок вычитается на стороне JS: без него клик может уйти
            # в статичное overflow-меню тулбара с такими же подписями (issue #9).
            res = connector.click_menu_item(
                wanted, baseline=getattr(self, "_cdp_ui_baseline", None),
                timeout=self.CDP_OP_TIMEOUT_SEC)
        except Exception:
            res = None
        finally:
            if charge_pace:
                self._paced_total += time.time() - t0
        if isinstance(res, dict) and res.get("clicked"):
            log_cb(f"   🧩 CDP: нажат пункт меню {res.get('text')!r} "
                   f"(совпало с {res.get('matched')!r})")
            return True
        return False

    def _close_update_dialog_if_exists(self, log_cb=None, search_timeout=5):
        """Looks for the R7-Office update dialog and closes it if found.

        Scans visible top-level windows for update-related title keywords AND
        owned by a Р7-Офис process (GetWindowThreadProcessId against
        _get_r7_processes()) — a title match alone used to be enough, which let
        this close "Центр обновления Windows", browser tabs containing "update"
        in the title, or a foreign installer window. Enumerates child windows to
        find a dismiss button and clicks it via Win32 messages (no pyautogui, no
        focus dependency). Falls back to WM_CLOSE + VK_ESCAPE if no button is
        matched. Logs nothing if no dialog is present.

        Args:
            log_cb: Callable for log output. Defaults to self.add_test_log.
            search_timeout: Seconds to poll for the dialog window (use a small
                value such as 1 when called from a monitor loop).

        Returns:
            bool: True if a dialog was found and dismissed.
        """
        if log_cb is None:
            log_cb = self.add_test_log

        if not WIN32_OK:
            return False

        import win32gui
        import win32con
        import win32process

        # Только составные фразы, специфичные для диалога обновления Р7-Офис.
        # Раньше список заканчивался голыми "обновление"/"update"/"доступна" —
        # под них подходило почти любое системное окно с таким словом в заголовке.
        UPDATE_TITLES = (
            'обновление программного обеспечения',
            'доступна новая версия',
            'р7-офис обновление',
            'update available',
            'software update',
            'новая версия доступна',
        )
        # Checked in order — first match wins
        DISMISS_PRIORITY = (
            'напомнить позже',
            'пропустить эту версию',
            'не сейчас',
            'remind me later',
            'skip this version',
            'not now',
            'later',
            'skip',
        )

        # ── Search for the dialog (up to search_timeout seconds) ─────────────
        # Сканируем минимум один раз, даже при search_timeout=0. Прежний цикл
        # с проверкой условия на входе при нулевом таймауте не выполнялся ни
        # разу, а при значении по умолчанию (5) сжигал все 5 секунд каждый раз,
        # когда диалога не было, — и эти секунды попадали в замер открытия файла.
        # Владелец окна обязан быть процессом Р7-Офис (не x2t — конвертер не
        # показывает диалогов). Пусто здесь значит, что Р7 сейчас не запущен —
        # в этом случае заголовок, каким бы он ни был, точно не наш диалог.
        r7_pids = {
            p.pid for p in self._get_r7_processes(log_cb=lambda _m: None)
            if "x2t" not in (p.name() or "").lower()
        } if PSUTIL_OK else set()

        def _owned_by_r7(hwnd):
            if not r7_pids:
                return False
            try:
                _, owner_pid = win32process.GetWindowThreadProcessId(hwnd)
            except Exception:
                return False
            return owner_pid in r7_pids

        found = []
        deadline = time.time() + search_timeout

        def _enum(hwnd, _):
            if win32gui.IsWindowVisible(hwnd):
                t = win32gui.GetWindowText(hwnd).lower()
                if any(s in t for s in UPDATE_TITLES) and _owned_by_r7(hwnd):
                    found.append(hwnd)

        while True:
            win32gui.EnumWindows(_enum, None)
            if found or time.time() >= deadline:
                break
            time.sleep(0.5)

        if not found:
            return False

        hwnd = found[0]
        actual_title = win32gui.GetWindowText(hwnd)
        log_cb(f"⚠️ Обнаружено окно обновления: {actual_title}")

        clicked, _ = self._click_priority_button(hwnd, DISMISS_PRIORITY, log_cb=log_cb)

        # ── Fallback when no button matched ──────────────────────────────────
        if not clicked:
            # Try WM_CLOSE first (clean dialog dismissal)
            try:
                win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
            except Exception:
                pass
            time.sleep(0.2)

            # If still visible — send VK_ESCAPE via message (no pyautogui)
            if win32gui.IsWindowVisible(hwnd):
                try:
                    # lParam for key-down: repeat=1, scan=0x01, other bits=0
                    win32gui.PostMessage(hwnd, win32con.WM_KEYDOWN,
                                         win32con.VK_ESCAPE, 0x00010001)
                    time.sleep(0.05)
                    win32gui.PostMessage(hwnd, win32con.WM_KEYUP,
                                         win32con.VK_ESCAPE, 0xC0010001)
                except Exception:
                    pass

        time.sleep(0.3)
        log_cb("✅ Окно обновления закрыто")
        return True

    def _find_r7_path(self):
        """Locates the R7-Office desktop executable, caching the result.

        Returns:
            str: Absolute path to DesktopEditors.exe, or None if not found.
        """
        if self._cached_r7_path:
            return self._cached_r7_path
        # Реальная раскладка установки: ...\R7-Office\Editors\DesktopEditors.exe
        # Вложенной папки DesktopEditors\ не существует — прежний список путей
        # промахивался всеми четырьмя вариантами, и поиск каждый раз уходил в
        # запасной rglob по всему Program Files. Тот отрабатывал (0.4 сек на
        # тестовой машине), но только потому, что каталог R7-Office попадается
        # обходу рано; при другом порядке имён или установке в Program Files
        # (x86) это полный обход дерева.
        possible_paths = [
            r"C:\Program Files\R7-Office\Editors\DesktopEditors.exe",
            r"C:\Program Files (x86)\R7-Office\Editors\DesktopEditors.exe",
            r"C:\Program Files\Р7-Офис\Editors\DesktopEditors.exe",
            r"C:\Program Files (x86)\Р7-Офис\Editors\DesktopEditors.exe",
            # Раскладки других сборок — оставлены как запасные варианты.
            r"C:\Program Files\R7-Office\Editors\DesktopEditors\DesktopEditors.exe",
            r"C:\Program Files (x86)\R7-Office\Editors\DesktopEditors\DesktopEditors.exe",
        ]
        for path in possible_paths:
            if Path(path).exists():
                self._cached_r7_path = path
                return path
        # Запасной поиск: сканируем только каталоги Р7, а не весь Program Files.
        for root in (r"C:\Program Files", r"C:\Program Files (x86)"):
            for brand in ("R7-Office", "Р7-Офис"):
                base = Path(root) / brand
                if not base.exists():
                    continue
                for exe_path in base.rglob("DesktopEditors.exe"):
                    self._cached_r7_path = str(exe_path)
                    return str(exe_path)
        return None


if __name__ == "__main__":
    if not ctypes.windll.shell32.IsUserAnAdmin():
        result = messagebox.askyesno("Права администратора", "Запустить от имени администратора?")
        if result:
            ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
            sys.exit()
    root = tk.Tk()
    app = R7Testovarka(root)
    root.mainloop()

